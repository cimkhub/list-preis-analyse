from pathlib import Path
import zipfile

import numpy as np
from openpyxl import Workbook, load_workbook
import pandas as pd

from match_competitor_products import (
    build_clusters,
    build_final_output_rows,
    build_output_rows,
    dedupe_supplier_offers,
    dedupe_exact_same_supplier_offers,
    dedupe_same_supplier_products,
    final_output_column_widths,
    format_offer_cell_short,
    generate_same_supplier_duplicate_candidates,
    normalize_confidence,
    origin_country_codes,
    validate_cluster_partition,
    validate_offer_accounting,
    write_excel,
    write_final_output_sheet,
)


def product_row(product_id: str, **overrides):
    row = {
        "product_id": product_id,
        "supplier": "metro",
        "supplier_norm": "Metro",
        "category": "fisch",
        "product_name": "Forelle Rot",
        "product": "Forelle Rot",
        "brand": "",
        "description": "ausgenommen, mit Kopf",
        "origin": "Deutschland",
        "unit": "Kiste",
        "quantity": "5",
        "price": "12.99",
        "price_per_kg": "",
        "price_tiers": "",
        "valid_from": "2026-08-03",
        "valid_to": "2026-08-08",
        "source_file": "metro.pdf",
        "source_page": "3",
        "extraction_confidence": "90",
    }
    row.update(overrides)
    return row


def product_attr(product_id: str, **overrides):
    attr = {
        "product_id": product_id,
        "base_product": "trout",
        "variant": "red",
        "processing": None,
        "brand": None,
        "origin": None,
        "quality_class": None,
        "calibre": None,
        "fresh_or_frozen": "unknown",
        "attribute_confidence": 90,
        "notes": "test",
    }
    attr.update(overrides)
    return attr


def exact_same_supplier_pair(a: str, b: str, **overrides):
    pair = {
        "product_id_a": a,
        "product_id_b": b,
        "supplier_a": "Metro",
        "supplier_b": "Metro",
        "candidate_source": "same_supplier_exact_product_name",
        "decision": "exact_match",
        "confidence": 99,
        "should_human_review": False,
        "canonical_name": "",
    }
    pair.update(overrides)
    return pair


def test_normalize_confidence_accepts_fractional_and_percent_scales():
    assert normalize_confidence(1) == 100
    assert normalize_confidence("0.92") == 92
    assert normalize_confidence("87") == 87
    assert normalize_confidence("") == 0


def test_dedupe_exact_same_supplier_offers_keeps_different_units():
    rows = [
        {
            "product_id": "p1",
            "supplier": "metro",
            "supplier_norm": "Metro",
            "category": "fisch",
            "product_name": "Marinaden",
            "product": "Marinaden",
            "brand": "",
            "description": "Verschiedene Sorten",
            "origin": "",
            "unit": "schale",
            "quantity": "2.5",
            "price": "19.99",
            "price_per_kg": "",
            "price_tiers": '[{"min_qty": 4, "price": 19.99}]',
            "valid_from": "2026-06-01",
            "valid_to": "2026-06-06",
            "source_file": "a.pdf",
            "source_page": "8",
        },
        {
            "product_id": "p2",
            "supplier": "metro",
            "supplier_norm": "Metro",
            "category": "fisch",
            "product_name": "Marinaden",
            "product": "Marinaden",
            "brand": "",
            "description": "Verschiedene Sorten",
            "origin": "",
            "unit": "kg",
            "quantity": "2.5",
            "price": "19.99",
            "price_per_kg": "",
            "price_tiers": '[{"price": 19.99, "min_qty": 4}]',
            "valid_from": "2026-06-01",
            "valid_to": "2026-06-06",
            "source_file": "b.pdf",
            "source_page": "9",
        },
    ]

    deduped, removed = dedupe_exact_same_supplier_offers(pd.DataFrame(rows))

    assert removed == 0
    assert len(deduped) == 2


def test_dedupe_exact_same_supplier_offers_removes_only_full_commercial_identity():
    base = {
        "supplier": "selgros",
        "supplier_norm": "Selgros",
        "category": "fisch",
        "product_name": "Forelle",
        "product": "Forelle",
        "brand": "",
        "description": "ausgenommen, 300-400 g",
        "origin": "",
        "unit": "Kiste",
        "quantity": "5",
        "price": "12.99",
        "price_per_kg": "",
        "price_tiers": "",
        "valid_from": "2026-08-03",
        "valid_to": "2026-08-08",
        "source_file": "a.pdf",
        "source_page": "3",
    }
    identical_from_other_source = {
        **base,
        "source_file": "copy.pdf",
        "source_page": "7",
    }
    rows = [
        {**base, "product_id": "p1"},
        {**identical_from_other_source, "product_id": "p2"},
    ]

    deduped, removed = dedupe_exact_same_supplier_offers(pd.DataFrame(rows))

    assert removed == 1
    assert len(deduped) == 1


def test_same_supplier_dedupe_keeps_all_critical_distinct_offers():
    cases = [
        ("Forelle", "fisch", "ausgenommen, mit Kopf", "11.99", "12.99", "5", "Kiste"),
        ("Cherry Strauchtomaten", "obst_gemuese", "Klasse I", "5.99", "6.99", "2", "Kiste"),
        ("Weißkohl", "obst_gemuese", "Klasse I", "4.99", "4.99", "10", "Sack"),
    ]

    for product, category, description, old_price, new_price, quantity, unit in cases:
        base = {
            "supplier": "selgros",
            "supplier_norm": "Selgros",
            "category": category,
            "product_name": product,
            "product": product,
            "brand": "",
            "description": description,
            "origin": "",
            "unit": unit,
            "quantity": quantity,
            "price_per_kg": "",
            "price_tiers": "",
        }
        rows = [
            {
                **base,
                "product_id": "old",
                "price": old_price,
                "valid_from": "2026-08-03",
                "valid_to": "2026-08-08",
            },
            {
                **base,
                "product_id": "new",
                "price": new_price,
                "valid_from": "2026-08-10",
                "valid_to": "2026-08-15",
            },
        ]

        deduped, removed = dedupe_same_supplier_products(
            pd.DataFrame(rows),
            attributes={},
            embeddings={},
            caches=None,
            pair_client=None,
            force=False,
            skip_llm=True,
            pair_workers=1,
            top_k=1,
        )

        assert removed == 0, product
        assert set(deduped["product_id"]) == {"old", "new"}, product


def test_final_output_moves_brand_into_description_and_removes_brand_column():
    rows = build_final_output_rows([
        {
            "category": "fleisch",
            "product": "Rinderfilet",
            "brand": "LIST Premium",
            "description": "Argentinisches Rinderfilet",
            "origin": "Argentinien",
        }
    ])

    assert "Marke" not in rows[0]
    assert rows[0]["Beschreibung"] == "Argentinisches Rinderfilet\nBrand: LIST Premium,"
    assert rows[0]["Herkunft"] == "AR"


def test_final_output_skips_brand_line_when_brand_is_empty():
    rows = build_final_output_rows([
        {
            "category": "fleisch",
            "product": "Rinderfilet",
            "brand": " ",
            "description": "Argentinisches Rinderfilet",
            "origin": "Argentinien",
        }
    ])

    assert rows[0]["Beschreibung"] == "Argentinisches Rinderfilet"


def test_origin_country_codes_maps_german_names_aliases_and_existing_codes():
    assert origin_country_codes("Deutschland / Niederlande") == "DE, NL"
    assert origin_country_codes("Holland") == "NL"
    assert origin_country_codes("de") == "DE"
    assert origin_country_codes("Spanien, Italien oder Frankreich") == "ES, IT, FR"
    assert origin_country_codes("Chile + Neuseeland") == "CL, NZ"


def test_short_offer_cell_uses_line_breaks_instead_of_wide_separator():
    cell = format_offer_cell_short({
        "price": "19.99",
        "quantity": "2.5",
        "unit": "kg",
        "price_tiers": '[{"min_qty": 4, "price": 18.5}]',
        "valid_from": "2026-06-01",
        "valid_to": "2026-06-06",
    })

    assert cell == "19,99 €\n2,5 kg\nab 4: 18,5 €\n01.06. bis 06.06."
    assert " | " not in cell


def test_compact_final_output_widths_are_narrower_for_editing():
    widths = final_output_column_widths(compact_rows=True)

    assert sum(widths.values()) <= 170
    assert widths["Metro"] == 17
    assert widths["Beschreibung"] == 24
    assert widths["Herkunft"] == 7


def test_final_output_formats_ek_and_vk_as_euro_values():
    wb = Workbook()
    ws = wb.active
    write_final_output_sheet(
        ws,
        [{
            "Kategorie": "Fleisch",
            "Produkt": "Rinderfilet",
            "Beschreibung": "Argentinisches Rinderfilet",
            "Herkunft": "AR",
            "Metro": "",
            "Selgros": "",
            "Handelshof": "",
            "Edeka": "",
            "EK": 12.3456,
            "VK": 19.99,
            "Notizen": "",
        }],
        Path("Artikelvergleich KW01.xlsx"),
        None,
    )
    headers = {str(cell.value): cell.column for cell in ws[7]}

    assert ws.cell(8, headers["EK"]).number_format == '#,##0.000 "€"'
    assert ws.cell(8, headers["VK"]).number_format == '#,##0.00 "€"'


def test_compact_final_output_row_height_grows_with_wrapped_content():
    wb = Workbook()
    ws = wb.active
    write_final_output_sheet(
        ws,
        [{
            "Kategorie": "Fleisch",
            "Produkt": "Rinderfilet mit sehr langem Namen",
            "Beschreibung": "Sehr lange Beschreibung mit vielen Details zur Qualität, zum Zuschnitt und zur Verarbeitung\nBrand: LIST Premium,",
            "Herkunft": "AR",
            "Metro": "19,99 €\n2,5 kg\nab 4: 18,5 €\n01.06. bis 06.06.",
            "Selgros": "20,99 €\n2,5 kg\n01.06. bis 06.06.",
            "Handelshof": "",
            "Edeka": "",
            "EK": 12.3456,
            "VK": 19.99,
            "Notizen": "Lange Bearbeitungsnotiz fuer die Disposition",
        }],
        Path("Artikelvergleich KW01.xlsx"),
        None,
        compact_rows=True,
    )

    assert ws.row_dimensions[8].height > 80
    assert ws.row_dimensions[8].height <= 360


def test_same_supplier_candidates_include_same_page_and_ignore_description_differences():
    rows = [
        product_row("p1", description="ausgenommen"),
        product_row("p2", description="ausgenommen, mit Kopf, Kaliber 300-400 g"),
    ]
    attrs = {row["product_id"]: product_attr(row["product_id"]) for row in rows}
    embeddings = {"p1": np.array([1.0, 0.0]), "p2": np.array([0.9, 0.1])}

    candidates = generate_same_supplier_duplicate_candidates(pd.DataFrame(rows), attrs, embeddings, top_k=5)

    assert len(candidates) == 1
    assert candidates[0]["candidate_source"] == "same_supplier_exact_product_name"


def test_four_extraction_representations_become_one_row_and_one_offer():
    rows = [
        product_row("p1", description="ausgenommen", quantity="", unit="Kiste"),
        product_row("p2", description="ausgenommen, mit Kopf", quantity="5", unit="Kiste"),
        product_row("p3", description="ausgenommen, mit Kopf, Kaliber 300-400 g", quantity="5", unit="Kiste"),
        product_row("p4", description="ausgenommen", quantity="", unit="Kiste"),
    ]
    df = pd.DataFrame(rows)
    attrs = {row["product_id"]: product_attr(row["product_id"]) for row in rows}
    pairs = [exact_same_supplier_pair("p1", "p2"), exact_same_supplier_pair("p2", "p3"), exact_same_supplier_pair("p3", "p4")]
    clusters = build_clusters(df, attrs, pairs)
    audit = []

    output, _ = build_output_rows(df, attrs, pairs, clusters, grouping_audit=audit)
    validate_offer_accounting(df, audit)

    assert len(output) == 1
    assert output[0]["Metro_short"].count("12,99 €") == 1
    assert output[0]["description"] == "ausgenommen, mit Kopf, Kaliber 300-400 g"
    assert len(audit) == 1
    assert audit[0]["rule"] == "same commercial offer; source extraction representations coalesced"
    assert set(audit[0]["merged_product_ids"]) == {"p1", "p2", "p3", "p4"}


def test_same_variant_keeps_different_prices_and_periods_as_two_offer_blocks():
    rows = [
        product_row("old", price="11.99", valid_from="2026-08-03", valid_to="2026-08-08"),
        product_row("new", price="12.99", valid_from="2026-08-10", valid_to="2026-08-15"),
    ]
    df = pd.DataFrame(rows)
    attrs = {row["product_id"]: product_attr(row["product_id"], calibre=calibre) for row, calibre in zip(rows, ["200-400 g", "300-400 g"])}
    pairs = [exact_same_supplier_pair("old", "new")]
    clusters = build_clusters(df, attrs, pairs)

    output, _ = build_output_rows(df, attrs, pairs, clusters)

    assert len(output) == 1
    assert output[0]["Metro_short"].count("€") == 2
    assert "03.08. bis 08.08.\n\n12,99 €" in output[0]["Metro_short"]


def test_same_source_quantity_noise_is_coalesced_but_distinct_offer_is_kept():
    products = [
        product_row("p1", quantity="", unit="kg"),
        product_row("p2", quantity="2.5", unit="kg", description="ausgenommen, mit Kopf, 2,5 kg"),
        product_row("p3", price="13.99", quantity="2.5", unit="kg"),
    ]
    audit = []

    offers = dedupe_supplier_offers(products, audit, "c1", "Metro")

    assert len(offers) == 2
    assert {offer["price"] for offer in offers} == {"12.99", "13.99"}
    retained = next(offer for offer in offers if offer["price"] == "12.99")
    assert retained["quantity"] == "2.5"
    assert retained["description"] == "ausgenommen, mit Kopf, 2,5 kg"
    assert len(audit) == 2
    assert {entry["rule"] for entry in audit} == {
        "same commercial offer; source extraction representations coalesced",
        "unique commercial offer",
    }


def test_two_explicitly_different_pack_sizes_remain_distinct_offers():
    products = [
        product_row("p1", quantity="1", unit="Kiste"),
        product_row("p2", quantity="3", unit="Kiste"),
    ]

    offers = dedupe_supplier_offers(products)

    assert len(offers) == 2
    assert {offer["quantity"] for offer in offers} == {"1", "3"}


def test_missing_origin_cannot_bridge_two_different_known_origins():
    rows = [
        product_row("de", origin="Deutschland"),
        product_row("missing", origin=""),
        product_row("es", origin="Spanien"),
    ]
    df = pd.DataFrame(rows)
    attrs = {row["product_id"]: product_attr(row["product_id"], origin=row["origin"] or None) for row in rows}
    pairs = [
        exact_same_supplier_pair("de", "missing"),
        exact_same_supplier_pair("missing", "es"),
        exact_same_supplier_pair("de", "es"),
    ]

    clusters = build_clusters(df, attrs, pairs)
    validate_cluster_partition(df, clusters)

    assert len(clusters) == 2
    assert any(pair.get("cluster_merge_blocked") for pair in pairs)


def test_different_known_brands_and_temperature_states_stay_separate():
    brand_rows = [
        product_row("aro", category="mopro", product="Speisequark", product_name="Aro Speisequark", brand="Aro"),
        product_row("milram", category="mopro", product="Speisequark", product_name="Milram Speisequark", brand="Milram"),
    ]
    brand_attrs = {row["product_id"]: product_attr(row["product_id"], base_product="quark", variant=None) for row in brand_rows}
    brand_pair = [exact_same_supplier_pair("aro", "milram", candidate_source="same_supplier_vector_similarity")]

    assert len(build_clusters(pd.DataFrame(brand_rows), brand_attrs, brand_pair)) == 2

    temperature_rows = [
        product_row("fresh", product="Rinder-Roastbeef", product_name="Rinder-Roastbeef", description="frisch"),
        product_row("frozen", category="tk", product="Rinder-Roastbeef", product_name="Rinder-Roastbeef", description="gefroren"),
    ]
    temperature_attrs = {row["product_id"]: product_attr(row["product_id"], base_product="beef roastbeef", variant=None) for row in temperature_rows}
    temperature_pair = [exact_same_supplier_pair("fresh", "frozen")]

    assert len(build_clusters(pd.DataFrame(temperature_rows), temperature_attrs, temperature_pair)) == 2


def test_duroc_and_origin_combinations_create_separate_variants():
    premium_rows = [
        product_row("generic", category="fleisch", product="Schweinenacken", product_name="Schweinenacken", description="frisch"),
        product_row("duroc", category="fleisch", product="Schweinenacken", product_name="Schweinenacken", description="Duroc, frisch"),
    ]
    premium_attrs = {row["product_id"]: product_attr(row["product_id"], base_product="pork neck", variant=None) for row in premium_rows}
    premium_pairs = [exact_same_supplier_pair("generic", "duroc")]

    assert len(build_clusters(pd.DataFrame(premium_rows), premium_attrs, premium_pairs)) == 2

    origin_rows = [
        product_row("combo", origin="UA, RO"),
        product_row("ua", origin="UA"),
    ]
    origin_attrs = {row["product_id"]: product_attr(row["product_id"], origin=row["origin"]) for row in origin_rows}
    origin_pairs = [exact_same_supplier_pair("combo", "ua")]

    assert len(build_clusters(pd.DataFrame(origin_rows), origin_attrs, origin_pairs)) == 2


def test_grouped_workbook_is_valid_and_contains_offer_audit(tmp_path):
    rows = [product_row("p1", unit="kg"), product_row("p2", quantity="", unit="kg")]
    df = pd.DataFrame(rows)
    attrs = {row["product_id"]: product_attr(row["product_id"]) for row in rows}
    pairs = [exact_same_supplier_pair("p1", "p2")]
    clusters = build_clusters(df, attrs, pairs)
    audit = []
    matched, review = build_output_rows(df, attrs, pairs, clusters, grouping_audit=audit)
    output_path = tmp_path / "Artikelvergleich KW32.xlsx"

    write_excel(output_path, matched, review, pairs, [], None, offer_grouping_audit=audit)

    with zipfile.ZipFile(output_path) as archive:
        assert archive.testzip() is None
    workbook = load_workbook(output_path, read_only=False, data_only=False)
    assert "offer_grouping_audit" in workbook.sheetnames
    assert workbook["offer_grouping_audit"].max_row == 2
