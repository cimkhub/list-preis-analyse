import json

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

import final_row_quality_check as final_review


def build_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Output Short"
    ws["D3"] = "Erstellt am 09.08.2026 | Artikelgruppen: 3"
    headers = ["Kategorie", "Produkt", "Beschreibung", "Metro"]
    for col, header in enumerate(headers, 1):
        ws.cell(7, col, header)
    rows = [
        ["Sonstiges", "Lachsfilet", "roh, ohne Haut", "12,99 €"],
        ["Sonstiges", "Riesling Sekt", "trocken", "4,99 €"],
        ["Obst & Gemüse", "Tomaten", "Klasse I", "2,99 €"],
    ]
    for row_idx, values in enumerate(rows, 8):
        for col_idx, value in enumerate(values, 1):
            ws.cell(row_idx, col_idx, value)
    ws.add_table(Table(displayName="Final_Output_Short_tbl", ref="A7:D10"))
    ws.auto_filter.ref = "A7:D10"
    wb.save(path)


def test_final_review_corrects_category_deletes_only_hard_error_and_writes_audit(tmp_path):
    workbook_path = tmp_path / "Artikelvergleich KW32.xlsx"
    audit_path = tmp_path / "audit.jsonl"
    build_workbook(workbook_path)

    def fake_caller(**kwargs):
        product = kwargs["row"]["Produkt"]
        if product == "Lachsfilet":
            return {
                "action": "correct_category",
                "category": "Fisch",
                "hard_error": True,
                "reason": "Fisch falsch kategorisiert",
                "confidence": 0.97,
            }
        if product == "Riesling Sekt":
            return {
                "action": "delete",
                "category": None,
                "hard_error": True,
                "reason": "Getränk außerhalb Zielsortiment",
                "confidence": 0.99,
            }
        return {
            "action": "keep",
            "category": None,
            "hard_error": False,
            "reason": "Logisch",
            "confidence": 0.96,
        }

    stats = final_review.review_workbook_rows(
        workbook_path,
        audit_path=audit_path,
        api_key="test-key",
        workers=2,
        max_retries=1,
        caller=fake_caller,
    )

    assert stats == {
        "checked": 3,
        "corrected": 1,
        "deleted": 1,
        "kept": 1,
        "sheet": "Final Output Short",
        "model": final_review.DEFAULT_FLASH_MODEL,
        "audit_path": str(audit_path),
    }
    wb = load_workbook(workbook_path)
    ws = wb["Final Output Short"]
    assert ws["A8"].value == "Fisch"
    assert ws["B8"].value == "Lachsfilet"
    assert ws["B9"].value == "Tomaten"
    assert ws["B10"].value is None
    assert ws.tables["Final_Output_Short_tbl"].ref == "A7:D9"
    assert "Artikelgruppen: 2" in ws["D3"].value

    audit = [json.loads(line) for line in audit_path.read_text(encoding="utf-8").splitlines()]
    assert [item["input"]["Produkt"] for item in audit] == ["Lachsfilet", "Riesling Sekt", "Tomaten"]
    assert [item["applied_action"] for item in audit] == ["correct_category", "delete", "keep"]


def test_low_confidence_delete_is_not_applied():
    decision = final_review.normalize_decision(
        {
            "action": "delete",
            "category": None,
            "hard_error": True,
            "reason": "Unsicher",
            "confidence": 0.70,
        },
        "Sonstiges",
    )
    assert decision["applied_action"] == "keep"


def test_high_confidence_category_correction_is_applied_even_if_hard_error_flag_is_inconsistent():
    decision = final_review.normalize_decision(
        {
            "action": "correct_category",
            "category": "Fisch",
            "hard_error": False,
            "reason": "Lachs ist Fisch",
            "confidence": 0.95,
        },
        "Sonstiges",
    )
    assert decision["applied_action"] == "correct_category"
    assert decision["applied_category"] == "Fisch"


def test_final_flash_call_uses_json_and_disables_thinking(monkeypatch):
    captured = {}

    class FakeResponse:
        def raise_for_status(self):
            return None

        def json(self):
            return {
                "choices": [
                    {
                        "message": {
                            "content": json.dumps(
                                {
                                    "action": "keep",
                                    "category": None,
                                    "hard_error": False,
                                    "reason": "Logisch",
                                    "confidence": 0.99,
                                }
                            )
                        }
                    }
                ]
            }

    class FakeSession:
        def post(self, url, *, headers, json, timeout):
            captured.update({"url": url, "headers": headers, "json": json, "timeout": timeout})
            return FakeResponse()

    monkeypatch.setattr(final_review, "get_session", lambda: FakeSession())
    result = final_review.call_deepseek_flash(
        api_key="secret",
        model=final_review.DEFAULT_FLASH_MODEL,
        base_url="https://example.invalid",
        timeout_seconds=12,
        row={"Kategorie": "Fisch", "Produkt": "Lachsfilet"},
    )

    assert result["action"] == "keep"
    assert captured["json"]["model"] == final_review.DEFAULT_FLASH_MODEL
    assert captured["json"]["thinking"] == {"type": "disabled"}
    assert captured["json"]["response_format"] == {"type": "json_object"}
    assert '"Produkt":"Lachsfilet"' in captured["json"]["messages"][1]["content"]


def test_review_failure_aborts_without_modifying_workbook(tmp_path):
    workbook_path = tmp_path / "Artikelvergleich KW32.xlsx"
    build_workbook(workbook_path)

    def failing_caller(**kwargs):
        raise RuntimeError("API unavailable")

    with pytest.raises(RuntimeError, match="Final DeepSeek Flash check failed"):
        final_review.review_workbook_rows(
            workbook_path,
            api_key="test-key",
            workers=1,
            max_retries=1,
            caller=failing_caller,
        )

    wb = load_workbook(workbook_path)
    assert wb["Final Output Short"]["B9"].value == "Riesling Sekt"
