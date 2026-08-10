#!/usr/bin/env python3
"""Final customer-row sanity check with DeepSeek Flash.

The checker runs after market-forecast enrichment and before the customer
workbook is finalized/uploaded. It may only keep a row, correct its category,
or delete a clearly out-of-scope row. Source-backed product, description,
origin, offer, price, validity, and forecast cells are immutable here.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from threading import local
from typing import Any, Callable

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


LOGGER = logging.getLogger("birkenhof.final_row_quality")
DEFAULT_DEEPSEEK_BASE_URL = "https://api.deepseek.com"
DEFAULT_FLASH_MODEL = os.environ.get("DEEPSEEK_V4_FLASH_MODEL", "deepseek-v4-flash")
DEFAULT_WORKERS = 10
DEFAULT_TIMEOUT_SECONDS = 120
DEFAULT_MAX_RETRIES = 3
PROMPT_VERSION = "final-row-quality-v2"
DELETE_CONFIDENCE_THRESHOLD = 0.90
CATEGORY_CONFIDENCE_THRESHOLD = 0.80
ALLOWED_CATEGORIES = (
    "Fisch",
    "Fleisch",
    "Obst & Gemüse",
    "Mopro",
    "TK",
    "Wurst",
    "Sonstiges",
)
_THREAD_LOCAL = local()


def require_flash_model(model: str) -> None:
    if model != DEFAULT_FLASH_MODEL:
        raise RuntimeError(
            "Final row quality review is locked to the configured DeepSeek Flash model "
            f"{DEFAULT_FLASH_MODEL!r}; got {model!r}."
        )


def get_session() -> requests.Session:
    session = getattr(_THREAD_LOCAL, "session", None)
    if session is None:
        session = requests.Session()
        _THREAD_LOCAL.session = session
    return session


def build_review_prompt(row: dict[str, Any]) -> str:
    compact_json = json.dumps(row, ensure_ascii=False, separators=(",", ":"))
    return "\n".join(
        [
            "Prüfe genau eine finale Zeile eines Wettbewerbsvergleichs auf harte fachliche Fehler.",
            "Die Eingabe ist JSON. Nutze ausschließlich die darin enthaltenen Informationen.",
            "Bei Unsicherheit immer KEEP. Keine neuen Produktdaten erfinden.",
            "",
            "Zulässige Aktionen:",
            "- keep: Zeile ist logisch oder nur unvollständig/unklar.",
            "- correct_category: Kategorie ist eindeutig falsch; nur category setzen.",
            "- delete: Produkt ist eindeutig nicht im Zielsortiment. Nur bei einem harten, klaren Fehler.",
            "",
            "Zielsortiment:",
            "- rohes/einfaches Fleisch und Geflügel; auch geschnitten, zerkleinert, geformt, gewürzt, mariniert, als BBQ-Zuschnitt oder als Spieß",
            "- roher/einfacher Fisch und Seafood ohne Sauce/Marinade/Feinkostzubereitung",
            "- frisches Obst, Gemüse, Salat, Kräuter, Pilze und Kartoffeln",
            "- zusätzlich Öl, Sahne, Quark, Milch, TK-Gemüse und Pommes nur mit belegter Marke aus ARO, Chef, Metro, Milram, Schleiz, Quality, economy, Edeka, Foodservice, Henkelmann, Meemken oder Aviko",
            "- Käse und Wurst zusätzlich nur mit einer dieser Marken und Verpackungsgröße ab 500 g",
            "",
            "Für Fleisch und Geflügel hat ein explizites frisch/roh/gekühlt Vorrang vor mariniert, Kräutermarinade, gewürzt oder BBQ. Solche rohen Fleischprodukte niemals allein wegen Marinade oder BBQ löschen. Beispiele: Short Ribs BBQ, frisch => keep; Lammhüftsteaks, in Kräutermarinade, frisch => keep.",
            "Eindeutig außerhalb sind insbesondere Non-Food, Getränke, Saucen, Dips, Pizzasauce, Guacamole, Gewürzgurken, Konserven, Feinkostsalate, fertige Gerichte und Fisch in Sauce/Sahne/Marinade. Gekochtes oder verzehrfertiges Fleisch bleibt ebenfalls außerhalb.",
            "Die Händlernamen Metro/Edeka in Spaltenüberschriften sind kein Markenbeleg. Ein Markenbeleg muss im Produkt oder in der Beschreibung stehen.",
            "",
            "Erlaubte Kategorien: " + ", ".join(ALLOWED_CATEGORIES) + ".",
            "Kategorie-Leitlinie: frischer/gekühlter Fisch=Fisch; frisches/gekühltes Fleisch=Fleisch; frisches Obst/Gemüse=Obst & Gemüse; gefrorene Produkte=TK; Wurst=Wurst; Milchprodukte=Mopro; sonstige zulässige Zusatzprodukte=Sonstiges.",
            "",
            "Antworte ausschließlich als JSON-Objekt:",
            '{"action":"keep|correct_category|delete","category":null,"hard_error":false,"reason":"kurze Begründung","confidence":0.0}',
            "Bei correct_category muss category eine erlaubte Kategorie sein. Bei keep/delete muss category null sein. confidence liegt zwischen 0 und 1.",
            "Setze hard_error bei correct_category und delete immer auf true; nur keep darf hard_error=false haben.",
            "",
            "Eingabe:",
            compact_json,
        ]
    )


def call_deepseek_flash(
    *,
    api_key: str,
    model: str,
    base_url: str,
    timeout_seconds: int,
    row: dict[str, Any],
) -> dict[str, Any]:
    require_flash_model(model)
    response = get_session().post(
        base_url.rstrip("/") + "/chat/completions",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json={
            "model": model,
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "Du bist die letzte strenge Qualitätskontrolle für Produktzeilen. "
                        "Gib ausschließlich ein gültiges JSON-Objekt zurück."
                    ),
                },
                {"role": "user", "content": build_review_prompt(row)},
            ],
            "temperature": 0,
            "max_tokens": 300,
            "stream": False,
            "response_format": {"type": "json_object"},
            # The final check is intentionally cheap and fast; thinking is
            # reserved for the Pro relevance decision requested by the user.
            "thinking": {"type": "disabled"},
        },
        timeout=timeout_seconds,
    )
    response.raise_for_status()
    payload = response.json()
    choices = payload.get("choices") or []
    if not choices:
        raise RuntimeError("DeepSeek Flash response has no choices")
    message = choices[0].get("message") or {}
    content = message.get("content")
    if isinstance(content, list):
        content = "".join(
            str(part.get("text") or "") if isinstance(part, dict) else str(part)
            for part in content
        )
    if not str(content or "").strip():
        raise RuntimeError("DeepSeek Flash response has no content")
    return parse_json_object(str(content))


def parse_json_object(text: str) -> dict[str, Any]:
    cleaned = str(text or "").strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?", "", cleaned).strip()
        cleaned = re.sub(r"```$", "", cleaned).strip()
    try:
        value = json.loads(cleaned)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
        if not match:
            raise
        value = json.loads(match.group(0))
    if not isinstance(value, dict):
        raise RuntimeError("DeepSeek Flash result must be a JSON object")
    return value


def normalize_decision(data: dict[str, Any], current_category: str) -> dict[str, Any]:
    action = str(data.get("action") or "").strip().casefold()
    if action not in {"keep", "correct_category", "delete"}:
        raise RuntimeError(f"Invalid final-review action: {action!r}")

    hard_error = _as_bool(data.get("hard_error"))
    confidence = _as_confidence(data.get("confidence"))
    reason = re.sub(r"\s+", " ", str(data.get("reason") or "").strip())[:300]
    requested_category = canonical_category(data.get("category"))
    applied_action = action
    applied_category: str | None = None

    if action == "correct_category":
        if requested_category is None:
            raise RuntimeError(f"Invalid corrected category: {data.get('category')!r}")
        # A category correction is low-impact and restricted to a closed enum.
        # Some Flash responses correctly request the change but inconsistently
        # leave hard_error=false. The explicit action plus confidence is enough
        # to apply it; destructive deletion remains strictly fail-closed below.
        if confidence < CATEGORY_CONFIDENCE_THRESHOLD:
            applied_action = "keep"
        elif requested_category.casefold() == str(current_category or "").strip().casefold():
            applied_action = "keep"
        else:
            applied_category = requested_category
    elif action == "delete":
        if not hard_error or confidence < DELETE_CONFIDENCE_THRESHOLD:
            applied_action = "keep"
    else:
        requested_category = None

    return {
        "action": action,
        "category": requested_category,
        "hard_error": hard_error,
        "reason": reason or "Keine Begründung",
        "confidence": confidence,
        "applied_action": applied_action,
        "applied_category": applied_category,
    }


def canonical_category(value: Any) -> str | None:
    text = re.sub(r"\s+", " ", str(value or "").strip()).casefold()
    aliases = {
        "fisch": "Fisch",
        "fleisch": "Fleisch",
        "obst & gemüse": "Obst & Gemüse",
        "obst & gemuese": "Obst & Gemüse",
        "obst und gemüse": "Obst & Gemüse",
        "obst und gemuese": "Obst & Gemüse",
        "obst_gemuese": "Obst & Gemüse",
        "mopro": "Mopro",
        "tk": "TK",
        "wurst": "Wurst",
        "sonstiges": "Sonstiges",
    }
    return aliases.get(text)


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().casefold() in {"true", "1", "yes", "ja"}


def _as_confidence(value: Any) -> float:
    try:
        confidence = float(value)
    except (TypeError, ValueError) as exc:
        raise RuntimeError(f"Invalid confidence: {value!r}") from exc
    if confidence > 1 and confidence <= 100:
        confidence /= 100
    if not 0 <= confidence <= 1:
        raise RuntimeError(f"Confidence outside 0..1: {confidence!r}")
    return confidence


def review_one_row(
    *,
    row_idx: int,
    row: dict[str, Any],
    api_key: str,
    model: str,
    base_url: str,
    timeout_seconds: int,
    max_retries: int,
    caller: Callable[..., dict[str, Any]] = call_deepseek_flash,
) -> dict[str, Any]:
    last_error: Exception | None = None
    for attempt in range(1, max_retries + 1):
        try:
            raw = caller(
                api_key=api_key,
                model=model,
                base_url=base_url,
                timeout_seconds=timeout_seconds,
                row=row,
            )
            decision = normalize_decision(raw, str(row.get("Kategorie") or ""))
            return {
                "prompt_version": PROMPT_VERSION,
                "model": model,
                "row_number": row_idx,
                "input": row,
                **decision,
            }
        except Exception as exc:
            last_error = exc
            if attempt < max_retries:
                time.sleep(min(2 ** (attempt - 1), 8))
    raise RuntimeError(
        f"Final DeepSeek Flash check failed for Excel row {row_idx} "
        f"({row.get('Produkt') or '<unknown product>'}): {last_error}"
    ) from last_error


def find_customer_sheet(workbook) -> str:
    if "Final Output Short" in workbook.sheetnames:
        return "Final Output Short"
    for name in workbook.sheetnames:
        if re.fullmatch(r"KW\d{1,2}", name):
            return name
    if "Final Output" in workbook.sheetnames:
        return "Final Output"
    raise RuntimeError("No customer-facing output sheet found")


def find_header(ws) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 15) + 1):
        headers = {
            str(ws.cell(row_idx, col_idx).value or "").strip(): col_idx
            for col_idx in range(1, ws.max_column + 1)
            if str(ws.cell(row_idx, col_idx).value or "").strip()
        }
        required = {"Kategorie", "Produkt", "Beschreibung"}
        supplier_headers = {"Metro", "Selgros", "Handelshof", "Edeka"}
        if required.issubset(headers) and supplier_headers.intersection(headers):
            return row_idx, headers
    raise RuntimeError("Could not find Kategorie/Produkt header row")


def collect_rows(ws, header_row: int, headers: dict[str, int]) -> list[tuple[int, dict[str, Any]]]:
    ordered_headers = sorted(headers.items(), key=lambda item: item[1])
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = {
            header: ws.cell(row_idx, col_idx).value
            for header, col_idx in ordered_headers
        }
        if not str(row.get("Produkt") or "").strip():
            continue
        rows.append((row_idx, row))
    return rows


def apply_decisions(ws, header_row: int, headers: dict[str, int], decisions: list[dict[str, Any]]) -> dict[str, int]:
    category_col = headers["Kategorie"]
    corrected = 0
    deleted_rows: list[int] = []
    for decision in decisions:
        row_idx = int(decision["row_number"])
        if decision["applied_action"] == "correct_category":
            ws.cell(row_idx, category_col).value = decision["applied_category"]
            corrected += 1
        elif decision["applied_action"] == "delete":
            deleted_rows.append(row_idx)

    for row_idx in sorted(deleted_rows, reverse=True):
        ws.delete_rows(row_idx, 1)

    product_col = headers["Produkt"]
    remaining_rows = [
        row_idx
        for row_idx in range(header_row + 1, ws.max_row + 1)
        if str(ws.cell(row_idx, product_col).value or "").strip()
    ]
    remaining = len(remaining_rows)
    if remaining != len(decisions) - len(deleted_rows):
        raise RuntimeError(
            "Final row accounting mismatch after applying decisions: "
            f"expected {len(decisions) - len(deleted_rows)}, found {remaining}."
        )
    _resize_sheet_tables(ws, header_row, max(remaining_rows) if remaining_rows else None)
    _update_article_count(ws, remaining)
    return {"checked": len(decisions), "corrected": corrected, "deleted": len(deleted_rows), "kept": remaining}


def _resize_sheet_tables(ws, header_row: int, last_data_row: int | None) -> None:
    if last_data_row is None:
        for name in list(ws.tables):
            del ws.tables[name]
        ws.auto_filter.ref = None
        return
    for table in ws.tables.values():
        min_col, _min_row, max_col, _max_row = range_boundaries(table.ref)
        table_ref = (
            f"{get_column_letter(min_col)}{header_row}:"
            f"{get_column_letter(max_col)}{last_data_row}"
        )
        table.ref = table_ref
        if table.autoFilter is not None:
            table.autoFilter.ref = table_ref
    if ws.tables:
        # The table owns its AutoFilter. A second worksheet AutoFilter is both
        # redundant and a frequent source of Excel repair warnings.
        ws.auto_filter.ref = None
    else:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(ws.max_column)}{last_data_row}"
        )


def _update_article_count(ws, row_count: int) -> None:
    value = str(ws["D3"].value or "")
    if "Artikelgruppen:" in value:
        ws["D3"] = re.sub(r"Artikelgruppen:\s*\d+", f"Artikelgruppen: {row_count}", value)


def write_audit(path: Path, decisions: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with tmp_path.open("w", encoding="utf-8") as handle:
        for decision in sorted(decisions, key=lambda item: int(item["row_number"])):
            handle.write(json.dumps(decision, ensure_ascii=False, sort_keys=True) + "\n")
    os.replace(tmp_path, path)


def verify_customer_workbook(
    workbook_path: Path,
    *,
    expected_product_rows: int | None = None,
) -> dict[str, Any]:
    """Fail closed when row accounting or Excel table metadata is inconsistent."""
    workbook = load_workbook(workbook_path)
    sheet_name = find_customer_sheet(workbook)
    ws = workbook[sheet_name]
    header_row, headers = find_header(ws)
    product_rows = collect_rows(ws, header_row, headers)
    product_count = len(product_rows)
    if expected_product_rows is not None and product_count != expected_product_rows:
        raise RuntimeError(
            "Saved customer workbook row mismatch: "
            f"expected {expected_product_rows}, found {product_count}."
        )

    last_data_row = max((row_idx for row_idx, _row in product_rows), default=None)
    for table in ws.tables.values():
        _min_col, min_row, _max_col, max_row = range_boundaries(table.ref)
        if min_row != header_row or max_row != last_data_row:
            raise RuntimeError(
                f"Excel table {table.displayName!r} has invalid ref {table.ref!r}; "
                f"expected header row {header_row} and last data row {last_data_row}."
            )
        if table.autoFilter is None or table.autoFilter.ref != table.ref:
            raise RuntimeError(
                f"Excel table {table.displayName!r} AutoFilter does not match "
                f"its table range {table.ref!r}."
            )
    if ws.tables and ws.auto_filter.ref:
        raise RuntimeError(
            "Customer sheet contains a redundant worksheet AutoFilter in addition to its table filter."
        )
    return {
        "sheet": sheet_name,
        "product_rows": product_count,
        "header_row": header_row,
        "tables": len(ws.tables),
    }


def review_workbook_rows(
    workbook_path: Path,
    *,
    audit_path: Path | None = None,
    api_key: str | None = None,
    model: str = DEFAULT_FLASH_MODEL,
    base_url: str | None = None,
    workers: int = DEFAULT_WORKERS,
    timeout_seconds: int = DEFAULT_TIMEOUT_SECONDS,
    max_retries: int = DEFAULT_MAX_RETRIES,
    caller: Callable[..., dict[str, Any]] = call_deepseek_flash,
) -> dict[str, Any]:
    require_flash_model(model)
    if workers < 1 or max_retries < 1:
        raise RuntimeError("workers and max_retries must be at least 1")
    resolved_api_key = api_key or os.environ.get("DEEPSEEK_API_KEY", "")
    if not resolved_api_key:
        raise RuntimeError("DEEPSEEK_API_KEY is required for final row quality review")
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    workbook = load_workbook(workbook_path)
    sheet_name = find_customer_sheet(workbook)
    ws = workbook[sheet_name]
    header_row, headers = find_header(ws)
    rows = collect_rows(ws, header_row, headers)
    if not rows:
        raise RuntimeError("Customer workbook contains no product rows to review")

    decisions: list[dict[str, Any] | None] = [None] * len(rows)
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {
            executor.submit(
                review_one_row,
                row_idx=row_idx,
                row=row,
                api_key=resolved_api_key,
                model=model,
                base_url=base_url or os.environ.get("DEEPSEEK_BASE_URL", DEFAULT_DEEPSEEK_BASE_URL),
                timeout_seconds=timeout_seconds,
                max_retries=max_retries,
                caller=caller,
            ): position
            for position, (row_idx, row) in enumerate(rows)
        }
        try:
            for completed, future in enumerate(as_completed(future_map), 1):
                decisions[future_map[future]] = future.result()
                if completed == len(rows) or completed % workers == 0:
                    LOGGER.info("Final row quality checked %d/%d rows", completed, len(rows))
        except Exception:
            executor.shutdown(wait=False, cancel_futures=True)
            raise

    final_decisions = [decision for decision in decisions if decision is not None]
    if len(final_decisions) != len(rows):
        raise RuntimeError("Final row quality review did not return one decision per row")

    stats = apply_decisions(ws, header_row, headers, final_decisions)
    resolved_audit_path = audit_path or workbook_path.with_name("final_row_quality_audit.jsonl")
    write_audit(Path(resolved_audit_path), final_decisions)

    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=workbook_path.stem + ".review-",
            suffix=".xlsx",
            dir=workbook_path.parent,
            delete=False,
        ) as handle:
            temporary_path = Path(handle.name)
        workbook.save(temporary_path)
        verify_customer_workbook(
            temporary_path,
            expected_product_rows=stats["kept"],
        )
        os.replace(temporary_path, workbook_path)
    finally:
        if temporary_path and temporary_path.exists():
            temporary_path.unlink()

    return {
        **stats,
        "sheet": sheet_name,
        "model": model,
        "audit_path": str(resolved_audit_path),
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Final DeepSeek Flash row quality review")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--audit", type=Path)
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS)
    parser.add_argument("--max-retries", type=int, default=DEFAULT_MAX_RETRIES)
    parser.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT_SECONDS)
    parser.add_argument("--model", default=DEFAULT_FLASH_MODEL, choices=[DEFAULT_FLASH_MODEL])
    parser.add_argument("--base-url", default=os.environ.get("DEEPSEEK_BASE_URL", DEFAULT_DEEPSEEK_BASE_URL))
    return parser


def main() -> None:
    args = build_parser().parse_args()
    stats = review_workbook_rows(
        args.workbook,
        audit_path=args.audit,
        model=args.model,
        base_url=args.base_url,
        workers=args.workers,
        timeout_seconds=args.timeout,
        max_retries=args.max_retries,
    )
    print(json.dumps(stats, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
