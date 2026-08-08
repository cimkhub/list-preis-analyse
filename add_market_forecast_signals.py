#!/usr/bin/env python3
"""Add search-based market forecast signals to the final LIST workbook.

This is intentionally a lightweight first version: it uses the final product
list, asks DeepSeek for a small set of market search groups, searches Brave
and Tavily once per unique query, summarizes the combined evidence per group,
and writes the group signal back to the workbook as a new column.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import time
from copy import copy
from datetime import date, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import TableColumn


ROOT = Path(__file__).resolve().parent
FORECAST_COLUMN = "Marktprognose"
FINAL_SHEETS = ["Final Output", "Final Output Short"]
TARGET_CATEGORIES = {"fleisch", "fisch", "obst & gemüse", "obst & gemuese", "obst_gemuese"}
DEEPSEEK_BASE_URL = os.environ.get("DEEPSEEK_BASE_URL", "https://api.deepseek.com")
DEEPSEEK_MODEL = os.environ.get("DEEPSEEK_MODEL", "deepseek-v4-pro")
DEEPSEEK_SIGNAL_MODEL = os.environ.get("DEEPSEEK_SIGNAL_MODEL", "deepseek-v4-pro")
BRAVE_NEWS_URL = "https://api.search.brave.com/res/v1/news/search"
TAVILY_SEARCH_URL = "https://api.tavily.com/search"
REQUEST_TIMEOUT = 90
SIGNAL_CACHE_VERSION = f"brave-tavily-source-label-reasoner-v3:{DEEPSEEK_SIGNAL_MODEL}"


def main() -> None:
    load_env_file(ROOT / ".env")
    args = build_parser().parse_args()
    configure_deepseek_models(args)

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    cache_dir = Path(args.cache_dir) if args.cache_dir else workbook_path.parent / "market_forecast_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)

    brave_api_key = os.environ.get("BRAVE_API_KEY", "")
    tavily_api_key = (
        os.environ.get("TAVILY_API_KEY", "")
        or os.environ.get("TAVILY_KEY", "")
        or os.environ.get("tavily_key", "")
    )
    deepseek_api_key = os.environ.get("DEEPSEEK_API_KEY", "")
    if not deepseek_api_key or (not brave_api_key and not tavily_api_key):
        message = "DEEPSEEK_API_KEY und mindestens ein Suchanbieter (BRAVE_API_KEY oder tavily_key) werden benötigt; Marktprognose wird übersprungen."
        if args.fail_on_missing_api:
            raise RuntimeError(message)
        print(message)
        return
    if not brave_api_key:
        print("BRAVE_API_KEY fehlt; Marktprognose nutzt nur Tavily.")
    if not tavily_api_key:
        print("tavily_key/TAVILY_API_KEY fehlt; Marktprognose nutzt nur Brave.")

    wb = load_workbook(workbook_path)
    rows = collect_final_output_rows(wb)
    if not rows:
        print("Keine Fleisch/Fisch/Obst & Gemüse Zeilen für Marktprognose gefunden.")
        return

    groups = build_or_load_groups(
        rows=rows,
        cache_dir=cache_dir,
        deepseek_api_key=deepseek_api_key,
        force_refresh=args.force_refresh_groups,
        max_groups_per_category=args.max_groups_per_category,
    )
    print(f"Market forecast groups: {len(groups)}")

    freshness = build_freshness_range(args.days_back)
    query_cache = load_json(cache_dir / "brave_search_cache.json", {})
    tavily_query_cache = load_json(cache_dir / "tavily_search_cache.json", {})
    group_signals = build_group_signals(
        groups=groups,
        cache_dir=cache_dir,
        query_cache=query_cache,
        tavily_query_cache=tavily_query_cache,
        brave_api_key=brave_api_key,
        tavily_api_key=tavily_api_key,
        deepseek_api_key=deepseek_api_key,
        freshness=freshness,
        tavily_time_range=tavily_time_range(args.days_back),
        max_results_per_query=args.max_results_per_query,
        force_refresh_search=args.force_refresh_search,
        force_refresh_signals=args.force_refresh_signals,
    )
    save_json(cache_dir / "brave_search_cache.json", query_cache)
    save_json(cache_dir / "tavily_search_cache.json", tavily_query_cache)

    product_signals = map_products_to_signals(rows, groups, group_signals)
    cleared_cells = clear_forecast_columns(wb)
    updated_cells = write_forecast_columns(wb, product_signals)
    final_sheet_name = finalize_customer_workbook(wb, workbook_path)
    wb.save(workbook_path)

    print(
        f"Market forecast signals written to {workbook_path} "
        f"({updated_cells} cells updated, {cleared_cells} previous cells cleared, final sheet={final_sheet_name})"
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Add search-based market forecast signals to LIST final workbook.")
    parser.add_argument("--workbook", required=True, help="Path to Artikelvergleich workbook")
    parser.add_argument("--cache-dir", default="", help="Cache directory; default: workbook folder/market_forecast_cache")
    parser.add_argument("--deepseek-model", default=os.environ.get("DEEPSEEK_MODEL", DEEPSEEK_MODEL))
    parser.add_argument("--deepseek-signal-model", default=os.environ.get("DEEPSEEK_SIGNAL_MODEL", DEEPSEEK_SIGNAL_MODEL))
    parser.add_argument("--deepseek-base-url", default=os.environ.get("DEEPSEEK_BASE_URL", DEEPSEEK_BASE_URL))
    parser.add_argument("--days-back", type=int, default=7, help="Brave news freshness window")
    parser.add_argument("--max-results-per-query", type=int, default=8)
    parser.add_argument("--max-groups-per-category", type=int, default=8)
    parser.add_argument("--force-refresh-groups", action="store_true")
    parser.add_argument("--force-refresh-search", action="store_true")
    parser.add_argument("--force-refresh-signals", action="store_true")
    parser.add_argument("--fail-on-missing-api", action="store_true")
    return parser


def configure_deepseek_models(args: argparse.Namespace) -> None:
    global DEEPSEEK_BASE_URL, DEEPSEEK_MODEL, DEEPSEEK_SIGNAL_MODEL, SIGNAL_CACHE_VERSION
    DEEPSEEK_BASE_URL = args.deepseek_base_url
    DEEPSEEK_MODEL = args.deepseek_model
    DEEPSEEK_SIGNAL_MODEL = args.deepseek_signal_model
    SIGNAL_CACHE_VERSION = f"brave-tavily-source-label-reasoner-v3:{DEEPSEEK_MODEL}:{DEEPSEEK_SIGNAL_MODEL}"


def finalize_customer_workbook(wb, workbook_path: Path) -> str:
    """Keep only the customer-facing short output sheet and name it KWXX."""
    source_name = None
    for candidate in ["Final Output Short", final_sheet_name_from_path(workbook_path), "Final Output"]:
        if candidate in wb.sheetnames:
            source_name = candidate
            break
    if source_name is None:
        return wb.sheetnames[0] if wb.sheetnames else ""

    ws = wb[source_name]
    target_name = final_sheet_name_from_path(workbook_path)
    for sheet_name in list(wb.sheetnames):
        if sheet_name != source_name:
            del wb[sheet_name]
    ws.title = target_name
    return target_name


def final_sheet_name_from_path(path: Path) -> str:
    match = re.search(r"\bKW\s*(\d{1,2})\b", path.stem, flags=re.IGNORECASE)
    if not match:
        match = re.search(r"KW(\d{1,2})", path.stem, flags=re.IGNORECASE)
    return f"KW{int(match.group(1)):02d}" if match else "KW"


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def collect_final_output_rows(wb) -> list[dict[str, Any]]:
    sheet_name = next(iter(forecast_sheet_names(wb)), None)
    if not sheet_name:
        return []
    ws = wb[sheet_name]
    header_row, headers = find_header(ws)
    category_col = headers.get("Kategorie")
    product_col = headers.get("Produkt")
    if not category_col or not product_col:
        return []
    rows: list[dict[str, Any]] = []
    seen_keys = set()
    for row_idx in range(header_row + 1, ws.max_row + 1):
        category = str(ws.cell(row_idx, category_col).value or "").strip()
        product = str(ws.cell(row_idx, product_col).value or "").strip()
        if not product or normalize_category(category) not in TARGET_CATEGORIES:
            continue
        item = {
            "row_idx": row_idx,
            "category": category,
            "category_norm": normalize_category(category),
            "product": product,
            "description": str(ws.cell(row_idx, headers.get("Beschreibung", 0)).value or "").strip() if headers.get("Beschreibung") else "",
            "origin": str(ws.cell(row_idx, headers.get("Herkunft", 0)).value or "").strip() if headers.get("Herkunft") else "",
        }
        key = product_key(item["category_norm"], product)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        rows.append(item)
    return rows


def build_or_load_groups(
    rows: list[dict[str, Any]],
    cache_dir: Path,
    deepseek_api_key: str,
    force_refresh: bool,
    max_groups_per_category: int,
) -> list[dict[str, Any]]:
    cache_path = cache_dir / "market_groups.json"
    if cache_path.exists() and not force_refresh:
        cached = load_json(cache_path, {})
        groups = cached.get("groups") if isinstance(cached, dict) else None
        if isinstance(groups, list):
            return refine_overbroad_groups(normalize_groups(groups, rows))

    grouped_products: dict[str, list[str]] = {}
    for row in rows:
        label = category_label(row["category_norm"])
        grouped_products.setdefault(label, []).append(row["product"])

    prompt = f"""
Du bist Marktanalyst für Lebensmittel-Großhandel.

Aufgabe:
Aus den Produktnamen sollst du wenige sinnvolle Marktgruppen für Suchmaschinen-Recherchen bilden.
Die Suche soll NICHT pro Artikel laufen. Ähnliche Artikel sollen dieselbe Marktgruppe bekommen.

Beispiele:
- Rinderfilet, Rinderroastbeef, Rinderhüfte, Entrecôte -> Rindfleisch
- Schweinenacken, Schweinelachs, Schweinefilet -> Schweinefleisch
- Kalbssteakhüfte, Kalbsfilet -> Kalbfleisch
- Lammkeule, Lammlachs -> Lammfleisch
- Lachsfilet, ASC Lachs -> Lachs
- Spargel weiß, Spargel violett -> Spargel

Regeln:
- Nur Kategorien Fleisch, Fisch, Obst & Gemüse verwenden.
- Pro Kategorie maximal {max_groups_per_category} Marktgruppen.
- Jedes Produkt aus der Eingabe soll möglichst genau einer Gruppe zugeordnet werden.
- Suchbegriffe sollen Deutsch oder Deutsch+Englisch sein und auf die konkrete Warengruppe plus Preisentwicklung, Marktpreise, Angebot/Nachfrage zielen.
- Keine generischen Begriffe wie "Lebensmittelpreise", "Großhandelspreise allgemein" oder "Inflation".
- Keine Marken als eigene Marktgruppe, außer die Marke ist wirklich marktbestimmend.
- Gib striktes JSON zurück, keine Markdown-Erklärung.

JSON-Format:
{{
  "groups": [
    {{
      "group_id": "beef",
      "category": "Fleisch",
      "label": "Rindfleisch",
      "search_terms": [
        "Rindfleisch Preisentwicklung Deutschland Großhandel",
        "beef prices Europe Germany market"
      ],
      "products": ["exakter Produktname aus der Eingabe"]
    }}
  ]
}}

Produktliste:
{json.dumps(grouped_products, ensure_ascii=False, indent=2)}
""".strip()

    try:
        response = call_deepseek_json(
            api_key=deepseek_api_key,
            prompt=prompt,
            system="Du gruppierst Produktnamen für Marktpreis-Recherchen und antwortest nur mit JSON.",
        )
        groups = refine_overbroad_groups(normalize_groups(response.get("groups", []), rows))
    except Exception as exc:
        print(f"DeepSeek grouping failed, using rule fallback: {exc}")
        groups = fallback_groups(rows)

    save_json(cache_path, {"groups": groups, "created_at": time.strftime("%Y-%m-%d %H:%M:%S")})
    return groups


def normalize_groups(groups: list[dict[str, Any]], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    known_products = {row["product"] for row in rows}
    normalized = []
    used_products: set[str] = set()
    for index, group in enumerate(groups, start=1):
        label = str(group.get("label") or "").strip()
        category = str(group.get("category") or "").strip()
        products = [p for p in group.get("products", []) if isinstance(p, str) and p in known_products]
        terms = [str(term).strip() for term in group.get("search_terms", []) if str(term).strip()]
        if not label or not products or not terms:
            continue
        group_id = slugify(str(group.get("group_id") or label or f"group-{index}"))
        normalized.append({
            "group_id": group_id,
            "category": category or category_label(normalize_category(rows[0]["category"])),
            "label": label,
            "search_terms": unique_list(terms)[:3],
            "products": unique_list(products),
            "keywords": keywords_for_label(label),
        })
        used_products.update(products)

    if missing := [row for row in rows if row["product"] not in used_products]:
        normalized.extend(fallback_groups(missing))

    return merge_duplicate_groups(normalized)


def fallback_groups(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        label, group_id, keywords = infer_group(row["category_norm"], row["product"])
        key = (row["category_norm"], group_id)
        bucket = buckets.setdefault(key, {
            "group_id": group_id,
            "category": category_label(row["category_norm"]),
            "label": label,
            "search_terms": search_terms_for_group(label, row["category_norm"]),
            "products": [],
            "keywords": keywords,
        })
        bucket["products"].append(row["product"])
    return [dict(group, products=unique_list(group["products"])) for group in buckets.values()]


def merge_duplicate_groups(groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for group in groups:
        key = group["group_id"]
        target = merged.setdefault(key, {**group, "products": [], "search_terms": []})
        target["products"].extend(group.get("products", []))
        target["search_terms"].extend(group.get("search_terms", []))
        target["products"] = unique_list(target["products"])
        target["search_terms"] = unique_list(target["search_terms"])[:3]
    return list(merged.values())


def refine_overbroad_groups(groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    refined: list[dict[str, Any]] = []
    for group in groups:
        fallback_rows = [
            {
                "category_norm": normalize_category(group.get("category")),
                "product": product,
            }
            for product in group.get("products", [])
        ]
        fallback = [candidate for candidate in fallback_groups(fallback_rows) if not is_broad_group(candidate)]
        fallback_ids = {candidate["group_id"] for candidate in fallback}

        if not is_broad_group(group) and len(fallback_ids) <= 1:
            refined.append(group)
            continue

        refined.extend(fallback)
    return merge_duplicate_groups([group for group in refined if not is_broad_group(group)])


def is_broad_group(group: dict[str, Any]) -> bool:
    label = normalize_text(group.get("label", ""))
    group_id = normalize_text(group.get("group_id", ""))
    broad_labels = {
        "obst",
        "gemüse",
        "gemuese",
        "obst & gemüse",
        "obst & gemuese",
        "obst und gemüse",
        "obst & gemüse allgemein",
        "obst & gemuese allgemein",
        "meeresfrüchte",
        "meeresfruechte",
        "fisch allgemein",
        "fleisch allgemein",
        "fisch",
        "fleisch",
    }
    if label in broad_labels or "allgemein" in label:
        return True
    return group_id.endswith("_general") or group_id in {"fruit", "produce_general", "fish_general", "meat_general"}


def build_group_signals(
    groups: list[dict[str, Any]],
    cache_dir: Path,
    query_cache: dict[str, Any],
    tavily_query_cache: dict[str, Any],
    brave_api_key: str,
    tavily_api_key: str,
    deepseek_api_key: str,
    freshness: str,
    tavily_time_range: str,
    max_results_per_query: int,
    force_refresh_search: bool,
    force_refresh_signals: bool,
) -> dict[str, dict[str, str]]:
    signal_cache_path = cache_dir / "group_signals.json"
    signal_cache = load_json(signal_cache_path, {})
    signals: dict[str, dict[str, str]] = {}

    for group in groups:
        group_id = group["group_id"]
        signal_key = group_signal_cache_key(group, freshness)
        if not force_refresh_signals and signal_key in signal_cache:
            signals[group_id] = signal_cache[signal_key]
            continue

        results = []
        for query in group.get("search_terms", []):
            query_results = []
            if brave_api_key:
                query_results.append(fetch_brave_cached(
                    api_key=brave_api_key,
                    query=query,
                    freshness=freshness,
                    count=max_results_per_query,
                    query_cache=query_cache,
                    force_refresh=force_refresh_search,
                ))
            if tavily_api_key:
                query_results.append(fetch_tavily_cached(
                    api_key=tavily_api_key,
                    query=query,
                    time_range=tavily_time_range,
                    count=max_results_per_query,
                    query_cache=tavily_query_cache,
                    force_refresh=force_refresh_search,
                ))
            results.extend(interleave_result_sets(query_results))
        evidence = dedupe_search_results(results)[:18]
        signal = summarize_group_signal(group, evidence, deepseek_api_key)
        signal_cache[signal_key] = signal
        signals[group_id] = signal
        save_json(signal_cache_path, signal_cache)

    return signals


def fetch_brave_cached(
    api_key: str,
    query: str,
    freshness: str,
    count: int,
    query_cache: dict[str, Any],
    force_refresh: bool,
) -> list[dict[str, Any]]:
    cache_key = hashlib.sha1(json.dumps({"q": query, "freshness": freshness, "count": count}, sort_keys=True).encode()).hexdigest()
    if not force_refresh and cache_key in query_cache:
        return query_cache[cache_key].get("results", [])

    print(f"Brave search: {query}")
    response = requests.get(
        BRAVE_NEWS_URL,
        headers={
            "Accept": "application/json",
            "Accept-Encoding": "gzip",
            "X-Subscription-Token": api_key,
        },
        params={
            "q": query,
            "freshness": freshness,
            "country": "ALL",
            "ui_lang": "de-DE",
            "count": count,
            "offset": 0,
            "spellcheck": False,
            "extra_snippets": True,
            "safesearch": "moderate",
        },
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    data = response.json()
    results = data.get("results", []) if isinstance(data, dict) else []
    query_cache[cache_key] = {
        "query": query,
        "freshness": freshness,
        "created_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "results": results,
    }
    return results


def fetch_tavily_cached(
    api_key: str,
    query: str,
    time_range: str,
    count: int,
    query_cache: dict[str, Any],
    force_refresh: bool,
) -> list[dict[str, Any]]:
    cache_key = hashlib.sha1(json.dumps(
        {"provider": "tavily", "q": query, "time_range": time_range, "count": count},
        sort_keys=True,
    ).encode()).hexdigest()
    if not force_refresh and cache_key in query_cache:
        return clean_tavily_results(query_cache[cache_key].get("results", []))

    print(f"Tavily search: {query}")
    try:
        response = requests.post(
            TAVILY_SEARCH_URL,
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json={
                "query": query,
                "include_answer": "advanced",
                "search_depth": "advanced",
                "time_range": time_range,
                "max_results": min(max(count, 1), 20),
                "topic": "news",
            },
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        payload = response.json()
    except Exception as exc:
        print(f"Tavily search failed for {query}: {exc}")
        return []
    results = normalize_tavily_response(payload, query)[:count + 1]
    query_cache[cache_key] = {
        "query": query,
        "time_range": time_range,
        "created_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "results": results,
    }
    return results


def clean_tavily_results(results: Any) -> list[dict[str, Any]]:
    cleaned = []
    for item in results or []:
        if not isinstance(item, dict):
            continue
        if not str(item.get("url") or "").strip():
            continue
        cleaned.append(item)
    return cleaned


def normalize_tavily_response(response: Any, query: str) -> list[dict[str, Any]]:
    if not isinstance(response, dict):
        return []
    normalized: list[dict[str, Any]] = []
    for item in response.get("results") or []:
        if not isinstance(item, dict):
            continue
        if not str(item.get("url") or "").strip():
            continue
        normalized.append({
            "title": item.get("title", ""),
            "url": item.get("url", ""),
            "description": item.get("content") or item.get("snippet") or item.get("description") or "",
            "age": item.get("published_date") or "",
            "extra_snippets": [item.get("raw_content", "")] if item.get("raw_content") else [],
            "search_provider": "Tavily",
            "score": item.get("score", ""),
        })
    return normalized


def summarize_group_signal(group: dict[str, Any], evidence: list[dict[str, Any]], deepseek_api_key: str) -> dict[str, str]:
    if not evidence:
        return {"signal_text": "", "direction": "unknown", "confidence": "0", "sources": ""}

    prompt = f"""
Du bist Marktanalyst für konkrete Lebensmittel-Warengruppen in Deutschland und Europa.

Nutze AUSSCHLIESSLICH die unten bereitgestellten Suchmaschinen-/News-Treffer aus Brave und Tavily.
Erfinde nichts hinzu. Nutze kein Weltwissen außerhalb dieser Treffer.

Warengruppe: {group.get("label")}
Kategorie: {group.get("category")}
Beispielprodukte: {", ".join(group.get("products", [])[:12])}

Aufgabe:
- Bewerte nur Treffer, die wirklich diese konkrete Warengruppe betreffen: {group.get("label")}.
- Die News muss konkret eine Preisentwicklung, Marktpreise, Angebot/Nachfrage, Produktion, Ernte, Fangmenge, Export-/Importlage oder Tierseuchenlage für diese Warengruppe betreffen.
- Ignoriere generische Meldungen zu Inflation, allgemeinen Lebensmittelpreisen, allgemeinen Großhandelspreisen, Verbraucherpreisen oder Gastro-/Retail-Trends, wenn die konkrete Warengruppe nicht im Fokus steht.
- Ignoriere Treffer, die nur entfernte Kategorien betreffen. Beispiel: Gemüse allgemein ist nicht automatisch Spargel; Fisch allgemein ist nicht automatisch Lachs; Fleisch allgemein ist nicht automatisch Rindfleisch.
- Ignoriere irrelevante, werbliche oder zu indirekte Treffer.
- Wenn kein belastbares produktspezifisches Preissignal vorhanden ist, gib signal_text als leeren String zurück.
- Nutze nur Treffer, die als aktuelle News innerhalb der letzten 7 Tage erkennbar sind.
- Schreibe vollständig auf Deutsch. Übersetze englische Treffer sinngemäß ins Deutsche.
- signal_text enthält maximal 1 kurzen Bullet Point.
- Jeder Bullet Point beginnt mit "+" für eher steigende Preise, "-" für eher sinkende Preise oder "=" für stabil/unklar.
- Jeder Bullet Point endet mit einer Quelle in eckigen Klammern.
- Verwende dafür exakt das Feld source_label aus den verwendeten Treffern, z.B. [Brave / topagrar.com] oder [Tavily / topagrar.com].
- Gib niemals nur eine Domain ohne Anbieter aus.
- Wenn Brave und Tavily dieselbe konkrete Aussage stützen, zitiere beide Labels, z.B. [Brave / topagrar.com; Tavily / topagrar.com].
- Keine Einleitung.
Do not mention a source if it is not directly relevant to the concrete group.

Return strict JSON:
{{
  "signal_text": "+ ... [Brave / domain.de]",
  "direction": "up|down|mixed|stable|unknown",
  "confidence": "0-100",
  "sources": "domain1.de, domain2.de"
}}

Treffer:
{json.dumps([compact_search_item(item) for item in evidence], ensure_ascii=False, indent=2)}
""".strip()

    try:
        data = call_deepseek_json(
            api_key=deepseek_api_key,
            prompt=prompt,
            system="Du bist Marktanalyst und antwortest nur mit strikt gültigem JSON.",
            model=DEEPSEEK_SIGNAL_MODEL,
            timeout=240,
        )
        signal_text = normalize_signal_text(str(data.get("signal_text") or "").strip())
        if signal_text and looks_english_signal(signal_text):
            signal_text = translate_signal_to_german(signal_text, deepseek_api_key)
        if signal_text and is_generic_or_offtopic_signal(signal_text, group):
            signal_text = ""
        return {
            "signal_text": signal_text,
            "direction": str(data.get("direction") or "unknown").strip(),
            "confidence": str(data.get("confidence") or "0").strip(),
            "sources": str(data.get("sources") or "").strip(),
        }
    except Exception as exc:
        print(f"DeepSeek signal failed for {group.get('label')}: {exc}")
        return {"signal_text": "", "direction": "unknown", "confidence": "0", "sources": ""}


def normalize_signal_text(value: str) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    if not text:
        return ""
    # DeepSeek sometimes returns multiple semicolon-separated bullets despite the prompt.
    text = re.sub(r";\s*([+\-=])\s+", r"\n\1 ", text)
    lines = [normalize_signal_bullet(line) for line in text.splitlines() if line.strip()]
    lines = [line for line in lines if line]
    bullet_lines = [line for line in lines if line[:1] in {"+", "-", "±"}]
    if bullet_lines:
        return bullet_lines[0]
    return lines[0] if lines else ""


def normalize_signal_bullet(value: str) -> str:
    text = str(value or "").strip()
    # "=" is a valid market shorthand for stable prices, but Excel treats a
    # leading equals sign as a formula. Store stable signals with a text marker.
    text = re.sub(r"^=\s*@?\s*", "± ", text)
    text = re.sub(r"^@\s*", "", text)
    return text.strip()


def looks_english_signal(signal_text: str) -> bool:
    text = f" {normalize_text(signal_text)} "
    english_markers = [
        " prices ",
        " price ",
        " demand ",
        " supply ",
        " exports ",
        " imports ",
        " global ",
        " projected ",
        " amid ",
        " weak ",
        " boost ",
        " fall ",
        " rise ",
        " rising ",
    ]
    return any(marker in text for marker in english_markers)


def translate_signal_to_german(signal_text: str, deepseek_api_key: str) -> str:
    prompt = f"""
Übersetze den folgenden Marktprognose-Bullet vollständig ins Deutsche.
Behalte das führende Zeichen (+, - oder =) und die Quellen in eckigen Klammern unverändert.
Keine Zusatzinfos, keine Erklärung. Gib strikt JSON zurück.

JSON:
{{"signal_text": "+ deutscher Text [Quelle]"}}

Bullet:
{signal_text}
""".strip()
    try:
        data = call_deepseek_json(
            api_key=deepseek_api_key,
            prompt=prompt,
            system="Du übersetzt kurze Marktprognose-Bullets ins Deutsche und antwortest nur mit JSON.",
        )
        return normalize_signal_text(str(data.get("signal_text") or signal_text).strip())
    except Exception:
        return signal_text


def is_generic_or_offtopic_signal(signal_text: str, group: dict[str, Any]) -> bool:
    text = normalize_text(signal_text)
    if not text:
        return True
    price_terms = [
        "preis",
        "preise",
        "preisentwicklung",
        "erzeugerpreise",
        "großhandelspreise",
        "grosshandelspreise",
        "marktpreise",
        "teuer",
        "günstig",
        "guenstig",
        "billig",
        "euro",
        "€/kg",
        "price",
        "prices",
    ]
    if not any(term in text for term in price_terms):
        return True
    price_or_market_terms = [
        *price_terms,
        "markt",
        "angebot",
        "nachfrage",
        "ernte",
        "fang",
        "produktion",
        "export",
        "import",
        "seuche",
        "überangebot",
        "ueberangebot",
        "knapp",
        "gestiegen",
        "gesunken",
        "fallen",
        "rückläufig",
        "ruecklaeufig",
        "demand",
        "supply",
        "glut",
        "prices",
        "price",
    ]
    if not any(term in text for term in price_or_market_terms):
        return True
    generic_terms = [
        "lebensmittelpreise",
        "verbraucherpreise",
        "inflation",
        "grosshandelspreise insgesamt",
        "großhandelspreise insgesamt",
        "großhandelspreise stiegen",
        "grosshandelspreise stiegen",
        "preise für waren insgesamt",
        "waren insgesamt",
        "gastro",
    ]
    if any(term in text for term in generic_terms):
        keywords = set(group.get("keywords") or []) | set(keywords_for_label(group.get("label", "")))
        if not any(keyword and keyword in text for keyword in keywords):
            return True
    label = normalize_text(group.get("label", ""))
    keywords = set(group.get("keywords") or []) | set(keywords_for_label(label))
    strong_terms = {term for term in keywords if len(term) >= 4}
    if strong_terms and not any(term in text for term in strong_terms):
        # Allow explicit English market labels from common groups.
        english_aliases = {
            "Rindfleisch": ["beef", "cattle"],
            "Schweinefleisch": ["pork", "pig", "swine"],
            "Kalbfleisch": ["veal", "calf"],
            "Lammfleisch": ["lamb", "sheep"],
            "Geflügelfleisch": ["poultry", "chicken", "turkey"],
            "Lachs": ["salmon"],
            "Weißfisch": ["whitefish", "cod", "pollock"],
            "Meeresfrüchte": ["seafood", "shrimp", "prawn"],
            "Gemüse": ["vegetable", "vegetables"],
            "Obst": ["fruit", "fruits"],
        }.get(str(group.get("label") or ""), [])
        if not any(alias in text for alias in english_aliases):
            return True
    return False


def map_products_to_signals(rows: list[dict[str, Any]], groups: list[dict[str, Any]], group_signals: dict[str, dict[str, str]]) -> dict[tuple[str, str], str]:
    product_to_group: dict[str, str] = {}
    for group in groups:
        for product in group.get("products", []):
            product_to_group[normalize_text(product)] = group["group_id"]

    signals: dict[tuple[str, str], str] = {}
    for row in rows:
        group_id = product_to_group.get(normalize_text(row["product"])) or infer_group(row["category_norm"], row["product"])[1]
        signal = group_signals.get(group_id, {}).get("signal_text", "")
        if signal:
            signals[product_key(row["category_norm"], row["product"])] = signal
    return signals


def write_forecast_columns(wb, product_signals: dict[tuple[str, str], str]) -> int:
    updated = 0
    for sheet_name in forecast_sheet_names(wb):
        ws = wb[sheet_name]
        header_row, headers = find_header(ws)
        if not headers:
            continue
        forecast_col = headers.get(FORECAST_COLUMN)
        if not forecast_col:
            forecast_col = ws.max_column + 1
            headers[FORECAST_COLUMN] = forecast_col
            copy_column_style(ws, forecast_col - 1, forecast_col, header_row)
            ws.cell(header_row, forecast_col).value = FORECAST_COLUMN
            ws.column_dimensions[get_column_letter(forecast_col)].width = 48
        category_col = headers.get("Kategorie")
        product_col = headers.get("Produkt")
        if not category_col or not product_col:
            continue

        style_forecast_column(ws, forecast_col, header_row)
        ws.cell(header_row, forecast_col).value = FORECAST_COLUMN
        for row_idx in range(header_row + 1, ws.max_row + 1):
            category = normalize_category(ws.cell(row_idx, category_col).value)
            product = str(ws.cell(row_idx, product_col).value or "").strip()
            signal = product_signals.get(product_key(category, product), "")
            cell = ws.cell(row_idx, forecast_col)
            set_excel_text(cell, signal)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if signal:
                updated += 1
                ws.row_dimensions[row_idx].height = min(180, max(float(ws.row_dimensions[row_idx].height or 22), 16 * (signal.count("\n") + 2)))

        last_col = get_column_letter(ws.max_column)
        table_ref = f"A{header_row}:{last_col}{ws.max_row}"
        sync_tables_to_range(ws, header_row, table_ref)
        if ws.tables:
            ws.auto_filter.ref = None
        elif ws.auto_filter and ws.auto_filter.ref:
            ws.auto_filter.ref = table_ref
        ws.cell(header_row, forecast_col).value = FORECAST_COLUMN
    return updated


def set_excel_text(cell, value: Any) -> None:
    text = "" if value is None else str(value)
    if text.startswith("="):
        text = normalize_signal_bullet(text)
    cell.value = text
    cell.data_type = "s"


def sync_tables_to_range(ws, header_row: int, table_ref: str) -> None:
    """Keep Excel table XML in sync after adding the forecast column.

    Excel repairs workbooks when a sheet/table filter range points to a
    different number of columns than the table definition. This can happen when
    the forecast step appends "Marktprognose" after the workbook was already
    generated by the matching step.
    """
    for table_name in list(ws.tables.keys()):
        table = ws.tables[table_name]
        if int(re.search(r"\d+", table.ref or "0").group(0)) != header_row:
            continue
        table.ref = table_ref
        if table.autoFilter:
            table.autoFilter.ref = table_ref
        ensure_table_columns(ws, table, header_row)


def ensure_table_columns(ws, table, header_row: int) -> None:
    start_col, end_col = table.ref.split(":", 1)
    start_idx = column_index_from_cell(start_col)
    end_idx = column_index_from_cell(end_col)
    expected_count = end_idx - start_idx + 1
    existing = list(table.tableColumns)
    existing_by_id = {int(column.id): column for column in existing}
    columns = []
    for offset, col_idx in enumerate(range(start_idx, end_idx + 1), start=1):
        header = str(ws.cell(header_row, col_idx).value or f"Column{offset}").strip() or f"Column{offset}"
        column = existing_by_id.get(offset) or TableColumn(id=offset, name=header)
        column.id = offset
        column.name = header
        columns.append(column)
    table.tableColumns = columns


def column_index_from_cell(cell_ref: str) -> int:
    letters = re.sub(r"[^A-Z]", "", str(cell_ref).upper())
    total = 0
    for char in letters:
        total = total * 26 + (ord(char) - 64)
    return total


def clear_forecast_columns(wb) -> int:
    cleared = 0
    for sheet_name in forecast_sheet_names(wb):
        ws = wb[sheet_name]
        header_row, headers = find_header(ws)
        forecast_col = headers.get(FORECAST_COLUMN)
        if not forecast_col:
            continue
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if ws.cell(row_idx, forecast_col).value:
                cleared += 1
            ws.cell(row_idx, forecast_col).value = ""
    return cleared


def find_header(ws) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 15) + 1):
        headers = {
            str(ws.cell(row_idx, col_idx).value or "").strip(): col_idx
            for col_idx in range(1, ws.max_column + 1)
            if str(ws.cell(row_idx, col_idx).value or "").strip()
        }
        if "Kategorie" in headers and "Produkt" in headers and any(col in headers for col in ["Metro", "Selgros", "Handelshof", "Edeka"]):
            return row_idx, headers
    return 7, {}


def forecast_sheet_names(wb) -> list[str]:
    names = [name for name in FINAL_SHEETS if name in wb.sheetnames]
    names.extend(name for name in wb.sheetnames if re.fullmatch(r"KW\d{1,2}", name) and name not in names)
    return names


def copy_column_style(ws, source_col: int, target_col: int, header_row: int) -> None:
    if source_col < 1:
        return
    for row_idx in range(1, ws.max_row + 1):
        source = ws.cell(row_idx, source_col)
        target = ws.cell(row_idx, target_col)
        if source.has_style:
            target._style = copy(source._style)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.border:
            target.border = copy(source.border)
        if source.fill:
            target.fill = copy(source.fill)
        if source.font:
            target.font = copy(source.font)
    ws.cell(header_row, target_col).value = FORECAST_COLUMN


def style_forecast_column(ws, forecast_col: int, header_row: int) -> None:
    ws.cell(header_row, forecast_col).alignment = Alignment(wrap_text=True, vertical="center")
    for row_idx in range(header_row + 1, ws.max_row + 1):
        ws.cell(row_idx, forecast_col).alignment = Alignment(wrap_text=True, vertical="top")


def call_deepseek_json(
    api_key: str,
    prompt: str,
    system: str,
    model: str | None = None,
    timeout: int = 180,
) -> dict[str, Any]:
    response = requests.post(
        DEEPSEEK_BASE_URL.rstrip("/") + "/chat/completions",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={
            "model": model or DEEPSEEK_MODEL,
            "messages": [
                {"role": "system", "content": system},
                {"role": "user", "content": prompt},
            ],
            "stream": False,
            "response_format": {"type": "json_object"},
        },
        timeout=timeout,
    )
    response.raise_for_status()
    payload = response.json()
    content = payload["choices"][0]["message"]["content"]
    return parse_json_object(content)


def parse_json_object(text: str) -> dict[str, Any]:
    cleaned = str(text or "").strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?", "", cleaned).strip()
        cleaned = re.sub(r"```$", "", cleaned).strip()
    try:
        data = json.loads(cleaned)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
        if not match:
            raise
        data = json.loads(match.group(0))
    if not isinstance(data, dict):
        raise RuntimeError("DeepSeek JSON response was not an object")
    return data


def compact_search_item(item: dict[str, Any]) -> dict[str, Any]:
    provider = str(item.get("search_provider") or "Brave").strip() or "Brave"
    source = hostname_from_item(item) or item.get("source", "")
    return {
        "source_label": f"{provider} / {source or 'Quelle'}",
        "provider": provider,
        "title": item.get("title", ""),
        "source": source,
        "age": item.get("age", ""),
        "url": item.get("url", ""),
        "description": item.get("description", ""),
        "snippets": (item.get("extra_snippets") or [])[:2],
    }


def dedupe_search_results(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen_urls = set()
    seen_titles = set()
    deduped = []
    for item in items:
        url = normalize_url(item.get("url", ""))
        title = normalize_text(item.get("title", ""))
        if (url and url in seen_urls) or (title and title in seen_titles):
            continue
        if url:
            seen_urls.add(url)
        if title:
            seen_titles.add(title)
        deduped.append(item)
    return deduped


def interleave_result_sets(result_sets: list[list[dict[str, Any]]]) -> list[dict[str, Any]]:
    """Mix providers so the LLM does not see all Brave evidence before Tavily."""
    mixed: list[dict[str, Any]] = []
    max_len = max((len(items) for items in result_sets), default=0)
    for index in range(max_len):
        for items in result_sets:
            if index < len(items):
                mixed.append(items[index])
    return mixed


def infer_group(category: str, product: str) -> tuple[str, str, list[str]]:
    text = normalize_text(product)
    if category == "fleisch":
        rules = [
            ("Kalbfleisch", "veal", ["kalb"]),
            ("Rindfleisch", "beef", ["rind", "beef", "roastbeef", "entrecote", "entrecôte", "hüfte", "huefte", "tafelspitz"]),
            ("Schweinefleisch", "pork", ["schwein", "kasseler", "duroc"]),
            ("Geflügel", "poultry", ["pute", "hähnchen", "haehnchen", "huhn", "geflügel", "gefluegel"]),
            ("Lammfleisch", "lamb", ["lamm", "schaf"]),
        ]
        for label, group_id, keywords in rules:
            if any(keyword in text for keyword in keywords):
                return label, group_id, keywords
        return "Fleisch allgemein", "meat_general", ["fleisch"]
    if category == "fisch":
        rules = [
            ("Weißfisch", "whitefish", ["seelachs", "kabeljau", "skrei"]),
            ("Forelle", "trout", ["forelle"]),
            ("Dorade", "sea_bream", ["dorade"]),
            ("Lachs", "salmon", ["lachs", "salmon"]),
            ("Thunfisch", "tuna", ["thun", "tuna"]),
            ("Garnelen", "shrimp", ["garnele", "shrimp", "scampi"]),
        ]
        for label, group_id, keywords in rules:
            if any(keyword in text for keyword in keywords):
                return label, group_id, keywords
        return "Fisch allgemein", "fish_general", ["fisch"]
    rules = [
        ("Spargel", "asparagus", ["spargel"]),
        ("Tomaten", "tomatoes", ["tomate"]),
        ("Gurken", "cucumbers", ["gurke"]),
        ("Süßkartoffeln", "sweet_potatoes", ["süßkartoffel", "suesskartoffel"]),
        ("Kartoffeln", "potatoes", ["kartoffel"]),
        ("Zwiebeln", "onions", ["zwiebel"]),
        ("Salat", "lettuce", ["salat", "rucola", "spinat"]),
        ("Kräuter", "herbs", ["kräuter", "kraeuter"]),
        ("Beeren", "berries", ["erdbeere", "beere"]),
        ("Zitrusfrüchte", "citrus", ["zitrone", "limette", "orange"]),
        ("Melonen", "melons", ["melone"]),
        ("Bananen", "bananas", ["banane"]),
        ("Avocados", "avocados", ["avocado"]),
        ("Kiwi", "kiwi", ["kiwi"]),
        ("Steinobst", "stone_fruit", ["aprikose", "pfirsich", "nektarine"]),
        ("Birnen", "pears", ["birne"]),
        ("Austernpilze", "oyster_mushrooms", ["austernpilz"]),
        ("Pilze", "mushrooms", ["champignon", "pilz"]),
        ("Paprika", "peppers", ["paprika"]),
        ("Möhren", "carrots", ["möhre", "moehre", "karotte"]),
        ("Kohl", "cabbage", ["kohl", "kohlrabi"]),
        ("Radieschen", "radishes", ["radieschen"]),
        ("Zucchini", "zucchini", ["zucchini"]),
        ("Auberginen", "eggplant", ["aubergine"]),
    ]
    for label, group_id, keywords in rules:
        if any(keyword in text for keyword in keywords):
            return label, group_id, keywords
    return "Obst & Gemüse allgemein", "produce_general", ["obst", "gemüse"]


def search_terms_for_group(label: str, category: str) -> list[str]:
    suffix = "Preisentwicklung Marktpreise Deutschland"
    english = {
        "Rindfleisch": "beef prices Europe Germany market",
        "Schweinefleisch": "pork prices Europe Germany market",
        "Kalbfleisch": "veal prices Europe Germany market",
        "Lammfleisch": "lamb prices Europe Germany market",
        "Geflügel": "poultry prices Europe Germany market",
        "Lachs": "salmon prices Europe Germany market",
        "Garnelen": "shrimp prices Europe Germany market",
        "Spargel": "asparagus prices Germany market",
        "Tomaten": "tomato prices Germany market",
    }.get(label)
    terms = [f"{label} {suffix}", f"{label} Angebot Nachfrage Preis"]
    if english:
        terms.append(english)
    return terms[:3]


def keywords_for_label(label: str) -> list[str]:
    return [part for part in re.split(r"\W+", normalize_text(label)) if len(part) > 2]


def category_label(category_norm: str) -> str:
    if category_norm == "fleisch":
        return "Fleisch"
    if category_norm == "fisch":
        return "Fisch"
    return "Obst & Gemüse"


def normalize_category(value: Any) -> str:
    text = normalize_text(value).replace("_", " ")
    if text == "fleisch":
        return "fleisch"
    if text == "fisch":
        return "fisch"
    if "obst" in text and ("gemüse" in text or "gemuese" in text):
        return "obst & gemüse"
    return text


def product_key(category: str, product: str) -> tuple[str, str]:
    return normalize_category(category), normalize_text(product)


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().casefold())


def slugify(value: str) -> str:
    text = normalize_text(value)
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_") or "group"


def unique_list(values: list[str]) -> list[str]:
    seen = set()
    result = []
    for value in values:
        value = str(value or "").strip()
        key = normalize_text(value)
        if not value or key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def build_freshness_range(days_back: int) -> str:
    end_date = date.today()
    start_date = end_date - timedelta(days=days_back)
    return f"{start_date.isoformat()}to{end_date.isoformat()}"


def tavily_time_range(days_back: int) -> str:
    if days_back <= 7:
        return "week"
    if days_back <= 31:
        return "month"
    return "year"


def group_signal_cache_key(group: dict[str, Any], freshness: str) -> str:
    payload = {
        "version": SIGNAL_CACHE_VERSION,
        "group_id": group.get("group_id"),
        "search_terms": group.get("search_terms"),
        "freshness": freshness,
    }
    return hashlib.sha1(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()


def hostname_from_item(item: dict[str, Any]) -> str:
    meta = item.get("meta_url", {}) if isinstance(item, dict) else {}
    host = (meta.get("hostname") or "").lower()
    if host:
        return host
    host = urlparse(str(item.get("url", ""))).netloc.lower()
    if host:
        return host
    return str(item.get("source") or "").lower()


def normalize_url(url: Any) -> str:
    return str(url or "").split("?", 1)[0].rstrip("/")


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def save_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
