#!/usr/bin/env python3
"""Extract one cached PDF into a standalone CSV/XLSX using per-page subprocess timeouts."""

from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))

from src.config import load_config
from src.convert.pdf_to_images import pdf_to_images
from src.extract.cached_batch import CachedExtractionTarget, load_downloaded_documents
from src.extract.text_extract import classify_category
from src.extract.vision import (
    QUALITY_RETRY_MODEL,
    PageExtractionOutcome,
    _raw_items_to_products,
    apply_quality_retry_once,
    configure_gemini,
    extraction_quality_issues,
    extract_products_from_image,
    file_sha256,
)
from src.report.parsed_csv import PARSED_CSV_FIELDS, serialize_product_row
from src.utils.logging_setup import log_event, setup_logging


ROOT = Path(__file__).resolve().parent


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extract a single cached PDF into a standalone CSV/XLSX."
    )
    parser.add_argument("--pdf", type=Path, help="Cached PDF path")
    parser.add_argument(
        "--output",
        type=Path,
        help="Standalone CSV output path. Defaults to parsed/KWxx_yyyy/<pdf_stem>_single.csv",
    )
    parser.add_argument("--workers", type=int, default=10, help="Concurrent page workers")
    parser.add_argument(
        "--dpi",
        type=int,
        default=180,
        help="Render DPI for page images. Default: 180",
    )
    parser.add_argument(
        "--page-timeout",
        type=int,
        default=300,
        help="Timeout per page subprocess in seconds",
    )
    parser.add_argument(
        "--page-retries",
        type=int,
        default=3,
        help="Retry count per page subprocess",
    )
    parser.add_argument("--config", type=str, default="config.yaml", help="Config file path")

    parser.add_argument("--worker", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument(
        "--worker-mode",
        choices=["primary", "quality_retry"],
        default="primary",
        help=argparse.SUPPRESS,
    )
    parser.add_argument("--quality-issues-json", type=str, help=argparse.SUPPRESS)
    parser.add_argument("--page-image", type=Path, help=argparse.SUPPRESS)
    parser.add_argument("--supplier", type=str, help=argparse.SUPPRESS)
    parser.add_argument("--source-file", type=str, help=argparse.SUPPRESS)
    parser.add_argument("--valid-from", type=str, help=argparse.SUPPRESS)
    parser.add_argument("--valid-to", type=str, help=argparse.SUPPRESS)
    parser.add_argument("--week", type=int, help=argparse.SUPPRESS)
    parser.add_argument("--year", type=int, help=argparse.SUPPRESS)
    parser.add_argument("--location", type=str, help=argparse.SUPPRESS)
    parser.add_argument("--title", type=str, help=argparse.SUPPRESS)
    parser.add_argument("--tab", type=str, help=argparse.SUPPRESS)
    parser.add_argument("--fallback-category", type=str, help=argparse.SUPPRESS)
    return parser


def infer_cached_document(pdf_path: Path, config) -> object:
    try:
        supplier = pdf_path.parts[-4]
        year = int(pdf_path.parts[-3])
        week = int(pdf_path.parts[-2])
    except (IndexError, ValueError) as exc:
        raise RuntimeError(f"Expected cached PDF path like data/<supplier>/<year>/<week>/<file>.pdf: {pdf_path}") from exc

    target = CachedExtractionTarget(
        supplier=supplier,
        week=week,
        year=year,
        data_dir=pdf_path.parent,
    )
    documents = load_downloaded_documents(target, config.suppliers[supplier].model_dump())
    for document in documents:
        if document.file_path and Path(document.file_path).resolve() == pdf_path.resolve():
            return document
    raise RuntimeError(f"Could not resolve cached document metadata for {pdf_path}")


def default_output_path(document) -> Path:
    week = int(document.calendar_week or 0)
    year = int(document.year or 0)
    return ROOT / "parsed" / f"KW{week:02d}_{year}" / f"{Path(document.file_path).stem}_single.csv"


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=PARSED_CSV_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in PARSED_CSV_FIELDS})


def write_xlsx(csv_path: Path) -> Path:
    xlsx_path = csv_path.with_suffix(".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "products"

    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))

    for row in rows:
        ws.append(row)

    if rows:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    for idx, column in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
        start=1,
    ):
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 40)

    wb.save(xlsx_path)
    return xlsx_path


def page_number_from_path(page_image: Path) -> int:
    stem = page_image.stem
    suffix = stem.split("-")[-1]
    return int(suffix)


def run_worker(args) -> None:
    config = load_config(args.config)
    configure_gemini(
        config.vision_api.api_key,
        model_name=config.vision_api.model,
        max_retries=config.vision_api.max_retries,
        temperature=config.vision_api.temperature,
        max_concurrent_requests=1,
        min_request_interval_seconds=config.vision_api.min_request_interval_seconds,
    )

    worker_mode = getattr(args, "worker_mode", "primary")
    if worker_mode == "quality_retry":
        try:
            quality_issues = json.loads(args.quality_issues_json or "[]")
        except (TypeError, ValueError, json.JSONDecodeError):
            quality_issues = []
        raw_items = extract_products_from_image(
            str(args.page_image),
            supplier=args.supplier,
            model_name=QUALITY_RETRY_MODEL,
            max_retries=1,
            temperature=None,
            operation="product_extraction_quality_retry",
            quality_issues=quality_issues,
        )
        selected_model = QUALITY_RETRY_MODEL
    else:
        raw_items = extract_products_from_image(
            str(args.page_image),
            supplier=args.supplier,
            model_name=config.vision_api.model,
            max_retries=config.vision_api.max_retries,
            temperature=config.vision_api.temperature,
            operation="product_extraction",
        )
        selected_model = config.vision_api.model

    sys.stdout.write(
        json.dumps(
            {
                "status": "ok" if raw_items is not None else "failed",
                "items": raw_items,
                "model_name": selected_model,
                "worker_mode": worker_mode,
            },
            ensure_ascii=False,
        )
    )


def run_page_subprocess(
    script_path: Path,
    config_path: str,
    page_image: Path,
    document,
    fallback_category: str | None,
    timeout_seconds: int,
    retries: int,
    worker_mode: str = "primary",
    quality_issues: list[str] | None = None,
) -> list[dict] | None:
    command = [
        sys.executable,
        str(script_path),
        "--worker",
        "--worker-mode",
        worker_mode,
        "--config",
        config_path,
        "--page-image",
        str(page_image),
        "--supplier",
        document.supplier,
        "--source-file",
        str(document.file_path),
        "--week",
        str(document.calendar_week or 0),
        "--year",
        str(document.year or 0),
        "--location",
        document.location or "",
        "--title",
        document.title or "",
        "--tab",
        document.tab or "",
    ]
    if document.valid_from:
        command.extend(["--valid-from", document.valid_from.isoformat()])
    if document.valid_to:
        command.extend(["--valid-to", document.valid_to.isoformat()])
    if fallback_category:
        command.extend(["--fallback-category", fallback_category])
    if worker_mode == "quality_retry":
        command.extend(
            [
                "--quality-issues-json",
                json.dumps(quality_issues or [], ensure_ascii=False),
            ]
        )

    last_error = None
    attempt_limit = max(1, retries) if worker_mode == "primary" else 1
    for attempt in range(1, attempt_limit + 1):
        try:
            result = subprocess.run(
                command,
                cwd=str(ROOT),
                capture_output=True,
                text=True,
                encoding="utf-8",
                timeout=timeout_seconds,
                check=True,
            )
            payload = json.loads(result.stdout or "{}")
            if isinstance(payload, list):
                # Compatibility with pre-P6 worker output.
                return payload
            if not isinstance(payload, dict) or payload.get("status") != "ok":
                raise RuntimeError("worker returned failed extraction status")
            items = payload.get("items")
            if not isinstance(items, list):
                raise RuntimeError("worker returned invalid items payload")
            return items
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            print(
                f"{worker_mode} attempt {attempt}/{attempt_limit} failed for {page_image.name}: {exc}",
                flush=True,
            )

    print(
        f"{worker_mode} extraction failed for {page_image.name}: {last_error}",
        flush=True,
    )
    return None


def apply_single_page_quality_retry(
    *,
    primary_items: list[dict] | None,
    primary_model: str,
    script_path: Path,
    config_path: str,
    page_image: Path,
    document,
    fallback_category: str | None,
    timeout_seconds: int,
) -> PageExtractionOutcome:
    """Use the shared coordinator while isolating the one retry in a subprocess."""

    def quality_retry_executor(_image_path, _supplier, **kwargs):
        return run_page_subprocess(
            script_path,
            config_path,
            page_image,
            document,
            fallback_category,
            timeout_seconds,
            1,
            worker_mode="quality_retry",
            quality_issues=kwargs.get("quality_issues") or [],
        )

    return apply_quality_retry_once(
        primary_items,
        image_path=str(page_image),
        supplier=document.supplier,
        primary_model=primary_model,
        retry_executor=quality_retry_executor,
    )


def write_page_outcome_manifest(
    csv_path: Path,
    records: list[dict],
) -> Path:
    manifest_path = csv_path.with_suffix(".extraction_outcomes.jsonl")
    tmp_path = manifest_path.with_suffix(manifest_path.suffix + ".tmp")
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    with tmp_path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n")
    tmp_path.replace(manifest_path)
    return manifest_path


def run_parent(args) -> None:
    config = load_config(args.config)
    logger = setup_logging(config.storage.logs_dir)
    pdf_path = args.pdf.resolve()
    document = infer_cached_document(pdf_path, config)

    week_part = f"KW{document.calendar_week:02d}" if document.calendar_week is not None else "KW00"
    year_part = str(document.year or "unknown")
    image_dir_name = pdf_path.stem if args.dpi == 300 else f"{pdf_path.stem}_dpi{args.dpi}"
    images_dir = ROOT / "images" / document.supplier / year_part / week_part / image_dir_name
    image_paths = [Path(path) for path in pdf_to_images(str(pdf_path), str(images_dir), dpi=args.dpi)]
    if not image_paths:
        raise RuntimeError(f"No page images available for {pdf_path}")

    fallback_category = classify_category(document.category) if document.category else None
    output_csv = args.output.resolve() if args.output else default_output_path(document)
    script_path = Path(__file__).resolve()
    source_document_sha256 = file_sha256(pdf_path)
    if not source_document_sha256:
        raise RuntimeError(f"Could not hash source PDF: {pdf_path}")

    print(f"Extracting {len(image_paths)} pages from {pdf_path.name}", flush=True)
    primary_items_by_page: dict[int, list[dict] | None] = {}

    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
        futures = {
            executor.submit(
                run_page_subprocess,
                script_path,
                args.config,
                page_image,
                document,
                fallback_category,
                args.page_timeout,
                args.page_retries,
            ): page_image
            for page_image in image_paths
        }

        completed = 0
        for future in as_completed(futures):
            page_image = futures[future]
            primary_items = future.result()
            primary_items_by_page[page_number_from_path(page_image)] = primary_items
            completed += 1
            print(f"Primary processed {completed}/{len(image_paths)} pages", flush=True)

    outcomes_by_page: dict[int, PageExtractionOutcome] = {}
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
        outcome_futures = {
            executor.submit(
                apply_single_page_quality_retry,
                primary_items=primary_items_by_page.get(page_number_from_path(page_image)),
                primary_model=config.vision_api.model,
                script_path=script_path,
                config_path=args.config,
                page_image=page_image,
                document=document,
                fallback_category=fallback_category,
                timeout_seconds=args.page_timeout,
            ): page_image
            for page_image in image_paths
        }
        for future in as_completed(outcome_futures):
            page_image = outcome_futures[future]
            page_number = page_number_from_path(page_image)
            outcomes_by_page[page_number] = future.result()

    rows_by_page: dict[int, list[dict[str, object]]] = {}
    manifest_records: list[dict] = []
    image_by_page = {page_number_from_path(path): path for path in image_paths}
    for page_num in sorted(image_by_page):
        page_image = image_by_page[page_num]
        outcome = outcomes_by_page[page_num]
        products = _raw_items_to_products(
            outcome.selected_items,
            supplier=document.supplier,
            source_file=str(document.file_path or pdf_path),
            source_page=page_num,
            valid_from=document.valid_from,
            valid_to=document.valid_to,
            calendar_week=document.calendar_week,
            year=document.year,
            location=document.location,
            source_title=document.title,
            source_tab=document.tab,
            fallback_category=fallback_category,
            source_document_sha256=source_document_sha256,
            extraction_outcome=outcome,
        )
        rows_by_page[page_num] = [serialize_product_row(product) for product in products]
        manifest_record = outcome.to_manifest_record(
            supplier=document.supplier,
            source_file=str(document.file_path or pdf_path),
            source_document_sha256=source_document_sha256,
            source_page=page_num,
            image_path=str(page_image),
        )
        manifest_record["accepted_product_count"] = len(products)
        manifest_record["page_complete"] = bool(outcome.selected_items) and (
            len(products) == len(outcome.selected_items)
        ) and not extraction_quality_issues(outcome.selected_items)
        manifest_records.append(manifest_record)
        log_event(
            logger,
            f"Single-PDF page outcome recorded for page {page_num}",
            event="vision_page_outcome",
            status=outcome.quality_retry_status,
            **manifest_record,
        )

    all_rows: list[dict[str, str]] = []
    for page_num in sorted(rows_by_page):
        all_rows.extend(rows_by_page[page_num])

    write_csv(output_csv, all_rows)
    output_xlsx = write_xlsx(output_csv)
    outcome_manifest = write_page_outcome_manifest(output_csv, manifest_records)
    logger.info("Standalone single-PDF extraction completed: %s (%d rows)", output_csv, len(all_rows))
    print(f"CSV={output_csv}")
    print(f"XLSX={output_xlsx}")
    print(f"OUTCOMES={outcome_manifest}")
    print(f"ROWS={len(all_rows)}")


def main() -> None:
    args = build_parser().parse_args()
    if args.worker:
        run_worker(args)
        return
    if not args.pdf:
        raise SystemExit("--pdf is required unless --worker is used.")
    run_parent(args)


if __name__ == "__main__":
    main()
