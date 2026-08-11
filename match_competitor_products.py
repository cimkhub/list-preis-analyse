#!/usr/bin/env python3
"""Match comparable competitor products into one Excel workbook.

Requirements note:
  pandas, numpy, openpyxl, openai, pydantic, scikit-learn, tqdm, python-dotenv

The script is intentionally standalone. It can run with OpenAI for optional
attribute extraction, OpenRouter for embeddings, Pinecone for vector search,
and DeepSeek for pair judgement. Use --skip-llm for a local rule-based dry run.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import subprocess
import sys
import tempfile
import textwrap
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from tqdm import tqdm

try:
    from dotenv import load_dotenv
except Exception:  # pragma: no cover
    load_dotenv = None

try:
    from openai import OpenAI
except Exception:  # pragma: no cover
    OpenAI = None

try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
except Exception:  # pragma: no cover
    TfidfVectorizer = None
    cosine_similarity = None


INPUT_FILE = "all_suppliers_relevant.csv"
OUTPUT_FILE = "matched_competitor_products.xlsx"
EMBEDDING_DIR = "embeddings/product_matching"
PROJECT_ROOT = Path(__file__).resolve().parent
LOGO_FILE = str(PROJECT_ROOT / "assets" / "list_logo.png")
ATTRIBUTE_MODEL = os.environ.get("DEEPSEEK_ATTRIBUTE_MODEL", os.environ.get("DEEPSEEK_MODEL", "deepseek-v4-flash"))
PAIR_JUDGE_MODEL = os.environ.get("DEEPSEEK_PAIR_JUDGE_MODEL", os.environ.get("DEEPSEEK_MODEL", "deepseek-v4-flash"))
DEEPSEEK_BASE_URL = os.environ.get("DEEPSEEK_BASE_URL", "https://api.deepseek.com")
EMBEDDING_MODEL = "qwen/qwen3-embedding-8b"
OPENROUTER_EMBEDDING_URL = "https://openrouter.ai/api/v1/embeddings"
PINECONE_INDEX_NAME = "product-matching"
PINECONE_CLOUD = "aws"
PINECONE_REGION = "us-east-1"
TOP_K = 25
SIMILARITY_THRESHOLD = 0.0
AUTO_MERGE_EXACT_THRESHOLD = 85
AUTO_MERGE_CLOSE_THRESHOLD = 88
PAIR_JUDGE_WORKERS = 25
ATTRIBUTE_CACHE_VERSION = 3
EMBEDDING_CACHE_VERSION = 3
PAIR_CACHE_VERSION = 5
SECOND_JUDGE_CACHE_VERSION = 2
SAME_SUPPLIER_DEDUPE_SIMILARITY = 0.82

SUPPLIER_ORDER = ["Metro", "Selgros", "Handelshof", "Edeka"]
RELEVANT_VALUES = {"yes", "ja", "true", "1", "relevant", "x"}


def main() -> None:
    if load_dotenv:
        load_dotenv()
    args = build_parser().parse_args()

    input_path = resolve_input_path(Path(args.input))
    output_path = Path(args.output)
    embedding_dir = Path(args.embedding_dir)
    embedding_dir.mkdir(parents=True, exist_ok=True)

    df = load_products(input_path, args.max_products)
    print(f"Loaded rows: {len(pd.read_csv(input_path, dtype=str, keep_default_na=False))}")
    print(f"Relevant rows: {len(df)}")
    if df.empty:
        raise RuntimeError("No relevant rows found.")

    pair_model = args.pair_judge_model or args.deepseek_model or PAIR_JUDGE_MODEL
    pair_client = make_deepseek_client(
        args.skip_llm,
        "pair judgement",
        model=pair_model,
        base_url=args.deepseek_base_url,
    )
    if pair_client:
        print(f"DeepSeek pair judge model: {pair_client['model']}")
    caches = Caches(embedding_dir)

    attributes = build_attributes(df, caches, None, args.force_refresh_attributes, True)
    print(f"Attributes extracted: {len(attributes)}")

    embeddings = build_embeddings(df, attributes, caches, args.force_refresh_embeddings, args.skip_llm)
    print(f"Embeddings loaded or created: {len(embeddings)}")

    df, removed_same_supplier = dedupe_same_supplier_products(
        df,
        attributes,
        embeddings,
        caches,
        pair_client,
        args.force_refresh_pairs,
        args.skip_llm,
        args.pair_workers,
        args.top_k,
    )
    print(f"Same-supplier rows removed before grouping: {removed_same_supplier}")

    pinecone_store = None
    if not args.disable_pinecone and not args.skip_llm:
        pinecone_store = PineconeVectorStore.from_env(args.pinecone_index)
        if pinecone_store:
            pinecone_store.ensure_index(next(iter(embeddings.values())))
            pinecone_store.upsert_products(df, attributes, embeddings)
            print(f"Pinecone vector store: {pinecone_store.index_name}")
    elif args.skip_llm:
        print("Pinecone vector store: skipped because --skip-llm is active")

    cross_supplier_candidates = generate_candidate_pairs(
        df,
        attributes,
        embeddings,
        args.top_k,
        args.similarity_threshold,
        pinecone_store=pinecone_store,
    )
    same_supplier_candidates = generate_same_supplier_duplicate_candidates(
        df,
        attributes,
        embeddings,
        args.top_k,
    )
    candidates = merge_candidate_lists(cross_supplier_candidates, same_supplier_candidates)
    print(
        "Candidate pairs generated: "
        f"{len(candidates)} total "
        f"({len(cross_supplier_candidates)} cross-supplier, "
        f"{len(same_supplier_candidates)} same-supplier)"
    )

    pair_rows, hard_blocked_count = judge_pairs(
        candidates,
        df,
        attributes,
        caches,
        pair_client,
        args.force_refresh_pairs,
        args.skip_llm,
        args.pair_workers,
    )
    print(f"Hard blocked pairs: {hard_blocked_count} (disabled; all vector candidates go to DeepSeek or cache)")
    print(f"LLM judged pairs: {sum(1 for row in pair_rows if not row.get('hard_blocked'))}")

    pair_rows = second_round_judge_pairs(
        pair_rows,
        df,
        attributes,
        caches,
        pair_client,
        args.force_refresh_pairs,
        args.skip_llm,
        args.pair_workers,
    )

    clusters = build_clusters(df, attributes, pair_rows)
    validate_cluster_partition(df, clusters)
    print(f"Clusters created: {len(clusters)}")

    offer_grouping_audit: list[dict[str, Any]] = []
    matched_rows, review_rows = build_output_rows(
        df,
        attributes,
        pair_rows,
        clusters,
        grouping_audit=offer_grouping_audit,
    )
    validate_offer_accounting(df, offer_grouping_audit)
    collapsed_offer_rows = sum(max(len(item.get("merged_product_ids", [])) - 1, 0) for item in offer_grouping_audit)
    print(f"Offer representations collapsed in output: {collapsed_offer_rows}")
    print(f"Unique offers written: {len(df) - collapsed_offer_rows}")
    print(f"Review rows created: {len(review_rows)}")

    attribute_rows = build_attribute_debug_rows(df, attributes)
    write_excel(
        output_path,
        matched_rows,
        review_rows,
        pair_rows,
        attribute_rows,
        Path(args.logo) if args.logo else None,
        offer_grouping_audit=offer_grouping_audit,
    )
    write_offer_grouping_audit(output_path, offer_grouping_audit)
    print(f"Excel written: {output_path}")
    if args.upload_onedrive:
        from upload_to_onedrive import upload_to_onedrive

        upload_to_onedrive(
            file_path=output_path,
            target_url=args.onedrive_url,
            target_folder=args.onedrive_folder,
            tenant=args.ms_tenant_id,
            client_id=args.ms_client_id or os.environ.get("MS_CLIENT_ID", ""),
        )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Match competitor products into canonical comparison rows.")
    parser.add_argument("--input", default=INPUT_FILE)
    parser.add_argument("--output", default=OUTPUT_FILE)
    parser.add_argument("--embedding-dir", default=EMBEDDING_DIR)
    parser.add_argument("--logo", default=LOGO_FILE)
    parser.add_argument("--top-k", type=int, default=TOP_K)
    parser.add_argument("--similarity-threshold", type=float, default=SIMILARITY_THRESHOLD)
    parser.add_argument("--force-refresh-attributes", action="store_true")
    parser.add_argument("--force-refresh-embeddings", action="store_true")
    parser.add_argument("--force-refresh-pairs", action="store_true")
    parser.add_argument("--max-products", type=int)
    parser.add_argument("--skip-llm", action="store_true")
    parser.add_argument("--pair-workers", type=int, default=PAIR_JUDGE_WORKERS)
    parser.add_argument("--deepseek-model", default=os.environ.get("DEEPSEEK_MODEL", PAIR_JUDGE_MODEL))
    parser.add_argument("--attribute-model", default=os.environ.get("DEEPSEEK_ATTRIBUTE_MODEL", ""))
    parser.add_argument("--pair-judge-model", default=os.environ.get("DEEPSEEK_PAIR_JUDGE_MODEL", ""))
    parser.add_argument("--deepseek-base-url", default=os.environ.get("DEEPSEEK_BASE_URL", DEEPSEEK_BASE_URL))
    parser.add_argument("--pinecone-index", default=PINECONE_INDEX_NAME)
    parser.add_argument("--disable-pinecone", action="store_true")
    parser.add_argument("--upload-onedrive", action="store_true")
    parser.add_argument("--onedrive-url", default="https://listgs-my.sharepoint.com/personal/l_kornblum_list-goslar_com/_layouts/15/onedrive.aspx?id=%2Fpersonal%2Fl%5Fkornblum%5Flist%2Dgoslar%5Fcom%2FDocuments%2FAI%2Dbasierte%20Preisgestaltung&ga=1")
    parser.add_argument("--onedrive-folder")
    parser.add_argument("--ms-client-id", default=os.environ.get("MS_CLIENT_ID", ""))
    parser.add_argument("--ms-tenant-id", default=os.environ.get("MS_TENANT_ID", "common"))
    return parser


def resolve_input_path(path: Path) -> Path:
    if path.exists():
        return path
    if path.name == INPUT_FILE:
        candidates = sorted(Path("parsed").glob("KW*_*/all_suppliers_relevant.csv"), key=lambda p: p.stat().st_mtime)
        if candidates:
            fallback = candidates[-1]
            print(f"Input {path} not found; using latest relevant CSV: {fallback}")
            return fallback
    raise FileNotFoundError(f"Input CSV not found: {path}")


def load_products(input_path: Path, max_products: int | None) -> pd.DataFrame:
    df = pd.read_csv(input_path, dtype=str, keep_default_na=False)
    df = ensure_columns(df)
    if "Relevant" in df.columns:
        mask = df["Relevant"].astype(str).str.strip().str.casefold().isin(RELEVANT_VALUES)
        df = df.loc[mask].copy()
    if "Relevant Time" in df.columns:
        mask = df["Relevant Time"].astype(str).str.strip().str.casefold().isin(RELEVANT_VALUES)
        df = df.loc[mask].copy()
    df = df.reset_index(drop=False).rename(columns={"index": "source_row_index"})
    if max_products:
        df = df.head(max_products).copy()
    df["supplier_norm"] = df["supplier"].map(normalize_supplier)
    df["product_id"] = [
        stable_product_id(row, idx)
        for idx, row in df.iterrows()
    ]
    df["rich_product_text"] = df.apply(rich_product_text, axis=1)
    return df


def ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "supplier", "location", "product_name", "description", "category", "origin",
        "brand", "product", "brand_product_confidence",
        "unit", "quantity", "price", "price_per_kg", "price_is_net", "price_gross",
        "price_tiers", "valid_from", "valid_to", "calendar_week", "year",
        "source_file", "source_title", "source_tab", "source_page",
        "extraction_confidence", "Relevant", "Relevant Time",
    ]
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    return df


def get(row: pd.Series | dict[str, Any], column: str, default: str = "") -> str:
    value = row.get(column, default)
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return default
    return str(value)


def normalize_supplier(value: str) -> str:
    text = str(value).strip().casefold()
    if "metro" in text:
        return "Metro"
    if "selgros" in text:
        return "Selgros"
    if "handelshof" in text:
        return "Handelshof"
    if "edeka" in text:
        return "Edeka"
    return "Unknown"


def stable_product_id(row: pd.Series, idx: int) -> str:
    parts = [
        get(row, "supplier"), get(row, "product_name"), get(row, "description"),
        get(row, "price"), get(row, "quantity"), get(row, "source_file"),
        get(row, "source_page"), str(idx),
    ]
    digest = hashlib.sha1("||".join(parts).encode("utf-8")).hexdigest()[:16]
    return f"p_{digest}"


def rich_product_text(row: pd.Series) -> str:
    fields = [
        ("Supplier", "supplier_norm"),
        ("Category", "category"),
        ("Product name", "product_name"),
        ("Brand", "brand"),
        ("Product", "product"),
        ("Description", "description"),
        ("Origin", "origin"),
        ("Unit", "unit"),
        ("Quantity", "quantity"),
        ("Price", "price"),
        ("Price per kg", "price_per_kg"),
        ("Valid from", "valid_from"),
        ("Valid to", "valid_to"),
        ("Source", "source_file"),
    ]
    return "\n".join(f"{label}: {get(row, col)}" for label, col in fields if get(row, col))


class Caches:
    def __init__(self, root: Path):
        self.root = root
        self.attributes_path = root / "attributes.jsonl"
        self.embeddings_path = root / "embeddings.jsonl"
        self.pairs_path = root / "pair_judgements.jsonl"
        self.second_pairs_path = root / "second_pair_judgements.jsonl"
        for path in [self.attributes_path, self.embeddings_path, self.pairs_path, self.second_pairs_path]:
            path.touch(exist_ok=True)
        self.attributes = read_jsonl_by_key(self.attributes_path, "product_id")
        self.embeddings = read_jsonl_by_key(self.embeddings_path, "product_id")
        self.pairs = read_jsonl_by_key(self.pairs_path, "pair_key")
        self.second_pairs = read_jsonl_by_key(self.second_pairs_path, "pair_key")

    def append_attribute(self, item: dict[str, Any]) -> None:
        append_jsonl(self.attributes_path, item)
        self.attributes[item["product_id"]] = item

    def append_embedding(self, item: dict[str, Any]) -> None:
        append_jsonl(self.embeddings_path, item)
        self.embeddings[item["product_id"]] = item

    def append_pair(self, item: dict[str, Any]) -> None:
        append_jsonl(self.pairs_path, item)
        self.pairs[item["pair_key"]] = item

    def append_second_pair(self, item: dict[str, Any]) -> None:
        append_jsonl(self.second_pairs_path, item)
        self.second_pairs[item["pair_key"]] = item


def read_jsonl_by_key(path: Path, key: str) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    out: dict[str, dict[str, Any]] = {}
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                continue
            if item.get(key):
                out[str(item[key])] = item
    return out


def append_jsonl(path: Path, item: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(item, ensure_ascii=False) + "\n")


def make_deepseek_client(
    skip_llm: bool,
    purpose: str,
    model: str | None = None,
    base_url: str | None = None,
) -> dict[str, str] | None:
    if skip_llm:
        return None
    api_key = get_env_any("DEEPSEEK_API_KEY", "deepseek_api_key", "Deepseek_api_key")
    if not api_key:
        print(f"DEEPSEEK_API_KEY missing; {purpose} will use rule-based fallback.")
        return None
    return {
        "api_key": api_key,
        "base_url": base_url or get_env_any("DEEPSEEK_BASE_URL", "deepseek_base_url") or DEEPSEEK_BASE_URL,
        "model": model or (ATTRIBUTE_MODEL if purpose == "attribute extraction" else PAIR_JUDGE_MODEL),
    }


def build_attributes(df: pd.DataFrame, caches: Caches, client, force: bool, skip_llm: bool) -> dict[str, dict[str, Any]]:
    attributes: dict[str, dict[str, Any]] = {}
    for _, row in tqdm(df.iterrows(), total=len(df), desc="attributes"):
        product_id = get(row, "product_id")
        cached = caches.attributes.get(product_id)
        if cached and not force and int(cached.get("schema_version") or 0) == ATTRIBUTE_CACHE_VERSION:
            attributes[product_id] = cached
            continue
        if client and not skip_llm:
            item = extract_attributes_llm(client, product_id, get(row, "rich_product_text"))
        else:
            item = fallback_attributes(row)
        caches.append_attribute(item)
        attributes[product_id] = item
    return attributes


def extract_attributes_llm(client, product_id: str, text: str) -> dict[str, Any]:
    prompt = ATTRIBUTE_PROMPT.format(rich_product_text=text)
    fallback = fallback_attributes({"product_id": product_id, "rich_product_text": text, "product_name": text})
    for attempt in range(2):
        try:
            data = call_deepseek_json(
                client,
                "You extract structured product attributes. Return only strict JSON.",
                prompt if attempt == 0 else prompt + "\nReturn only repaired strict JSON.",
                max_tokens=900,
            )
            return normalize_attribute_item(product_id, data)
        except Exception as exc:
            if attempt == 1:
                fallback["notes"] = f"LLM attribute fallback after error: {exc}"
    return fallback


ATTRIBUTE_PROMPT = """You are extracting structured product attributes for food wholesale competitor matching.

The goal is to decide later whether products from different suppliers are commercially comparable and can be placed in one comparison row.

Extract only attributes supported by the product text.
Use English normalized attribute values.
Do not guess beyond the evidence.
For German product text, translate normalized values into English.

Examples:
Spargel Weiß means white asparagus.
Spargel Grün means green asparagus.
geschält means peeled.
Klasse I means class I.
Karton, Kiste, Box are packaging types.
Schale means tray.
TK means frozen.
If the product is a sauce, soup, seasoning, potato, side product, or accessory related to the main ingredient, mark is_accessory_or_related_product true when relevant.

The field do_not_merge_with is very important.
Include variants or related products that must not be merged with this product.

Prefer conservative extraction.

Return strict JSON with this structure:
{{
  "product_id": "...",
  "base_product": "...",
  "product_family": "...",
  "variant": null,
  "processing": null,
  "brand": null,
  "origin": null,
  "quality_class": null,
  "calibre": null,
  "packaging": null,
  "quantity_value": null,
  "quantity_unit": "unknown",
  "normalized_quantity_kg": null,
  "normalized_quantity_liter": null,
  "unit_basis": "unknown",
  "fresh_or_frozen": "unknown",
  "is_accessory_or_related_product": false,
  "commercially_relevant_attributes": [],
  "do_not_merge_with": [],
  "attribute_confidence": 0,
  "notes": ""
}}

Input product:
{rich_product_text}
"""


def fallback_attributes(row: pd.Series | dict[str, Any]) -> dict[str, Any]:
    product_id = get(row, "product_id")
    name = get(row, "product") or get(row, "product_name") or get(row, "rich_product_text")
    desc = get(row, "description")
    category = get(row, "category")
    text = f"{name} {desc}".casefold()
    base = infer_base_product(name, desc, category)
    processing = infer_processing(text)
    variant = infer_variant(text)
    packaging = infer_packaging(text, get(row, "unit"))
    quantity_value = to_float(get(row, "quantity"))
    unit = get(row, "unit") or "unknown"
    normalized_kg = None
    normalized_liter = None
    if quantity_value is not None:
        unit_cf = unit.casefold()
        if unit_cf in {"kg", "kilogramm"}:
            normalized_kg = quantity_value
        elif unit_cf in {"g", "gramm"}:
            normalized_kg = quantity_value / 1000
        elif unit_cf in {"l", "liter"}:
            normalized_liter = quantity_value
        elif unit_cf in {"ml"}:
            normalized_liter = quantity_value / 1000
    brand = get(row, "brand") or infer_brand(get(row, "product_name") or name)
    attr_list = [x for x in [base, variant, processing, get(row, "origin"), packaging] if x]
    return normalize_attribute_item(product_id, {
        "product_id": product_id,
        "schema_version": ATTRIBUTE_CACHE_VERSION,
        "base_product": base,
        "product_family": category or "unknown",
        "variant": variant,
        "processing": processing,
        "brand": brand,
        "origin": get(row, "origin") or None,
        "quality_class": infer_quality_class(text),
        "calibre": infer_calibre(text),
        "packaging": packaging,
        "quantity_value": quantity_value,
        "quantity_unit": unit or "unknown",
        "normalized_quantity_kg": normalized_kg,
        "normalized_quantity_liter": normalized_liter,
        "unit_basis": infer_unit_basis(unit),
        "fresh_or_frozen": "frozen" if any(x in text for x in ["tk", "tiefkühl", "tiefkuehl", "gefroren", "frozen"]) else "fresh",
        "is_accessory_or_related_product": is_accessory(text),
        "commercially_relevant_attributes": attr_list,
        "do_not_merge_with": infer_do_not_merge(text),
        "attribute_confidence": 65,
        "notes": "rule-based fallback",
    })


def normalize_attribute_item(product_id: str, data: dict[str, Any]) -> dict[str, Any]:
    defaults = {
        "product_id": product_id,
        "schema_version": ATTRIBUTE_CACHE_VERSION,
        "base_product": "",
        "product_family": "unknown",
        "variant": None,
        "processing": None,
        "brand": None,
        "origin": None,
        "quality_class": None,
        "calibre": None,
        "packaging": None,
        "quantity_value": None,
        "quantity_unit": "unknown",
        "normalized_quantity_kg": None,
        "normalized_quantity_liter": None,
        "unit_basis": "unknown",
        "fresh_or_frozen": "unknown",
        "is_accessory_or_related_product": False,
        "commercially_relevant_attributes": [],
        "do_not_merge_with": [],
        "attribute_confidence": 0,
        "notes": "",
    }
    item = {**defaults, **(data or {})}
    item["product_id"] = product_id
    item["schema_version"] = ATTRIBUTE_CACHE_VERSION
    for key in ["commercially_relevant_attributes", "do_not_merge_with"]:
        if not isinstance(item.get(key), list):
            item[key] = []
    item["attribute_confidence"] = normalize_confidence(item.get("attribute_confidence"))
    return item


def infer_base_product(name: str, desc: str, category: str) -> str:
    text = f"{name} {desc}".casefold()
    meat_cut = infer_meat_cut_base(text)
    if meat_cut:
        return meat_cut
    patterns = [
        ("white asparagus", ["spargel weiß", "weisser spargel", "weißer spargel", "white asparagus"]),
        ("green asparagus", ["spargel grün", "gruener spargel", "grüner spargel", "green asparagus"]),
        ("asparagus", ["spargel"]),
        ("potato", ["kartoffel", "pommes", "fries"]),
        ("cream", ["sahne"]),
        ("quark", ["quark"]),
        ("milk", ["milch"]),
        ("cheese", ["käse", "kaese", "cheese"]),
        ("sausage", ["wurst", "würst", "salami"]),
        ("beef", ["rind", "beef", "entrecôte", "roastbeef"]),
        ("pork", ["schwein", "pork"]),
        ("chicken", ["hähnchen", "haehnchen", "chicken"]),
        ("fish", ["fisch", "lachs", "scholle", "garnelen", "thunfisch"]),
        ("oil", ["öl", "oel", "olio"]),
        ("ice cream", ["eis", "ice cream"]),
        ("vegetable mix", ["gemüse", "gemuese", "brunoise"]),
    ]
    for base, terms in patterns:
        if any(term in text for term in terms):
            return base
    words = re.findall(r"[A-Za-zÄÖÜäöüßéèôû]+", name)
    return " ".join(words[:3]).casefold() if words else category or "unknown"


def infer_meat_cut_base(text: str) -> str | None:
    pork_cuts = [
        ("pork neck steak", ["schweinenackensteak", "schweinenacken-steak", "nackensteak"]),
        ("pork back steak", ["schweinerücken-steak", "schweineruecken-steak", "schweinerücken steaks", "schweineruecken steaks"]),
        ("pork neck", ["schweinenacken", "schweine nacken", "pork neck"]),
        ("pork tenderloin", ["schweinefilet", "pork tenderloin", "filet vom schwein"]),
        ("pork shoulder", ["schweine schulter", "schweineschulter", "schulter schier"]),
        ("pork loin", ["schweinelachs", "schweinelachse", "pork loin"]),
        ("pork back", ["schweinerücken", "schweineruecken"]),
        ("pork strips", ["schweinegeschnetzeltes"]),
        ("pork french rack", ["french rack"]),
        ("pork knuckle", ["schweineschäufele", "schweineschaeufele"]),
        ("pork belly", ["schweinebauch", "pork belly"]),
        ("pork chop rack", ["schweinekaree", "karee"]),
        ("pork topside", ["schweineoberschale", "oberschale"]),
        ("pork silverside", ["schweineunterschale", "unterschale"]),
        ("pork nut", ["schweinenuss", "nuss"]),
        ("pork schnitzel", ["schweineschnitzel", "schnitzel"]),
    ]
    beef_cuts = [
        ("beef roastbeef", ["rinder-roastbeef", "rinder roastbeef", "roastbeef"]),
        ("beef entrecote", ["entrecôte", "entrecote"]),
        ("beef rump", ["rinderhüfte", "rinderhuefte", "hüftsteak", "hueftsteak"]),
        ("beef topside", ["rinderoberschale", "oberschale vom rind"]),
        ("beef knuckle", ["rinderkugel", "rinder kugel"]),
    ]
    poultry_cuts = [
        ("chicken breast", ["hähnchenbrust", "haehnchenbrust", "chicken breast"]),
        ("chicken thigh", ["hähnchenkeule", "haehnchenkeule", "oberkeule", "chicken thigh"]),
        ("chicken wing", ["hähnchenflügel", "haehnchenfluegel", "chicken wing"]),
    ]
    fish_cuts = [
        ("salmon fillet", ["lachsfilet", "salmon fillet"]),
        ("trout fillet", ["forellenfilet", "trout fillet"]),
        ("tuna steak", ["thunfischsteak", "tuna steak"]),
        ("plaice", ["scholle"]),
        ("shrimp", ["garnelen", "shrimp"]),
    ]
    for base, terms in pork_cuts + beef_cuts + poultry_cuts + fish_cuts:
        if any(term in text for term in terms):
            return base
    return None


def broad_base_product(base: Any) -> bool:
    return clean(base) in {"pork", "beef", "chicken", "fish", "meat", "seafood"}


def meat_family(base: Any) -> str | None:
    text = clean(base)
    for family in ["pork", "beef", "chicken"]:
        if text == family or text.startswith(f"{family} "):
            return family
    if text in {"fish", "seafood"} or any(term in text for term in ["salmon", "trout", "tuna", "plaice", "shrimp"]):
        return "fish"
    return None


def infer_variant(text: str) -> str | None:
    if "duroc" in text:
        return "Duroc"
    if "duke of berkshire" in text or "berkshire" in text:
        return "Berkshire"
    if "weiß" in text or "weiss" in text or "white" in text:
        return "white"
    if "grün" in text or "gruen" in text or "green" in text:
        return "green"
    return None


def infer_processing(text: str) -> str | None:
    if "geschält" in text or "geschaelt" in text or "peeled" in text:
        return "peeled"
    if "ungeschält" in text or "ungeschaelt" in text or "unpeeled" in text:
        return "unpeeled"
    if "geschnitten" in text:
        return "cut"
    if "geräuchert" in text or "geraeuchert" in text:
        return "smoked"
    return None


def infer_packaging(text: str, unit: str) -> str | None:
    for term, normalized in [
        ("karton", "box"), ("kiste", "box"), ("box", "box"), ("schale", "tray"),
        ("beutel", "bag"), ("eimer", "bucket"), ("dose", "can"), ("flasche", "bottle"),
        ("packung", "pack"),
    ]:
        if term in text or term in str(unit).casefold():
            return normalized
    return None


def infer_quality_class(text: str) -> str | None:
    if re.search(r"klasse\s*i\b|class\s*i\b", text):
        return "class I"
    if re.search(r"klasse\s*ii\b|class\s*ii\b", text):
        return "class II"
    return None


def infer_calibre(text: str) -> str | None:
    match = re.search(r"(\d{1,2})\s*(?:mm)?\s*(?:plus|\+)", text)
    return f"{match.group(1)} mm plus" if match else None


def infer_brand(name: str) -> str | None:
    text = name.casefold()
    if "duroc" in text:
        return "Duroc"
    if "duke of berkshire" in text or "berkshire" in text:
        return "The Duke Of Berkshire"
    words = name.split()
    if not words:
        return None
    first = words[0]
    if first.isupper() and len(first) > 2:
        return first.title()
    return None


def infer_unit_basis(unit: str) -> str:
    unit_cf = str(unit).casefold()
    if unit_cf in {"kg", "g", "kilogramm", "gramm"}:
        return "kg"
    if unit_cf in {"l", "liter", "ml"}:
        return "liter"
    if unit_cf:
        return unit_cf
    return "unknown"


def is_accessory(text: str) -> bool:
    return any(term in text for term in ["sauce", "soße", "suppe", "bouillon", "gewürz", "seasoning", "kartoffel", "potato"])


def infer_do_not_merge(text: str) -> list[str]:
    items = []
    if "spargel" in text:
        items.extend(["asparagus sauce", "asparagus soup", "potatoes with asparagus"])
        if "weiß" in text or "weiss" in text:
            items.append("green asparagus")
        if "grün" in text or "gruen" in text:
            items.append("white asparagus")
    return items


def to_float(value: Any) -> float | None:
    try:
        text = str(value).strip().replace(",", ".")
        return float(text) if text else None
    except Exception:
        return None


def normalize_confidence(value: Any) -> int:
    score = to_float(value)
    if score is None:
        return 0
    if 0 < score <= 1:
        score *= 100
    return max(0, min(100, int(round(score))))


def safe_json_loads(text: str) -> dict[str, Any]:
    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1].rsplit("```", 1)[0]
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        text = text[start:end + 1]
    return json.loads(text)


def build_embeddings(df: pd.DataFrame, attributes: dict[str, dict[str, Any]], caches: Caches, force: bool, skip_llm: bool) -> dict[str, np.ndarray]:
    texts = [embedding_text(row, attributes[get(row, "product_id")]) for _, row in df.iterrows()]
    ids = [get(row, "product_id") for _, row in df.iterrows()]
    text_hashes = [short_hash(text) for text in texts]
    embeddings: dict[str, np.ndarray] = {}

    openrouter_key = get_env_any("OPENROUTER_API_KEY", "openrouter_api_key", "Openrouter_api_key")
    if openrouter_key and not skip_llm:
        for product_id, text, text_hash in tqdm(list(zip(ids, texts, text_hashes)), desc="embeddings"):
            cached = caches.embeddings.get(product_id)
            if (
                cached and not force
                and cached.get("model") == EMBEDDING_MODEL
                and int(cached.get("schema_version") or 0) == EMBEDDING_CACHE_VERSION
                and cached.get("text_hash") == text_hash
            ):
                embeddings[product_id] = np.array(cached["embedding"], dtype=float)
                continue
            vector = embed_openrouter(openrouter_key, text)
            caches.append_embedding({
                "product_id": product_id,
                "schema_version": EMBEDDING_CACHE_VERSION,
                "text_hash": text_hash,
                "model": EMBEDDING_MODEL,
                "provider": "openrouter",
                "embedding": vector,
            })
            embeddings[product_id] = np.array(vector, dtype=float)
        return embeddings
    elif not skip_llm:
        print("OpenRouter key missing; using local TF-IDF embeddings. Set openrouter_api_key or OPENROUTER_API_KEY for Qwen embeddings.")

    cached_count = 0
    for product_id, text_hash in zip(ids, text_hashes):
        cached = caches.embeddings.get(product_id)
        if (
            cached and not force
            and cached.get("embedding")
            and int(cached.get("schema_version") or 0) == EMBEDDING_CACHE_VERSION
            and cached.get("text_hash") == text_hash
        ):
            embeddings[product_id] = np.array(cached["embedding"], dtype=float)
            cached_count += 1
    if cached_count == len(ids):
        return embeddings

    matrix = build_tfidf_matrix(texts)
    for product_id, vector in zip(ids, matrix):
        embeddings[product_id] = vector
    return embeddings


def embedding_text(row: pd.Series, attr: dict[str, Any]) -> str:
    return "\n".join([
        f"base_product: {attr.get('base_product')}",
        f"product_family: {attr.get('product_family')}",
        f"variant: {attr.get('variant')}",
        f"processing: {attr.get('processing')}",
        f"brand: {attr.get('brand')}",
        f"origin: {attr.get('origin') or get(row, 'origin')}",
        f"quality_class: {attr.get('quality_class')}",
        f"calibre: {attr.get('calibre')}",
        f"packaging: {attr.get('packaging')}",
        f"quantity: {attr.get('quantity_value')} {attr.get('quantity_unit')}",
        f"unit_basis: {attr.get('unit_basis')}",
        f"fresh_or_frozen: {attr.get('fresh_or_frozen')}",
        f"commercially_relevant_attributes: {', '.join(map(str, attr.get('commercially_relevant_attributes', [])))}",
        f"split_brand: {get(row, 'brand')}",
        f"split_product: {get(row, 'product')}",
        f"original_product_name: {get(row, 'product_name')}",
        f"original_description: {get(row, 'description')}",
    ])


def embed_openrouter(api_key: str, text: str) -> list[float]:
    for attempt in range(4):
        try:
            response = requests.post(
                OPENROUTER_EMBEDDING_URL,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                    "HTTP-Referer": "https://local.birkenhof",
                    "X-Title": "Birkenhof Product Matching",
                },
                json={"model": EMBEDDING_MODEL, "input": text},
                timeout=120,
            )
            response.raise_for_status()
            data = response.json()
            return data["data"][0]["embedding"]
        except Exception as exc:
            if attempt == 3:
                raise RuntimeError(f"OpenRouter embedding failed after retries: {exc}") from exc
            time.sleep(2 ** attempt)
    raise RuntimeError("unreachable")


def get_env_any(*names: str) -> str:
    for name in names:
        value = os.environ.get(name)
        if value:
            return value
    return ""


def short_hash(value: str) -> str:
    return hashlib.sha1(value.encode("utf-8")).hexdigest()[:16]


def build_tfidf_matrix(texts: list[str]) -> list[np.ndarray]:
    if TfidfVectorizer is not None:
        matrix = TfidfVectorizer(min_df=1, ngram_range=(1, 2)).fit_transform(texts).toarray()
        return [np.asarray(row, dtype=float) for row in matrix]
    vocab = sorted({token for text in texts for token in re.findall(r"\w+", text.casefold())})
    index = {token: i for i, token in enumerate(vocab)}
    rows = []
    for text in texts:
        vector = np.zeros(len(vocab), dtype=float)
        counts = Counter(re.findall(r"\w+", text.casefold()))
        for token, count in counts.items():
            vector[index[token]] = count
        norm = np.linalg.norm(vector)
        rows.append(vector / norm if norm else vector)
    return rows


class PineconeVectorStore:
    def __init__(self, api_key: str, index_name: str):
        self.api_key = api_key
        self.index_name = index_name
        self.host: str | None = None

    @classmethod
    def from_env(cls, index_name: str) -> "PineconeVectorStore | None":
        api_key = get_env_any("PINECONE_API_KEY", "Pinecone_api_key", "pinecone_api_key")
        if not api_key:
            print("Pinecone key missing; using local vector candidate retrieval.")
            return None
        return cls(api_key, index_name)

    @property
    def headers(self) -> dict[str, str]:
        return {"Api-Key": self.api_key, "Content-Type": "application/json"}

    def ensure_index(self, sample_vector: np.ndarray) -> None:
        dimension = int(len(sample_vector))
        existing = self.describe_index()
        if not existing:
            print(f"Creating Pinecone index {self.index_name} with dimension {dimension}")
            response = requests.post(
                "https://api.pinecone.io/indexes",
                headers=self.headers,
                json={
                    "name": self.index_name,
                    "dimension": dimension,
                    "metric": "cosine",
                    "spec": {
                        "serverless": {
                            "cloud": PINECONE_CLOUD,
                            "region": PINECONE_REGION,
                        }
                    },
                },
                timeout=120,
            )
            if response.status_code not in {200, 201, 202, 409}:
                raise RuntimeError(f"Pinecone index create failed: {response.status_code} {response.text}")
            for _ in range(30):
                time.sleep(2)
                existing = self.describe_index()
                if existing:
                    break
        if not existing:
            raise RuntimeError("Pinecone index not available after create.")
        self.host = existing.get("host")
        if not self.host:
            raise RuntimeError(f"Pinecone index {self.index_name} has no host in describe response.")

    def describe_index(self) -> dict[str, Any] | None:
        response = requests.get(
            f"https://api.pinecone.io/indexes/{self.index_name}",
            headers=self.headers,
            timeout=60,
        )
        if response.status_code == 404:
            return None
        if response.status_code >= 400:
            raise RuntimeError(f"Pinecone describe failed: {response.status_code} {response.text}")
        return response.json()

    def upsert_products(self, df: pd.DataFrame, attributes: dict[str, dict[str, Any]], embeddings: dict[str, np.ndarray]) -> None:
        if not self.host:
            return
        rows = df.to_dict("records")
        vectors = []
        for row in rows:
            product_id = row["product_id"]
            attr = attributes[product_id]
            vectors.append({
                "id": product_id,
                "values": embeddings[product_id].astype(float).tolist(),
                "metadata": {
                    "supplier_norm": row.get("supplier_norm", ""),
                    "product_name": row.get("product_name", ""),
                    "brand": row.get("brand", ""),
                    "product": row.get("product", ""),
                    "base_product": str(attr.get("base_product") or ""),
                    "variant": str(attr.get("variant") or ""),
                    "processing": str(attr.get("processing") or ""),
                    "category": row.get("category", ""),
                },
            })
        for start in tqdm(range(0, len(vectors), 100), desc="pinecone upsert"):
            batch = vectors[start:start + 100]
            response = requests.post(
                f"https://{self.host}/vectors/upsert",
                headers=self.headers,
                json={"vectors": batch, "namespace": "products"},
                timeout=120,
            )
            if response.status_code >= 400:
                raise RuntimeError(f"Pinecone upsert failed: {response.status_code} {response.text}")

    def query(self, product_id: str, vector: np.ndarray, top_k: int) -> list[dict[str, Any]]:
        if not self.host:
            return []
        response = requests.post(
            f"https://{self.host}/query",
            headers=self.headers,
            json={
                "vector": vector.astype(float).tolist(),
                "topK": top_k,
                "includeMetadata": True,
                "namespace": "products",
            },
            timeout=120,
        )
        if response.status_code >= 400:
            raise RuntimeError(f"Pinecone query failed: {response.status_code} {response.text}")
        return response.json().get("matches", [])


def generate_candidate_pairs(
    df: pd.DataFrame,
    attributes: dict[str, dict[str, Any]],
    embeddings: dict[str, np.ndarray],
    top_k: int,
    threshold: float,
    pinecone_store: PineconeVectorStore | None = None,
) -> list[dict[str, Any]]:
    rows = df.to_dict("records")
    ids = [row["product_id"] for row in rows]
    index_by_id = {row["product_id"]: idx for idx, row in enumerate(rows)}
    candidate_map: dict[tuple[str, str], dict[str, Any]] = {}
    for i, row_a in enumerate(rows):
        scored = pinecone_scores(row_a, embeddings[ids[i]], rows, index_by_id, pinecone_store, top_k)
        if not scored:
            scored = local_scores(i, row_a, rows, ids, attributes, embeddings, threshold)
        for sim, j, source in sorted(scored, reverse=True)[:top_k]:
            row_b = rows[j]
            key = tuple(sorted([row_a["product_id"], row_b["product_id"]]))
            if key not in candidate_map or sim > candidate_map[key]["similarity"]:
                attr_a = attributes[row_a["product_id"]]
                attr_b = attributes[row_b["product_id"]]
                candidate_map[key] = {
                    "product_id_a": row_a["product_id"],
                    "product_id_b": row_b["product_id"],
                    "supplier_a": row_a["supplier_norm"],
                    "supplier_b": row_b["supplier_norm"],
                    "name_a": display_product_name(row_a),
                    "name_b": display_product_name(row_b),
                    "base_product_a": attr_a.get("base_product"),
                    "base_product_b": attr_b.get("base_product"),
                    "variant_a": attr_a.get("variant"),
                    "variant_b": attr_b.get("variant"),
                    "processing_a": attr_a.get("processing"),
                    "processing_b": attr_b.get("processing"),
                    "similarity": round(float(sim), 4),
                    "candidate_source": source,
                }
    return list(candidate_map.values())


def pinecone_scores(
    row_a: dict[str, Any],
    vector: np.ndarray,
    rows: list[dict[str, Any]],
    index_by_id: dict[str, int],
    pinecone_store: PineconeVectorStore | None,
    top_k: int,
) -> list[tuple[float, int, str]]:
    if not pinecone_store:
        return []
    try:
        matches = pinecone_store.query(row_a["product_id"], vector, top_k * 5)
    except Exception as exc:
        print(f"Pinecone query failed, falling back to local vector search: {exc}")
        return []
    scored = []
    for match in matches:
        product_id = match.get("id")
        if product_id == row_a["product_id"] or product_id not in index_by_id:
            continue
        j = index_by_id[product_id]
        row_b = rows[j]
        if row_a["supplier_norm"] == row_b["supplier_norm"]:
            continue
        scored.append((float(match.get("score") or 0), j, "pinecone_similarity"))
    return scored


def local_scores(
    i: int,
    row_a: dict[str, Any],
    rows: list[dict[str, Any]],
    ids: list[str],
    attributes: dict[str, dict[str, Any]],
    embeddings: dict[str, np.ndarray],
    threshold: float,
) -> list[tuple[float, int, str]]:
    scored = []
    for j, row_b in enumerate(rows):
        if i == j or row_a["supplier_norm"] == row_b["supplier_norm"]:
            continue
        sim = vector_similarity(embeddings[ids[i]], embeddings[ids[j]])
        if sim >= threshold:
            scored.append((sim, j, "vector_similarity"))
    return scored


def dedupe_same_supplier_products(
    df: pd.DataFrame,
    attributes: dict[str, dict[str, Any]],
    embeddings: dict[str, np.ndarray],
    caches: Caches,
    pair_client,
    force: bool,
    skip_llm: bool,
    pair_workers: int,
    top_k: int,
) -> tuple[pd.DataFrame, int]:
    # Nothing is deleted before variant grouping. Repeated extraction rows are
    # coalesced losslessly at offer level after cluster assignment, where all
    # retained fields and source IDs can be audited.
    print(
        "Same-supplier pre-clustering deletion disabled; "
        "offer representations are coalesced after variant grouping"
    )
    return df.copy().reset_index(drop=True), 0


def dedupe_exact_same_supplier_offers(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    groups: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in df.to_dict("records"):
        key = exact_same_supplier_offer_key(row)
        if key:
            groups[key].append(row)

    remove_ids: set[str] = set()
    keep_ids: set[str] = set()
    row_by_id = {row["product_id"]: row for row in df.to_dict("records")}
    for group in groups.values():
        if len(group) < 2:
            continue
        keeper = choose_longest_valid_offer(group)
        keep_ids.add(keeper["product_id"])
        remove_ids.update(row["product_id"] for row in group if row["product_id"] != keeper["product_id"])

    if not remove_ids:
        return df, 0

    print_same_supplier_dedupe_summary(row_by_id, keep_ids, remove_ids)
    kept = df.loc[~df["product_id"].isin(remove_ids)].copy().reset_index(drop=True)
    return kept, len(remove_ids)


def exact_same_supplier_offer_key(row: dict[str, Any]) -> tuple[str, ...] | None:
    product = get(row, "product") or get(row, "product_name")
    supplier = normalize_supplier(get(row, "supplier_norm") or get(row, "supplier"))
    price = duplicate_key_number(get(row, "price"))
    quantity = duplicate_key_number(get(row, "quantity"))
    unit = duplicate_key_text(get(row, "unit"))
    valid_from = str(get(row, "valid_from")).strip()
    valid_to = str(get(row, "valid_to")).strip()
    # Missing commercial identity fields are not proof of equality. Keeping a
    # possible duplicate is safer than deleting a real offer.
    if not all([supplier, product.strip(), price, quantity, unit, valid_from, valid_to]):
        return None
    return (
        supplier,
        duplicate_key_text(get(row, "category")),
        duplicate_key_text(get(row, "brand")),
        duplicate_key_text(product),
        duplicate_key_text(get(row, "description")),
        duplicate_key_text(get(row, "origin")),
        price,
        quantity,
        unit,
        duplicate_key_number(get(row, "price_per_kg")),
        duplicate_key_price_tiers(get(row, "price_tiers")),
        valid_from,
        valid_to,
    )


def duplicate_key_text(value: Any) -> str:
    text = str(value or "").casefold()
    return re.sub(r"[^a-z0-9äöüß]+", " ", text).strip()


def duplicate_key_number(value: Any) -> str:
    number = to_float(value)
    return f"{number:.4f}" if number is not None else ""


def duplicate_key_price_tiers(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        tiers = json.loads(text)
    except Exception:
        return duplicate_key_text(text)
    if not isinstance(tiers, list):
        return duplicate_key_text(text)
    normalized = []
    for tier in tiers:
        if not isinstance(tier, dict):
            continue
        normalized.append({
            "min_qty": duplicate_key_number(tier.get("min_qty")),
            "price": duplicate_key_number(tier.get("price")),
        })
    return json.dumps(sorted(normalized, key=lambda item: (item["min_qty"], item["price"])), sort_keys=True)


def generate_same_supplier_duplicate_candidates(
    df: pd.DataFrame,
    attributes: dict[str, dict[str, Any]],
    embeddings: dict[str, np.ndarray],
    top_k: int,
) -> list[dict[str, Any]]:
    rows = df.to_dict("records")
    candidate_map: dict[tuple[str, str], dict[str, Any]] = {}
    by_supplier: dict[str, list[int]] = defaultdict(list)
    for idx, row in enumerate(rows):
        by_supplier[row["supplier_norm"]].append(idx)

    for supplier, indexes in by_supplier.items():
        # Exact product names are always candidates. Description, price,
        # quantity, calibre and validity belong to a variant/offer and must not
        # prevent two rows for the same supplier from being compared.
        for position, i in enumerate(indexes):
            row_a = rows[i]
            name_a = exact_product_name_key(row_a)
            if not name_a:
                continue
            for j in indexes[position + 1:]:
                row_b = rows[j]
                if name_a != exact_product_name_key(row_b):
                    continue
                sim = vector_similarity(
                    embeddings[row_a["product_id"]],
                    embeddings[row_b["product_id"]],
                )
                add_matching_candidate(
                    candidate_map,
                    row_a,
                    row_b,
                    attributes,
                    sim,
                    "same_supplier_exact_product_name",
                )

        # Compute all same-supplier similarities in one BLAS-backed matrix
        # operation. This avoids millions of Python-level vector dot products
        # on a normal weekly run.
        supplier_matrix = np.vstack([embeddings[rows[index]["product_id"]] for index in indexes]).astype(float)
        norms = np.linalg.norm(supplier_matrix, axis=1, keepdims=True)
        normalized_matrix = np.divide(
            supplier_matrix,
            norms,
            out=np.zeros_like(supplier_matrix),
            where=norms != 0,
        )
        similarity_matrix = normalized_matrix @ normalized_matrix.T
        for local_i, i in enumerate(indexes):
            row_a = rows[i]
            scored: list[tuple[float, int, str]] = []
            for local_j in np.argsort(similarity_matrix[local_i])[::-1]:
                if local_i == int(local_j):
                    continue
                sim = float(similarity_matrix[local_i, local_j])
                if sim < SAME_SUPPLIER_DEDUPE_SIMILARITY:
                    break
                scored.append((sim, indexes[int(local_j)], "same_supplier_vector_similarity"))
                if len(scored) >= top_k:
                    break

            for sim, j, source in scored:
                row_b = rows[j]
                add_matching_candidate(
                    candidate_map,
                    row_a,
                    row_b,
                    attributes,
                    sim,
                    source,
                )
    return list(candidate_map.values())


def same_source_offer(row_a: dict[str, Any], row_b: dict[str, Any]) -> bool:
    return (
        get(row_a, "source_file") == get(row_b, "source_file")
        and get(row_a, "source_page") == get(row_b, "source_page")
    )


def exact_product_name_key(row: dict[str, Any]) -> str:
    return duplicate_key_text(get(row, "product") or get(row, "product_name"))


def add_matching_candidate(
    candidate_map: dict[tuple[str, str], dict[str, Any]],
    row_a: dict[str, Any],
    row_b: dict[str, Any],
    attributes: dict[str, dict[str, Any]],
    similarity: float,
    source: str,
) -> None:
    key = tuple(sorted([row_a["product_id"], row_b["product_id"]]))
    existing = candidate_map.get(key)
    source_priority = source == "same_supplier_exact_product_name"
    if existing and existing.get("candidate_source") == "same_supplier_exact_product_name" and not source_priority:
        return
    if existing and not source_priority and similarity <= float(existing.get("similarity") or 0):
        return
    attr_a = attributes[row_a["product_id"]]
    attr_b = attributes[row_b["product_id"]]
    candidate_map[key] = {
        "product_id_a": row_a["product_id"],
        "product_id_b": row_b["product_id"],
        "supplier_a": row_a["supplier_norm"],
        "supplier_b": row_b["supplier_norm"],
        "name_a": display_product_name(row_a),
        "name_b": display_product_name(row_b),
        "base_product_a": attr_a.get("base_product"),
        "base_product_b": attr_b.get("base_product"),
        "variant_a": attr_a.get("variant"),
        "variant_b": attr_b.get("variant"),
        "processing_a": attr_a.get("processing"),
        "processing_b": attr_b.get("processing"),
        "similarity": round(float(similarity), 4),
        "candidate_source": source,
    }


def merge_candidate_lists(*candidate_lists: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[tuple[str, str], dict[str, Any]] = {}
    for candidates in candidate_lists:
        for candidate in candidates:
            key = tuple(sorted([candidate["product_id_a"], candidate["product_id_b"]]))
            current = merged.get(key)
            if current is None or float(candidate.get("similarity") or 0) > float(current.get("similarity") or 0):
                merged[key] = candidate
    return list(merged.values())


def same_supplier_duplicate_decision(row: dict[str, Any]) -> bool:
    decision = row.get("decision")
    confidence = normalize_confidence(row.get("confidence"))
    review = bool(row.get("should_human_review"))
    return (
        decision == "exact_match" and confidence >= 85
    ) or (
        decision == "close_comparable" and confidence >= 90 and not review
    )


def build_duplicate_groups(product_ids: list[str], edges: list[dict[str, Any]]) -> list[list[str]]:
    parent = {pid: pid for pid in product_ids}

    def find(x: str) -> str:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: str, b: str) -> None:
        if a not in parent or b not in parent:
            return
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for edge in edges:
        union(edge["product_id_a"], edge["product_id_b"])

    groups: dict[str, list[str]] = defaultdict(list)
    for pid in product_ids:
        groups[find(pid)].append(pid)
    return [group for group in groups.values() if len(group) > 1]


def choose_longest_valid_offer(products: list[dict[str, Any]]) -> dict[str, Any]:
    return sorted(products, key=validity_rank, reverse=True)[0]


def validity_rank(product: dict[str, Any]) -> tuple:
    valid_from = parse_date(get(product, "valid_from"))
    valid_to = parse_date(get(product, "valid_to"))
    duration = (valid_to - valid_from).days if valid_from and valid_to else -1
    return (
        valid_to or date.min,
        duration,
        to_float(product.get("price_per_kg")) is not None,
        to_float(product.get("price")) is not None,
        -int(to_float(product.get("source_row_index")) or 0),
    )


def parse_date(value: Any) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def print_same_supplier_dedupe_summary(row_by_id: dict[str, dict[str, Any]], keep_ids: set[str], remove_ids: set[str]) -> None:
    examples = []
    for removed_id in sorted(remove_ids)[:10]:
        removed = row_by_id.get(removed_id, {})
        examples.append(
            f"{removed.get('supplier_norm')}: removed {display_product_name(removed)} "
            f"valid {removed.get('valid_from')} to {removed.get('valid_to')}"
        )
    if examples:
        print("Same-supplier duplicate examples:")
        for example in examples:
            print(f"- {example}")


def vector_similarity(a: np.ndarray, b: np.ndarray) -> float:
    denom = np.linalg.norm(a) * np.linalg.norm(b)
    return float(np.dot(a, b) / denom) if denom else 0.0


def clean(value: Any) -> str:
    return str(value or "").strip().casefold()


def judge_pairs(
    candidates: list[dict[str, Any]],
    df: pd.DataFrame,
    attributes: dict[str, dict[str, Any]],
    caches: Caches,
    pair_client,
    force: bool,
    skip_llm: bool,
    pair_workers: int,
) -> tuple[list[dict[str, Any]], int]:
    row_by_id = {row["product_id"]: row for row in df.to_dict("records")}
    out: list[dict[str, Any] | None] = [None] * len(candidates)
    hard_blocked = 0
    llm_jobs = []

    for idx, cand in enumerate(tqdm(candidates, desc="pairs prepare")):
        a_id, b_id = cand["product_id_a"], cand["product_id_b"]
        pair_key = make_pair_key(a_id, b_id)
        hard_reasons: list[str] = []
        soft_warnings: list[str] = []
        if cand.get("candidate_source") == "same_supplier_exact_product_name" and cand.get("supplier_a") == cand.get("supplier_b"):
            decision = {
                "pair_key": pair_key,
                "schema_version": PAIR_CACHE_VERSION,
                "product_id_a": a_id,
                "product_id_b": b_id,
                "decision": "exact_match",
                "confidence": 99,
                "canonical_name": cand.get("name_a") or cand.get("name_b") or "",
                "matching_attributes": ["same supplier", "same normalized product name"],
                "conflicting_attributes": [],
                "reason": "Same supplier and identical normalized product name; variant conflicts are checked at cluster level.",
                "should_human_review": False,
                "model": "deterministic-identity",
            }
            out[idx] = pair_debug_row(cand, hard_reasons, soft_warnings, decision)
            continue
        cached = caches.pairs.get(pair_key)
        cached_model_matches = (
            not pair_client
            or str(cached.get("model") or "") == pair_client["model"]
        ) if cached else False
        if cached and not force and int(cached.get("schema_version") or 0) == PAIR_CACHE_VERSION and cached_model_matches:
            decision = cached
            out[idx] = pair_debug_row(cand, hard_reasons, soft_warnings, decision)
        elif pair_client and not skip_llm:
            llm_jobs.append((idx, pair_key, cand, row_by_id[a_id], row_by_id[b_id], attributes[a_id], attributes[b_id], hard_reasons, soft_warnings))
        else:
            decision = judge_pair_fallback(pair_key, cand, attributes[a_id], attributes[b_id], soft_warnings)
            caches.append_pair(decision)
            out[idx] = pair_debug_row(cand, hard_reasons, soft_warnings, decision)

    if llm_jobs:
        workers = max(1, pair_workers)
        print(f"DeepSeek pair judgement: {len(llm_jobs)} uncached pairs with {workers} workers")
        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_map = {
                executor.submit(judge_pair_llm, pair_client, pair_key, cand, row_a, row_b, attr_a, attr_b): (
                    idx, pair_key, cand, hard_reasons, soft_warnings
                )
                for idx, pair_key, cand, row_a, row_b, attr_a, attr_b, hard_reasons, soft_warnings in llm_jobs
            }
            for future in tqdm(as_completed(future_map), total=len(future_map), desc="deepseek pairs"):
                idx, pair_key, cand, hard_reasons, soft_warnings = future_map[future]
                try:
                    decision = future.result()
                except Exception as exc:
                    decision = conservative_no_match(pair_key, cand, f"DeepSeek worker failed: {exc}")
                if pair_client and not decision.get("model"):
                    decision["model"] = pair_client["model"]
                caches.append_pair(decision)
                out[idx] = pair_debug_row(cand, hard_reasons, soft_warnings, decision)

    return [row for row in out if row is not None], hard_blocked


def second_round_judge_pairs(
    pair_rows: list[dict[str, Any]],
    df: pd.DataFrame,
    attributes: dict[str, dict[str, Any]],
    caches: Caches,
    pair_client,
    force: bool,
    skip_llm: bool,
    pair_workers: int,
) -> list[dict[str, Any]]:
    if skip_llm or not pair_client or not pair_rows:
        return pair_rows

    row_by_id = {row["product_id"]: row for row in df.to_dict("records")}
    jobs = []
    for idx, row in enumerate(pair_rows):
        if not needs_second_round(row):
            continue
        pair_key = make_pair_key(row["product_id_a"], row["product_id_b"])
        cached = caches.second_pairs.get(pair_key)
        cached_model_matches = (
            not pair_client
            or str(cached.get("model") or "") == pair_client["model"]
        ) if cached else False
        if cached and not force and int(cached.get("schema_version") or 0) == SECOND_JUDGE_CACHE_VERSION and cached_model_matches:
            pair_rows[idx] = apply_second_judgement(row, cached)
            continue
        jobs.append((idx, pair_key, row))

    if not jobs:
        return pair_rows

    workers = max(1, pair_workers)
    print(f"Second DeepSeek judgement: {len(jobs)} unclear pairs with {workers} workers")
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {}
        for idx, pair_key, pair_row in jobs:
            row_a = row_by_id.get(pair_row["product_id_a"], {})
            row_b = row_by_id.get(pair_row["product_id_b"], {})
            attr_a = attributes.get(pair_row["product_id_a"], {})
            attr_b = attributes.get(pair_row["product_id_b"], {})
            future = executor.submit(second_round_judge_llm, pair_client, pair_key, pair_row, row_a, row_b, attr_a, attr_b)
            future_map[future] = (idx, pair_key, pair_row)

        for future in tqdm(as_completed(future_map), total=len(future_map), desc="second judge"):
            idx, pair_key, pair_row = future_map[future]
            try:
                judgement = future.result()
            except Exception as exc:
                judgement = second_round_fallback(pair_key, f"Second judge failed: {exc}")
            if pair_client and not judgement.get("model"):
                judgement["model"] = pair_client["model"]
            caches.append_second_pair(judgement)
            pair_rows[idx] = apply_second_judgement(pair_row, judgement)

    return pair_rows


def needs_second_round(row: dict[str, Any]) -> bool:
    decision = row.get("decision")
    confidence = normalize_confidence(row.get("confidence"))
    if decision == "no_match" and confidence >= 90:
        return False
    if decision == "exact_match" and confidence >= 92 and not bool(row.get("should_human_review")):
        return False
    return (
        bool(row.get("should_human_review"))
        or decision in {"close_comparable", "same_family_not_comparable"}
        or 70 <= confidence < 92
    )


SECOND_ROUND_PROMPT = """Du bist die zweite und finale Kontrollinstanz für ein automatisches Produkt-Matching.

Ziel: möglichst vollautomatisch entscheiden, damit pro Woche höchstens 1-2 Fälle manuell offen bleiben.
Du bekommst die erste LLM-Entscheidung und musst sie final überprüfen.

Erlaubte final_action:
- MERGE: Die Produkte sind derselbe Artikel oder so klar derselbe Einkaufstyp, dass sie in eine Zeile gehören.
- NO_MERGE: Nicht derselbe Artikel. Dann sollen sie getrennt bleiben.
- REVIEW: Nur wenn es wirklich nicht zuverlässig entscheidbar ist.

Bitte nutze REVIEW sehr sparsam.

Verbindliche Variantenregeln:
- Unterschiedliches Kaliber oder unterschiedliche Packungsgröße erzeugt KEINE eigene Variante.
- Ein fehlendes Attribut darf mit einem bekannten Attribut zusammengeführt werden, solange kein Widerspruch besteht.
- Zwei unterschiedliche bekannte Herkünfte bleiben getrennt. Herkunftskombinationen wie "UA, RO" sind ein eigener Wert.
- Zwei unterschiedliche bekannte Marken bleiben getrennt; eine fehlende Marke allein verhindert MERGE nicht.
- Duroc, Berkshire und vergleichbare klar benannte Rassen/Premiumlinien bleiben von generischer Ware getrennt.
- Frisch und tiefgekühlt bleiben getrennt.
- Preis, Händler, Menge, Verpackungsart und Gültigkeitszeitraum sind Angebotsdaten. Unterschiedliche Werte bleiben als mehrere Angebote in derselben Produktzeile erhalten.

Beispiele:
- "Kalbs Steakhüfte Hell 1.2 Kg" vs "Kalbs-Steakhüfte Hell" => MERGE.
- "Kalbs-Steakhüfte hell" vs "Kalbs Steakhüfte hell" mit unterschiedlicher Quelle/Woche => MERGE.
- "Schweinelachs" vs "Schweinelachse" => MERGE.
- "Dorade Royal" vs "Dorade Royal 5 kg Kiste" => MERGE.
- "Schweinenacken" vs "Schweinefilet" => NO_MERGE.
- "Milch" vs "Milch-Schnitte" => NO_MERGE.
- "Olivenöl" vs "Sonnenblumenöl" => NO_MERGE.

Erste Entscheidung:
{first_decision}

Produkt A:
{original_a}

Hinweise A:
{attributes_a}

Produkt B:
{original_b}

Hinweise B:
{attributes_b}

Gib ausschließlich striktes JSON zurück:
{{
  "pair_key": "...",
  "final_action": "MERGE",
  "confidence": 0,
  "canonical_name": "...",
  "reason": "...",
  "manual_review_needed": false
}}

confidence ist immer eine ganze Zahl von 0 bis 100, nicht 0 bis 1.
"""


def second_round_judge_llm(
    pair_client: dict[str, str],
    pair_key: str,
    pair_row: dict[str, Any],
    row_a: dict[str, Any],
    row_b: dict[str, Any],
    attr_a: dict[str, Any],
    attr_b: dict[str, Any],
) -> dict[str, Any]:
    prompt = SECOND_ROUND_PROMPT.format(
        first_decision=json.dumps({k: pair_row.get(k, "") for k in ["decision", "confidence", "reason", "conflicting_attributes", "should_human_review", "similarity"]}, ensure_ascii=False, indent=2),
        original_a=json.dumps(serializable_original(row_a), ensure_ascii=False, indent=2),
        attributes_a=json.dumps(attr_a, ensure_ascii=False, indent=2),
        original_b=json.dumps(serializable_original(row_b), ensure_ascii=False, indent=2),
        attributes_b=json.dumps(attr_b, ensure_ascii=False, indent=2),
    )
    data = call_deepseek_json(
        pair_client,
        "Du bist die finale deutsche Kontrollinstanz für Produkt-Matching. Gib ausschließlich striktes JSON zurück.",
        prompt,
        max_tokens=700,
    )
    return normalize_second_judgement(pair_key, data)


def normalize_second_judgement(pair_key: str, data: dict[str, Any]) -> dict[str, Any]:
    action = str(data.get("final_action") or "REVIEW").strip().upper()
    if action not in {"MERGE", "NO_MERGE", "REVIEW"}:
        action = "REVIEW"
    return {
        "pair_key": pair_key,
        "schema_version": SECOND_JUDGE_CACHE_VERSION,
        "final_action": action,
        "confidence": normalize_confidence(data.get("confidence")),
        "canonical_name": data.get("canonical_name", ""),
        "reason": data.get("reason", ""),
        "manual_review_needed": bool(data.get("manual_review_needed", action == "REVIEW")),
    }


def second_round_fallback(pair_key: str, reason: str) -> dict[str, Any]:
    return {
        "pair_key": pair_key,
        "schema_version": SECOND_JUDGE_CACHE_VERSION,
        "final_action": "REVIEW",
        "confidence": 0,
        "canonical_name": "",
        "reason": reason,
        "manual_review_needed": True,
    }


def apply_second_judgement(pair_row: dict[str, Any], judgement: dict[str, Any]) -> dict[str, Any]:
    row = dict(pair_row)
    action = judgement.get("final_action")
    confidence = normalize_confidence(judgement.get("confidence"))
    reason = judgement.get("reason") or ""
    row["second_judge_action"] = action
    row["second_judge_confidence"] = confidence
    row["second_judge_reason"] = reason

    if action == "MERGE" and confidence >= 80:
        row["decision"] = "exact_match" if confidence >= 88 else "close_comparable"
        row["confidence"] = confidence
        row["canonical_name"] = judgement.get("canonical_name") or row.get("canonical_name", "")
        row["reason"] = f"Second judge MERGE: {reason}"
        row["should_human_review"] = bool(judgement.get("manual_review_needed")) and confidence < 90
    elif action == "NO_MERGE" and confidence >= 80:
        row["decision"] = "no_match"
        row["confidence"] = confidence
        row["reason"] = f"Second judge NO_MERGE: {reason}"
        row["should_human_review"] = False
    else:
        row["reason"] = f"Second judge REVIEW: {reason}"
        row["should_human_review"] = True
    return row


def pair_debug_row(
    cand: dict[str, Any],
    hard_reasons: list[str],
    soft_warnings: list[str],
    decision: dict[str, Any],
) -> dict[str, Any]:
    return {**cand, **{
        "hard_blocked": bool(hard_reasons),
        "block_reasons": "; ".join(hard_reasons),
        "soft_warnings": "; ".join(soft_warnings),
        "decision": decision.get("decision", "no_match"),
        "confidence": normalize_confidence(decision.get("confidence")),
        "canonical_name": decision.get("canonical_name", ""),
        "reason": decision.get("reason", ""),
        "conflicting_attributes": "; ".join(map(str, decision.get("conflicting_attributes", []))),
        "should_human_review": bool(decision.get("should_human_review", True)),
    }}


def make_pair_key(a: str, b: str) -> str:
    return "||".join(sorted([a, b]))


def blockers_and_warnings(cand: dict[str, Any], attr_a: dict[str, Any], attr_b: dict[str, Any], row_a: dict[str, Any], row_b: dict[str, Any]) -> tuple[list[str], list[str]]:
    hard = []
    soft = []
    base_a, base_b = clean(attr_a.get("base_product")), clean(attr_b.get("base_product"))
    if base_a and base_b and base_a != base_b:
        hard.append("base product differs")
    family_a, family_b = meat_family(base_a), meat_family(base_b)
    if family_a and family_b and family_a == family_b and base_a != base_b:
        hard.append(f"different {family_a} cut")
    ff_a, ff_b = clean(attr_a.get("fresh_or_frozen")), clean(attr_b.get("fresh_or_frozen"))
    if {ff_a, ff_b} == {"fresh", "frozen"}:
        hard.append("fresh vs frozen conflict")
    if bool(attr_a.get("is_accessory_or_related_product")) != bool(attr_b.get("is_accessory_or_related_product")):
        hard.append("accessory/core conflict")
    if {clean(attr_a.get("variant")), clean(attr_b.get("variant"))} == {"white", "green"}:
        hard.append("white vs green asparagus")
    if {clean(attr_a.get("processing")), clean(attr_b.get("processing"))} == {"peeled", "unpeeled"}:
        hard.append("peeled vs unpeeled")
    if {clean(attr_a.get("quality_class")), clean(attr_b.get("quality_class"))} == {"class i", "class ii"}:
        hard.append("class I vs class II")
    brand_a, brand_b = clean(attr_a.get("brand")), clean(attr_b.get("brand"))
    if brand_a and brand_b and brand_a != brand_b and any(clean(attr.get("packaging")) for attr in [attr_a, attr_b]):
        hard.append("different clear brand")
    if asparagus_related_conflict(attr_a, attr_b):
        hard.append("asparagus accessory matched to asparagus")

    for key, label in [("quantity_unit", "unknown quantity unit"), ("packaging", "packaging mismatch"), ("origin", "origin mismatch"), ("calibre", "calibre mismatch")]:
        a, b = clean(attr_a.get(key)), clean(attr_b.get(key))
        if key == "quantity_unit" and ("unknown" in {a, b}):
            soft.append(label)
        elif a and b and a != b:
            soft.append(label)
    p_a, p_b = to_float(row_a.get("price_per_kg")), to_float(row_b.get("price_per_kg"))
    if p_a and p_b and max(p_a, p_b) / max(min(p_a, p_b), 0.01) > 1.8:
        soft.append("very different price_per_kg")
    return hard, soft


def asparagus_related_conflict(attr_a: dict[str, Any], attr_b: dict[str, Any]) -> bool:
    bases = {clean(attr_a.get("base_product")), clean(attr_b.get("base_product"))}
    accessory = bool(attr_a.get("is_accessory_or_related_product")) or bool(attr_b.get("is_accessory_or_related_product"))
    return "asparagus" in " ".join(bases) and accessory


PAIR_PROMPT = """Du vergleichst Produkte aus deutschen Großhandels-Werbeprospekten.

Entscheide, ob Produkt A und Produkt B in dieselbe Vergleichszeile gehören.
Die Vektorsuche hat diese zwei Produkte nur als Kandidaten vorgeschlagen. Du bist die finale Entscheidung.
Die Produkte können aus unterschiedlichen Wettbewerbern kommen oder vom selben Supplier aus demselben oder aus unterschiedlichen Prospekten. Wenn es derselbe Supplier und dieselbe Produktvariante ist, bewerte es trotzdem als exact_match. Der technische Prozess führt die Produktzeile zusammen und bewahrt alle unterschiedlichen Angebote vollständig auf.

Antworte inhaltlich auf Deutsch. Gib aber das JSON-Schema exakt mit den vorgegebenen englischen Schlüsseln zurück.

Entscheidungen:
- exact_match: praktisch derselbe Artikel. Kleine Unterschiede in Schreibweise, Bindestrichen, Großschreibung, fehlender Mengenangabe oder Formatierung sind egal.
- close_comparable: sehr wahrscheinlich derselbe Einkaufstyp, aber ein kleiner Unterschied bleibt unklar. Nur verwenden, wenn ein Preisvergleich noch sinnvoll ist.
- same_family_not_comparable: gleiche Produktfamilie, aber nicht derselbe Artikel.
- no_match: anderer Artikel.

Wichtige Beispiele:
- "Kalbs Steakhüfte Hell 1.2 Kg" und "Kalbs-Steakhüfte Hell" = exact_match. Die 1.2 kg sind nur Angebots-/Packungsinfo, der Artikel ist gleich.
- "Kalbs-Steakhüfte Hell" aus Prospekt KW18 und "Kalbs Steakhüfte Hell 1.2 Kg" aus Prospekt KW19 vom selben Supplier = exact_match.
- "Spargel Weiß" und "Weißer Spargel" = exact_match.
- "Schweinenacken" und "Schweine Nacken" = exact_match.
- "Schweinelachs" und "Schweinelachse" = exact_match.
- "Dorade Royal" und "Dorade Royal 5 kg Kiste" = exact_match oder close_comparable, wenn nur die Verpackung fehlt.
- "Schweinenacken" und "Schweinefilet" = same_family_not_comparable, nicht mergen.
- "Schweineoberschale" und "Schweineschnitzel" = same_family_not_comparable, nicht mergen.
- "Rinderfilet" und "Rinder-Rib-Eye" = same_family_not_comparable, nicht mergen.
- "Milch" und "Milch-Schnitte" = no_match.
- "Olivenöl" und "Sonnenblumenöl" = same_family_not_comparable, nicht mergen.
- "Pommes Frites" und "Süßkartoffel-Pommes" = same_family_not_comparable, nicht mergen.
- "Broccoli" und "Blumenkohl" = same_family_not_comparable, nicht mergen.

Verbindliche Variantenregeln:
- Unterschiedliches Kaliber erzeugt KEINE eigene Variante und ist kein Grund gegen einen Match.
- Unterschiedliche Packungsgröße, Verpackungsart, Preis oder Gültigkeitszeitraum erzeugen KEINE eigene Variante. Das sind getrennte Angebote derselben Variante.
- Fehlt ein Attribut nur auf einer Seite, darf gematcht werden, solange kein anderer Widerspruch besteht.
- Unterschiedliche bekannte Herkunft = nicht mergen. Gleiche Herkunft = mergen. Bekannte Herkunft plus fehlende Herkunft darf gemergt werden. Kombinationen wie "UA, RO" gelten als eigener Herkunftswert.
- Unterschiedliche bekannte Marken = nicht mergen. Eine fehlende Marke allein ist kein Trennungsgrund.
- Duroc, Berkshire und vergleichbare klar benannte Rassen/Premiumlinien = von generischer Ware getrennte Variante.
- Frisch und tiefgekühlt = nicht mergen.
- Zertifizierungen sind Attribute und dürfen nicht stillschweigend aus dem kanonischen Namen oder den Quelldaten entfernt werden.

Nutze vor allem:
- Produktname, getrennte Felder brand/product, Beschreibung, Herkunft, Qualität, Zuschnitt/Sorte, Verarbeitungsart und Menge.
- Der Preis ist kein Identitätsmerkmal.
- Wenn eine Mengenangabe nur bei einem Produkt fehlt, ist das kein Grund gegen einen Match.
- Wenn beide Produkte dieselbe klare Artikelbezeichnung haben, sollen sie gematcht werden.
- Wenn es nur dieselbe Kategorie ist, aber anderer Artikel/Zuschnitt/Sorte/Geschmack/Marke, nicht mergen.

Setze should_human_review auf false, wenn du sicher bist.
Setze should_human_review auf true, wenn der Match fachlich möglich ist, aber eine wichtige Unsicherheit bleibt.

Produkt A original:
{original_a}

Produkt A extrahierte Hinweise:
{attributes_a}

Produkt B original:
{original_b}

Produkt B extrahierte Hinweise:
{attributes_b}

Gib ausschließlich striktes JSON zurück:
{{
  "product_id_a": "...",
  "product_id_b": "...",
  "decision": "exact_match",
  "confidence": 0,
  "canonical_name": "...",
  "matching_attributes": [],
  "conflicting_attributes": [],
  "reason": "...",
  "should_human_review": true
}}

confidence ist immer eine ganze Zahl von 0 bis 100, nicht 0 bis 1.
"""


def judge_pair_llm(pair_client: dict[str, str], pair_key: str, cand: dict[str, Any], row_a: dict[str, Any], row_b: dict[str, Any], attr_a: dict[str, Any], attr_b: dict[str, Any]) -> dict[str, Any]:
    prompt = PAIR_PROMPT.format(
        original_a=json.dumps(serializable_original(row_a), ensure_ascii=False, indent=2),
        attributes_a=json.dumps(attr_a, ensure_ascii=False, indent=2),
        original_b=json.dumps(serializable_original(row_b), ensure_ascii=False, indent=2),
        attributes_b=json.dumps(attr_b, ensure_ascii=False, indent=2),
    )
    for attempt in range(2):
        try:
            data = call_deepseek_pair_judge(
                pair_client,
                prompt if attempt == 0 else prompt + "\nReturn only repaired strict JSON.",
            )
            return normalize_pair_decision(pair_key, cand, data)
        except Exception as exc:
            if attempt == 1:
                return conservative_no_match(pair_key, cand, f"LLM pair fallback after error: {exc}")
    return conservative_no_match(pair_key, cand, "unreachable")


def call_deepseek_pair_judge(pair_client: dict[str, str], prompt: str) -> dict[str, Any]:
    return call_deepseek_json(
        pair_client,
        "Du bist ein sehr genauer deutscher Produkt-Matching-Judge. Gib ausschließlich striktes JSON zurück.",
        prompt,
        max_tokens=800,
    )


def call_deepseek_json(pair_client: dict[str, str], system_prompt: str, prompt: str, max_tokens: int) -> dict[str, Any]:
    response = requests.post(
        pair_client["base_url"].rstrip("/") + "/chat/completions",
        headers={
            "Authorization": f"Bearer {pair_client['api_key']}",
            "Content-Type": "application/json",
        },
        json={
            "model": pair_client["model"],
            "messages": [
                {
                    "role": "system",
                    "content": system_prompt,
                },
                {
                    "role": "user",
                    "content": prompt,
                },
            ],
            "temperature": 0,
            "max_tokens": max_tokens,
            "stream": False,
        },
        timeout=120,
    )
    response.raise_for_status()
    payload = response.json()
    choices = payload.get("choices") or []
    if not choices:
        raise RuntimeError("DeepSeek response has no choices")
    content = choices[0].get("message", {}).get("content", "")
    if not content:
        raise RuntimeError("DeepSeek response has no content")
    return safe_json_loads(content)


def serializable_original(row: dict[str, Any]) -> dict[str, Any]:
    keys = ["product_id", "supplier_norm", "product_name", "brand", "product", "brand_product_confidence", "description", "category", "origin", "unit", "quantity", "price", "price_per_kg", "valid_from", "valid_to", "source_file", "source_page"]
    return {key: row.get(key, "") for key in keys}


def normalize_pair_decision(pair_key: str, cand: dict[str, Any], data: dict[str, Any]) -> dict[str, Any]:
    decision = data.get("decision", "no_match")
    if decision not in {"exact_match", "close_comparable", "same_family_not_comparable", "no_match"}:
        decision = "no_match"
    return {
        "pair_key": pair_key,
        "schema_version": PAIR_CACHE_VERSION,
        "product_id_a": cand["product_id_a"],
        "product_id_b": cand["product_id_b"],
        "decision": decision,
        "confidence": normalize_confidence(data.get("confidence")),
        "canonical_name": data.get("canonical_name", ""),
        "matching_attributes": data.get("matching_attributes", []) if isinstance(data.get("matching_attributes"), list) else [],
        "conflicting_attributes": data.get("conflicting_attributes", []) if isinstance(data.get("conflicting_attributes"), list) else [],
        "reason": data.get("reason", ""),
        "should_human_review": bool(data.get("should_human_review", True)),
    }


def conservative_no_match(pair_key: str, cand: dict[str, Any], reason: str) -> dict[str, Any]:
    return {
        "pair_key": pair_key,
        "schema_version": PAIR_CACHE_VERSION,
        "product_id_a": cand["product_id_a"],
        "product_id_b": cand["product_id_b"],
        "decision": "no_match",
        "confidence": 0,
        "canonical_name": "",
        "matching_attributes": [],
        "conflicting_attributes": [],
        "reason": reason,
        "should_human_review": True,
    }


def judge_pair_fallback(pair_key: str, cand: dict[str, Any], attr_a: dict[str, Any], attr_b: dict[str, Any], soft_warnings: list[str]) -> dict[str, Any]:
    sim = float(cand.get("similarity") or 0)
    if sim >= 0.92:
        decision, confidence, review = "exact_match", 91, False
    elif sim >= 0.84:
        decision, confidence, review = "close_comparable", 88, True
    else:
        decision, confidence, review = "no_match", 50, True
    return {
        "pair_key": pair_key,
        "schema_version": PAIR_CACHE_VERSION,
        "product_id_a": cand["product_id_a"],
        "product_id_b": cand["product_id_b"],
        "decision": decision,
        "confidence": confidence,
        "canonical_name": "",
        "matching_attributes": [],
        "conflicting_attributes": soft_warnings,
        "reason": "Fallback ohne LLM: nur Vektorähnlichkeit genutzt",
        "should_human_review": review,
    }


def build_clusters(df: pd.DataFrame, attributes: dict[str, dict[str, Any]], pair_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ids = df["product_id"].tolist()
    row_by_id = {row["product_id"]: row for row in df.to_dict("records")}
    parent = {pid: pid for pid in ids}
    members = {pid: {pid} for pid in ids}

    def find(x: str) -> str:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: str, b: str) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra
            members[ra].update(members.pop(rb))

    # Strongest decisions are applied first. Before every union the complete
    # two clusters are checked, so an item with missing attributes cannot act
    # as a bridge between two known, contradictory variants.
    accepted_rows = sorted(
        (row for row in pair_rows if accepted_merge_edge(row)),
        key=lambda row: normalize_confidence(row.get("confidence")),
        reverse=True,
    )
    for row in accepted_rows:
        a_id, b_id = row["product_id_a"], row["product_id_b"]
        if a_id not in parent or b_id not in parent:
            continue
        root_a, root_b = find(a_id), find(b_id)
        if root_a == root_b:
            continue
        conflicts = cluster_variant_conflicts(
            members[root_a],
            members[root_b],
            row_by_id,
            attributes,
        )
        if conflicts:
            row["cluster_merge_blocked"] = True
            row["cluster_block_reasons"] = "; ".join(conflicts)
            continue
        union(a_id, b_id)

    groups = defaultdict(list)
    for pid in ids:
        groups[find(pid)].append(pid)

    pair_by_cluster = defaultdict(list)
    for row in pair_rows:
        if not accepted_merge_edge(row) or row.get("cluster_merge_blocked"):
            continue
        root_a, root_b = find(row["product_id_a"]), find(row["product_id_b"])
        if root_a == root_b:
            pair_by_cluster[root_a].append(row)

    clusters = []
    for index, (root, product_ids) in enumerate(groups.items(), 1):
        pair_edges = pair_by_cluster[root]
        clusters.append({
            "canonical_product_id": f"c_{index:05d}",
            "product_ids": product_ids,
            "pair_edges": pair_edges,
        })
    return clusters


def cluster_variant_conflicts(
    member_ids_a: set[str],
    member_ids_b: set[str],
    row_by_id: dict[str, dict[str, Any]],
    attributes: dict[str, dict[str, Any]],
) -> list[str]:
    conflicts: list[str] = []
    for a_id in member_ids_a:
        for b_id in member_ids_b:
            pair_conflicts = product_variant_conflicts(
                row_by_id[a_id],
                row_by_id[b_id],
                attributes.get(a_id, {}),
                attributes.get(b_id, {}),
            )
            for reason in pair_conflicts:
                if reason not in conflicts:
                    conflicts.append(reason)
    return conflicts


def product_variant_conflicts(
    row_a: dict[str, Any],
    row_b: dict[str, Any],
    attr_a: dict[str, Any],
    attr_b: dict[str, Any],
) -> list[str]:
    conflicts: list[str] = []
    same_product_name = exact_product_name_key(row_a) == exact_product_name_key(row_b)

    base_a, base_b = clean(attr_a.get("base_product")), clean(attr_b.get("base_product"))
    if not same_product_name and base_a and base_b and base_a != base_b:
        conflicts.append("different known base product")

    temperature_a = explicit_temperature_state(row_a, attr_a)
    temperature_b = explicit_temperature_state(row_b, attr_b)
    if temperature_a and temperature_b and temperature_a != temperature_b:
        conflicts.append("fresh vs frozen")

    origin_a = canonical_origin_identity(get(row_a, "origin") or attr_a.get("origin"))
    origin_b = canonical_origin_identity(get(row_b, "origin") or attr_b.get("origin"))
    if origin_a and origin_b and not compatible_origin_identities(origin_a, origin_b):
        conflicts.append("different known origin")

    brand_a = verified_brand_identity(row_a)
    brand_b = verified_brand_identity(row_b)
    if brand_a and brand_b and brand_a != brand_b:
        conflicts.append("different known brand")

    variant_a, variant_b = clean(attr_a.get("variant")), clean(attr_b.get("variant"))
    if variant_a and variant_b and variant_a != variant_b:
        conflicts.append("different known variant")

    premium_a = premium_variant_signals(row_a)
    premium_b = premium_variant_signals(row_b)
    if premium_a != premium_b and (premium_a or premium_b):
        conflicts.append("premium line or breed differs")

    processing_a = explicit_processing_identity(row_a, attr_a)
    processing_b = explicit_processing_identity(row_b, attr_b)
    if processing_a and processing_b and processing_a != processing_b:
        conflicts.append("different known processing state")

    quality_a, quality_b = clean(attr_a.get("quality_class")), clean(attr_b.get("quality_class"))
    if quality_a and quality_b and quality_a != quality_b:
        conflicts.append("different known quality class")

    # Calibre, package size, packaging, price and validity are deliberately
    # absent here: by business decision they are offer attributes, not variant
    # separators.
    return conflicts


def explicit_temperature_state(row: dict[str, Any], attr: dict[str, Any]) -> str:
    text = " ".join([
        get(row, "category"),
        get(row, "product") or get(row, "product_name"),
        get(row, "description"),
    ]).casefold()
    if any(token in text for token in ["tiefkühl", "tiefkuehl", "gefroren", "frozen", " tk "]) or clean(get(row, "category")) == "tk":
        return "frozen"
    if any(token in text for token in ["frisch", "fresh", "gekühlt", "gekuehlt", "chilled"]):
        return "fresh"
    state = clean(attr.get("fresh_or_frozen"))
    return state if state in {"fresh", "frozen"} and clean(attr.get("notes")) != "rule-based fallback" else ""


ORIGIN_ADJECTIVE_CODES = {
    "deutsch": "DE", "spanisch": "ES", "franzoesisch": "FR", "franzosisch": "FR",
    "daenisch": "DK", "danisch": "DK", "irisch": "IE", "italienisch": "IT",
    "polnisch": "PL", "rumaenisch": "RO", "rumanisch": "RO", "ukrainisch": "UA",
    "argentinisch": "AR", "uruguayisch": "UY", "norwegisch": "NO",
    "niederlaendisch": "NL", "hollaendisch": "NL", "neuseelaendisch": "NZ",
    "australisch": "AU", "suedafrikanisch": "ZA", "chilenisch": "CL",
}


def canonical_origin_identity(value: Any) -> tuple[str, ...]:
    text = str(value or "").strip()
    if not text or clean(text) in {"unknown", "unbekannt", "none", "null"}:
        return ()
    converted = origin_country_codes(text)
    converted_codes = [part.strip() for part in converted.split(",") if part.strip() in VALID_COUNTRY_CODES]
    if converted_codes and len(", ".join(converted_codes)) == len(converted):
        return tuple(sorted(set(converted_codes)))
    normalized = normalize_country_alias(text)
    adjective_codes = {
        code
        for stem, code in ORIGIN_ADJECTIVE_CODES.items()
        if re.search(rf"(^| ){re.escape(stem)}[a-z]*($| )", normalized)
    }
    if adjective_codes:
        return tuple(sorted(adjective_codes))
    normalized = re.sub(r"\b(herkunft|origin|aus|from)\b", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return (f"text:{normalized}",) if normalized else ()


def compatible_origin_identities(origin_a: tuple[str, ...], origin_b: tuple[str, ...]) -> bool:
    if origin_a == origin_b:
        return True
    if len(origin_a) == len(origin_b) == 1 and origin_a[0].startswith("text:") and origin_b[0].startswith("text:"):
        text_a, text_b = origin_a[0][5:], origin_b[0][5:]
        return text_a in text_b or text_b in text_a
    return False


VERIFIED_BRAND_ALIASES = [
    ("edeka foodservice", ["edeka foodservice"]),
    ("metro professional", ["metro professional"]),
    ("metro chef", ["metro chef"]),
    ("aro", ["aro"]),
    ("milram", ["milram"]),
    ("schleiz", ["schleiz", "schleizer"]),
    ("quality", ["quality"]),
    ("economy", ["economy"]),
    ("henkelmann", ["henkelmann"]),
    ("meemken", ["meemken"]),
    ("aviko", ["aviko"]),
    ("foodservice", ["foodservice"]),
    ("edeka", ["edeka"]),
    ("metro", ["metro"]),
    ("chef", ["chef"]),
]


def verified_brand_identity(row: dict[str, Any]) -> str:
    explicit = duplicate_key_text(get(row, "brand"))
    evidence = duplicate_key_text(" ".join([get(row, "product_name"), get(row, "product"), get(row, "description")]))
    for canonical, aliases in VERIFIED_BRAND_ALIASES:
        for alias in aliases:
            pattern = rf"(^| ){re.escape(alias)}($| )"
            if (explicit and re.search(pattern, explicit)) or re.search(pattern, evidence):
                return canonical
    return explicit


def premium_variant_signals(row: dict[str, Any]) -> tuple[str, ...]:
    text = duplicate_key_text(" ".join([get(row, "product_name"), get(row, "product"), get(row, "description")]))
    signals = [
        signal
        for signal in ["duroc", "berkshire", "free range", "dry aged", "label rouge"]
        if re.search(rf"(^| ){re.escape(signal)}($| )", text)
    ]
    return tuple(signals)


def explicit_processing_identity(row: dict[str, Any], attr: dict[str, Any]) -> str:
    text = normalize_country_alias(
        " ".join([get(row, "product_name"), get(row, "product"), get(row, "description")])
    )
    signatures = [
        ("dried", ["getrocknet", "dried"]),
        ("pureed", ["passiert", "pueree", "puree"]),
        ("smoked", ["gerauchert", "geraeuchert", "smoked"]),
        ("marinated", ["mariniert", "marinated"]),
        ("cooked", ["gekocht", "cooked"]),
        ("cut", ["geschnitten", "gewurfelt", "gewuerfelt", "sliced", "diced"]),
        ("peeled", ["geschalt", "geschaelt", "peeled"]),
        ("unpeeled", ["ungeschalt", "ungeschaelt", "unpeeled"]),
    ]
    for identity, terms in signatures:
        if any(term in text for term in terms):
            return identity
    processing = clean(attr.get("processing"))
    return processing if processing not in {"", "unknown", "none", "null"} else ""


def validate_cluster_partition(df: pd.DataFrame, clusters: list[dict[str, Any]]) -> None:
    expected = df["product_id"].tolist()
    assigned = [product_id for cluster in clusters for product_id in cluster["product_ids"]]
    if len(assigned) != len(set(assigned)):
        raise RuntimeError("Cluster validation failed: a product was assigned more than once")
    if set(assigned) != set(expected):
        missing = sorted(set(expected) - set(assigned))
        extra = sorted(set(assigned) - set(expected))
        raise RuntimeError(f"Cluster validation failed: missing={missing[:5]} extra={extra[:5]}")


def accepted_merge_edge(row: dict[str, Any]) -> bool:
    if row.get("candidate_source") == "same_supplier_exact_product_name" and row.get("supplier_a") == row.get("supplier_b"):
        return True
    decision = row.get("decision")
    confidence = normalize_confidence(row.get("confidence"))
    review = bool(row.get("should_human_review"))
    return (
        decision == "exact_match" and confidence >= AUTO_MERGE_EXACT_THRESHOLD
    ) or (
        decision == "close_comparable" and confidence >= AUTO_MERGE_CLOSE_THRESHOLD and not review
    )


def build_output_rows(
    df: pd.DataFrame,
    attributes: dict[str, dict[str, Any]],
    pair_rows: list[dict[str, Any]],
    clusters: list[dict[str, Any]],
    grouping_audit: list[dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    row_by_id = {row["product_id"]: row for row in df.to_dict("records")}
    matched_rows = []
    review_rows = []
    for cluster in clusters:
        products = [row_by_id[pid] for pid in cluster["product_ids"]]
        attrs = [attributes[pid] for pid in cluster["product_ids"]]
        canonical_name = choose_canonical_name(cluster, attrs)
        suppliers = defaultdict(list)
        for product in products:
            suppliers[product["supplier_norm"]].append(product)
        unique_suppliers = {
            supplier: dedupe_supplier_offers(
                supplier_products,
                grouping_audit=grouping_audit,
                cluster_id=cluster["canonical_product_id"],
                supplier=supplier,
            )
            for supplier, supplier_products in suppliers.items()
        }
        unique_products = [product for supplier_products in unique_suppliers.values() for product in supplier_products]
        review_needed, review_reasons = cluster_review_reasons(unique_products, attrs, cluster["pair_edges"], unique_suppliers)
        primary_attr = choose_primary_attr(attrs)
        matched = {
            "canonical_product_id": cluster["canonical_product_id"],
            "canonical_product_name": canonical_name,
            "category": most_common([get(p, "category") for p in unique_products]),
            "product": canonical_name,
            "brand": choose_best_source_text([get(p, "brand") for p in unique_products]),
            "description": choose_best_source_text([get(p, "description") for p in unique_products]),
            "base_product": primary_attr.get("base_product"),
            "variant": primary_attr.get("variant"),
            "processing": primary_attr.get("processing"),
            "quality_class": primary_attr.get("quality_class"),
            "origin": choose_best_source_text([get(p, "origin") for p in unique_products]) or primary_attr.get("origin"),
            "packaging": primary_attr.get("packaging"),
            "unit_basis": primary_attr.get("unit_basis"),
            "match_confidence": cluster_confidence(cluster["pair_edges"]),
            "review_needed": str(bool(review_needed)).upper(),
            "match_reason": "; ".join(review_reasons) if review_reasons else "single product or accepted exact match",
        }
        for supplier in SUPPLIER_ORDER:
            offers = sorted(unique_suppliers.get(supplier, []), key=offer_sort_key)
            matched[supplier] = "\n\n".join(format_offer_cell(p) for p in offers)
            matched[f"{supplier}_short"] = "\n\n".join(format_offer_cell_short(p) for p in offers)
        matched_rows.append(matched)
        if review_needed:
            review_rows.append(make_review_row(cluster, canonical_name, products, review_reasons, cluster["pair_edges"]))
    return matched_rows, review_rows


def dedupe_supplier_offers(
    products: list[dict[str, Any]],
    grouping_audit: list[dict[str, Any]] | None = None,
    cluster_id: str = "",
    supplier: str = "",
) -> list[dict[str, Any]]:
    parent = list(range(len(products)))

    def find(index: int) -> int:
        while parent[index] != index:
            parent[index] = parent[parent[index]]
            index = parent[index]
        return index

    def union(left: int, right: int) -> None:
        root_left, root_right = find(left), find(right)
        if root_left != root_right:
            parent[root_right] = root_left

    exact_groups: dict[tuple[str, ...], list[int]] = defaultdict(list)
    source_groups: dict[tuple[str, ...], list[int]] = defaultdict(list)
    for index, product in enumerate(products):
        exact_key = exact_commercial_offer_key(product)
        if exact_key:
            exact_groups[exact_key].append(index)
        source_key = same_source_commercial_offer_key(product)
        if source_key:
            source_groups[source_key].append(index)
    for indexes in exact_groups.values():
        for index in indexes[1:]:
            union(indexes[0], index)
    for indexes in source_groups.values():
        for position, left in enumerate(indexes):
            for right in indexes[position + 1:]:
                if compatible_offer_extraction_representations(products[left], products[right]):
                    union(left, right)

    components: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for index, product in enumerate(products):
        components[find(index)].append(product)

    unique_offers = []
    for component in components.values():
        merged = merge_offer_representations(component)
        unique_offers.append(merged)
        if grouping_audit is not None:
            grouping_audit.append({
                "cluster_id": cluster_id,
                "supplier": supplier,
                "rule": (
                    "same commercial offer; source extraction representations coalesced"
                    if len(component) > 1
                    else "unique commercial offer"
                ),
                "kept_product_id": merged.get("product_id", ""),
                "merged_product_ids": [item.get("product_id", "") for item in component],
                "sources": [source_label(item) for item in component],
                "price": get(merged, "price"),
                "valid_from": get(merged, "valid_from"),
                "valid_to": get(merged, "valid_to"),
                "final_amount": amount_label(merged),
            })
    return unique_offers


def validate_offer_accounting(df: pd.DataFrame, audit_rows: list[dict[str, Any]]) -> None:
    expected = set(df["product_id"].tolist())
    accounted = [
        product_id
        for row in audit_rows
        for product_id in row.get("merged_product_ids", [])
        if product_id
    ]
    if len(accounted) != len(set(accounted)):
        raise RuntimeError("Offer accounting failed: an input row is represented by more than one output offer")
    if set(accounted) != expected:
        missing = sorted(expected - set(accounted))
        extra = sorted(set(accounted) - expected)
        raise RuntimeError(f"Offer accounting failed: missing={missing[:5]} extra={extra[:5]}")


def exact_commercial_offer_key(product: dict[str, Any]) -> tuple[str, ...] | None:
    price = duplicate_key_number(get(product, "price"))
    valid_from = str(get(product, "valid_from")).strip()
    valid_to = str(get(product, "valid_to")).strip()
    if not price or not (valid_from or valid_to):
        return None
    return (
        normalize_supplier(get(product, "supplier_norm") or get(product, "supplier")),
        exact_product_name_key(product),
        price,
        duplicate_key_number(get(product, "quantity")),
        duplicate_key_text(get(product, "unit")),
        duplicate_key_number(get(product, "price_per_kg")),
        duplicate_key_price_tiers(get(product, "price_tiers")),
        valid_from,
        valid_to,
    )


def same_source_commercial_offer_key(product: dict[str, Any]) -> tuple[str, ...] | None:
    source_file = duplicate_key_text(get(product, "source_file"))
    source_page = duplicate_key_text(get(product, "source_page"))
    price = duplicate_key_number(get(product, "price"))
    valid_from = str(get(product, "valid_from")).strip()
    valid_to = str(get(product, "valid_to")).strip()
    if not all([source_file, source_page, price]) or not (valid_from or valid_to):
        return None
    return (
        normalize_supplier(get(product, "supplier_norm") or get(product, "supplier")),
        exact_product_name_key(product),
        source_file,
        source_page,
        price,
        valid_from,
        valid_to,
    )


def compatible_offer_extraction_representations(
    left: dict[str, Any],
    right: dict[str, Any],
) -> bool:
    optional_fields = [
        (duplicate_key_number(get(left, "quantity")), duplicate_key_number(get(right, "quantity"))),
        (duplicate_key_text(get(left, "unit")), duplicate_key_text(get(right, "unit"))),
        (duplicate_key_number(get(left, "price_per_kg")), duplicate_key_number(get(right, "price_per_kg"))),
        (duplicate_key_price_tiers(get(left, "price_tiers")), duplicate_key_price_tiers(get(right, "price_tiers"))),
    ]
    return all(not left_value or not right_value or left_value == right_value for left_value, right_value in optional_fields)


def merge_offer_representations(products: list[dict[str, Any]]) -> dict[str, Any]:
    ranked = sorted(products, key=offer_representation_rank, reverse=True)
    merged = dict(ranked[0])
    for field in ["product", "product_name", "brand", "description", "origin"]:
        best = choose_best_source_text([get(product, field) for product in products])
        if best:
            merged[field] = best
    amount_source = max(products, key=amount_representation_rank)
    if get(amount_source, "quantity") or get(amount_source, "unit"):
        merged["quantity"] = get(amount_source, "quantity")
        merged["unit"] = get(amount_source, "unit")
    merged["_merged_product_ids"] = [product.get("product_id", "") for product in products]
    merged["_merged_sources"] = [source_label(product) for product in products]
    return merged


def offer_representation_rank(product: dict[str, Any]) -> tuple[int, int, int, int, int]:
    confidence = int(to_float(product.get("extraction_confidence")) or 0)
    return (
        bool(get(product, "description")),
        len(duplicate_key_text(get(product, "description")).split()),
        bool(get(product, "origin")),
        bool(get(product, "brand")),
        amount_representation_rank(product)[0] * 1000 + confidence,
    )


def amount_representation_rank(product: dict[str, Any]) -> tuple[int, int, int]:
    quantity = get(product, "quantity")
    unit = get(product, "unit")
    return (
        int(bool(quantity)) + int(bool(unit)),
        int(to_float(quantity) is not None),
        len(duplicate_key_text(unit)),
    )


def choose_best_source_text(values: list[Any]) -> str:
    candidates = [str(value).strip() for value in values if str(value or "").strip()]
    if not candidates:
        return ""
    return max(
        candidates,
        key=lambda value: (
            len(set(duplicate_key_text(value).split())),
            len(duplicate_key_text(value)),
            value,
        ),
    )


def offer_sort_key(product: dict[str, Any]) -> tuple:
    return (
        parse_date(get(product, "valid_from")) or date.min,
        parse_date(get(product, "valid_to")) or date.min,
        to_float(product.get("price")) if to_float(product.get("price")) is not None else math.inf,
        amount_label(product),
    )


def choose_canonical_name(cluster: dict[str, Any], attrs: list[dict[str, Any]]) -> str:
    best_pair = None
    for edge in cluster["pair_edges"]:
        if edge.get("canonical_name") and (
            best_pair is None
            or normalize_confidence(edge.get("confidence")) > normalize_confidence(best_pair.get("confidence"))
        ):
            best_pair = edge
    if best_pair:
        return best_pair["canonical_name"]
    attr = choose_primary_attr(attrs)
    parts = [
        attr.get("base_product"), attr.get("variant"), attr.get("processing"),
        attr.get("quality_class"), attr.get("origin"), attr.get("packaging"),
        quantity_label(attr),
    ]
    return title_case(" ".join(str(p) for p in parts if p))


def choose_primary_attr(attrs: list[dict[str, Any]]) -> dict[str, Any]:
    return sorted(attrs, key=lambda a: normalize_confidence(a.get("attribute_confidence")), reverse=True)[0]


def title_case(text: str) -> str:
    return " ".join(w[:1].upper() + w[1:] for w in text.split())


def quantity_label(attr: dict[str, Any]) -> str:
    value = attr.get("quantity_value")
    unit = attr.get("quantity_unit")
    return f"{value} {unit}" if value not in {None, ""} and unit else ""


def most_common(values: list[str]) -> str:
    values = [v for v in values if v]
    return Counter(values).most_common(1)[0][0] if values else ""


def cluster_confidence(edges: list[dict[str, Any]]) -> int:
    if not edges:
        return 100
    return int(round(sum(normalize_confidence(e.get("confidence")) for e in edges) / len(edges)))


def cluster_review_reasons(products: list[dict[str, Any]], attrs: list[dict[str, Any]], edges: list[dict[str, Any]], suppliers: dict[str, list[dict[str, Any]]]) -> tuple[bool, list[str]]:
    reasons = []
    if any(e.get("decision") == "close_comparable" and e.get("should_human_review") for e in edges):
        reasons.append("close comparable match needs review")
    if any(normalize_confidence(e.get("confidence")) < 80 and e.get("should_human_review") for e in edges):
        reasons.append("match confidence below exact threshold")
    if any(e.get("should_human_review") for e in edges):
        reasons.append("pair judge requested review")
    conflicts = [
        e.get("conflicting_attributes")
        for e in edges
        if e.get("conflicting_attributes") and e.get("should_human_review")
    ]
    if conflicts:
        reasons.append("conflicting attributes present")
    return bool(reasons), reasons


def make_review_row(cluster: dict[str, Any], canonical_name: str, products: list[dict[str, Any]], reasons: list[str], edges: list[dict[str, Any]]) -> dict[str, Any]:
    conflicts = []
    llm_reasons = []
    for edge in edges:
        if edge.get("conflicting_attributes"):
            conflicts.append(edge["conflicting_attributes"])
        if edge.get("reason"):
            llm_reasons.append(edge["reason"])
    return {
        "review_id": f"r_{cluster['canonical_product_id']}",
        "issue_type": "; ".join(reasons[:3]),
        "canonical_product_id": cluster["canonical_product_id"],
        "canonical_product_name": canonical_name,
        "suppliers": ", ".join(sorted({p["supplier_norm"] for p in products})),
        "product_names": "; ".join(display_product_name(p) for p in products),
        "prices": "; ".join(str(p.get("price", "")) for p in products),
        "price_per_kg_values": "; ".join(str(p.get("price_per_kg", "")) for p in products),
        "issue_summary": "; ".join(reasons),
        "recommended_action": "Review manually before using this cluster for price decisions.",
        "llm_reason": " | ".join(llm_reasons),
        "conflicting_attributes": " | ".join(conflicts),
        "product_ids": "; ".join(p["product_id"] for p in products),
    }


def format_offer_cell(product: dict[str, Any]) -> str:
    lines = [
        f"Preis: {price_label(product)}",
        f"Staffelpreise: {price_tiers_label(product)}",
        f"Preis/kg: {price_per_kg_label(product)}",
        f"Gültig: {valid_label(product)}",
        f"Menge: {amount_label(product)}",
        f"Produkt: {display_product_name(product) or 'unknown'}",
        f"Marke: {product.get('brand') or 'unknown'}",
        f"Beschreibung: {wrap_excel_text(product.get('description'), FINAL_TEXT_WRAP_WIDTH) or 'unknown'}",
        f"Herkunft: {product.get('origin') or 'unknown'}",
        f"Quelle: {source_label(product)}",
    ]
    return "\n".join(line for line in lines if not line.endswith("unknown") and not line.endswith(": "))


def format_offer_cell_short(product: dict[str, Any]) -> str:
    price = f"{format_german_number(product.get('price'))} €" if product.get("price") else ""
    amount = amount_label(product)
    valid = short_valid_label(product)
    tiers = price_tiers_label(product)
    parts = [
        price,
        amount if amount != "unknown" else "",
        tiers.replace(" EUR", " €") if tiers != "unknown" else "",
        valid,
    ]
    return "\n".join(part for part in parts if part and part != "unknown")


def display_product_name(product: dict[str, Any]) -> str:
    value = product.get("product") or product.get("product_name") or ""
    return str(value).strip()


def amount_label(product: dict[str, Any]) -> str:
    quantity, unit = product.get("quantity"), product.get("unit")
    quantity_text = clean_amount_value(quantity)
    unit_text = clean_amount_value(unit)
    return f"{quantity_text} {unit_text}".strip() if quantity_text or unit_text else "unknown"


def clean_amount_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if not text or text.casefold() in {"none", "nan", "null"}:
        return ""
    if re.fullmatch(r"\d+(?:[.,]\d+)?", text):
        return format_german_number(text)
    return text


def price_label(product: dict[str, Any]) -> str:
    price = product.get("price")
    return f"{format_german_number(price)} EUR" if price else "unknown"


def price_per_kg_label(product: dict[str, Any]) -> str:
    price = product.get("price_per_kg")
    return f"{format_german_number(price)} EUR" if price else "unknown"


def price_tiers_label(product: dict[str, Any]) -> str:
    raw = product.get("price_tiers")
    if raw is None:
        return "unknown"
    text = str(raw).strip()
    if not text:
        return "unknown"
    try:
        tiers = json.loads(text)
    except Exception:
        return humanize_price_tiers_text(text)
    if not isinstance(tiers, list) or not tiers:
        return "unknown"
    parts = []
    for tier in tiers:
        if not isinstance(tier, dict):
            continue
        qty = tier.get("min_qty")
        price = tier.get("price")
        if qty in {None, ""} or price in {None, ""}:
            continue
        parts.append(f"ab {format_german_number(qty)}: {format_german_number(price)} EUR")
    return "; ".join(parts) if parts else "unknown"


def humanize_price_tiers_text(text: str) -> str:
    text = text.strip()
    text = re.sub(r'["{}\[\]]+', "", text)
    text = text.replace("min_qty", "ab").replace("price", "Preis")
    text = text.replace(":", ": ").replace(",", ";")
    text = re.sub(r"\s+", " ", text)
    return text


def format_german_number(value: Any) -> str:
    try:
        number = float(str(value).strip().replace(",", "."))
    except Exception:
        return str(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:.2f}".rstrip("0").rstrip(".").replace(".", ",")


def valid_label(product: dict[str, Any]) -> str:
    start, end = product.get("valid_from"), product.get("valid_to")
    return f"{start} to {end}" if start or end else "unknown"


def short_valid_label(product: dict[str, Any]) -> str:
    start = short_date_label(product.get("valid_from"))
    end = short_date_label(product.get("valid_to"))
    if start and end:
        return f"{start} bis {end}"
    if start:
        return f"ab {start}"
    if end:
        return f"bis {end}"
    return ""


def short_date_label(value: Any) -> str:
    parsed = parse_date(value)
    if parsed:
        return parsed.strftime("%d.%m.")
    text = str(value or "").strip()
    return text[:10] if text else ""


def source_label(product: dict[str, Any]) -> str:
    source = Path(str(product.get("source_file") or "")).name
    page = product.get("source_page")
    return f"{source} page {page}".strip() if source or page else "unknown"


def wrap_excel_text(value: Any, width: int) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    if not text:
        return ""
    return "\n".join(textwrap.wrap(text, width=width, break_long_words=False, break_on_hyphens=False))


def build_attribute_debug_rows(df: pd.DataFrame, attributes: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for product in df.to_dict("records"):
        attr = attributes[product["product_id"]]
        row = {**product}
        row.pop("rich_product_text", None)
        for key, value in attr.items():
            row[f"attr_{key}" if key in row else key] = json.dumps(value, ensure_ascii=False) if isinstance(value, list) else value
        rows.append(row)
    return rows


LIST_RED = "D52B1E"
LIST_GREEN = "7AB800"
LIST_YELLOW = "F7BC60"
LIST_DARK = "27321F"
LIST_LIGHT = "F6FAF0"
LIST_LINE = "D9E6D0"
FINAL_OUTPUT_VERTICAL_SEPARATOR_COLUMNS = {"Herkunft", "Edeka", "VK"}
FINAL_TEXT_WRAP_WIDTH = 30
FINAL_OUTPUT_ROW_MIN_HEIGHT = 22
FINAL_OUTPUT_ROW_LINE_HEIGHT = 15
FINAL_OUTPUT_ROW_PADDING = 8
FINAL_OUTPUT_COMPACT_ROW_MAX_HEIGHT = 360
FINAL_OUTPUT_FULL_ROW_MAX_HEIGHT = 220
FINAL_OUTPUT_NUMBER_FORMATS = {
    "EK": '#,##0.000 "€"',
    "VK": '#,##0.00 "€"',
}
CATEGORY_LABELS = {
    "fisch": "Fisch",
    "fleisch": "Fleisch",
    "obst_gemuese": "Obst & Gemüse",
    "obst_gemüse": "Obst & Gemüse",
    "obst & gemüse": "Obst & Gemüse",
    "obst und gemüse": "Obst & Gemüse",
    "tk": "TK",
    "wurst": "Wurst",
    "mopro": "Mopro",
    "sonstiges": "Sonstiges",
}
COUNTRY_CODE_ALIASES = {
    "AL": ["albanien", "albania"],
    "AD": ["andorra"],
    "AM": ["armenien", "armenia"],
    "AT": ["oesterreich", "osterreich", "austria"],
    "AZ": ["aserbaidschan", "azerbaijan"],
    "BY": ["belarus", "weissrussland", "weissrussland", "weißrussland"],
    "BE": ["belgien", "belgium"],
    "BA": ["bosnien", "bosnien herzegowina", "bosnia", "bosnia and herzegovina"],
    "BG": ["bulgarien", "bulgaria"],
    "HR": ["kroatien", "croatia"],
    "CY": ["zypern", "cyprus"],
    "CZ": ["tschechien", "tschechische republik", "czech republic", "czechia"],
    "DK": ["daenemark", "danemark", "dänemark", "denmark"],
    "EE": ["estland", "estonia"],
    "FI": ["finnland", "finland"],
    "FR": ["frankreich", "france"],
    "GE": ["georgien", "georgia"],
    "DE": ["deutschland", "germany", "bundesrepublik deutschland"],
    "GR": ["griechenland", "greece"],
    "HU": ["ungarn", "hungary"],
    "IS": ["island", "iceland"],
    "IE": ["irland", "ireland"],
    "IT": ["italien", "italy"],
    "KZ": ["kasachstan", "kazakhstan"],
    "XK": ["kosovo"],
    "LV": ["lettland", "latvia"],
    "LI": ["liechtenstein"],
    "LT": ["litauen", "lithuania"],
    "LU": ["luxemburg", "luxembourg"],
    "MT": ["malta"],
    "MD": ["moldau", "moldawien", "moldova"],
    "MC": ["monaco"],
    "ME": ["montenegro"],
    "NL": ["niederlande", "holland", "netherlands"],
    "MK": ["nordmazedonien", "mazedonien", "north macedonia", "macedonia"],
    "NO": ["norwegen", "norway"],
    "PL": ["polen", "poland"],
    "PT": ["portugal"],
    "RO": ["rumaenien", "rumanien", "rumänien", "romania"],
    "RU": ["russland", "russia", "russian federation"],
    "SM": ["san marino"],
    "RS": ["serbien", "serbia"],
    "SK": ["slowakei", "slovakia"],
    "SI": ["slowenien", "slovenia"],
    "ES": ["spanien", "spain"],
    "SE": ["schweden", "sweden"],
    "CH": ["schweiz", "switzerland"],
    "TR": ["tuerkei", "turkei", "türkei", "turkey"],
    "UA": ["ukraine"],
    "GB": ["grossbritannien", "großbritannien", "vereinigtes koenigreich", "vereinigtes konigreich", "vereinigtes königreich", "uk", "united kingdom", "great britain", "england", "schottland", "scotland", "wales", "nordirland", "northern ireland"],
    "VA": ["vatikan", "vatican"],
    "US": ["usa", "us", "vereinigte staaten", "united states", "united states of america", "amerika"],
    "CA": ["kanada", "canada"],
    "MX": ["mexiko", "mexico"],
    "BR": ["brasilien", "brazil"],
    "AR": ["argentinien", "argentina"],
    "CL": ["chile"],
    "PE": ["peru"],
    "EC": ["ecuador"],
    "CO": ["kolumbien", "colombia"],
    "CR": ["costa rica"],
    "UY": ["uruguay"],
    "PY": ["paraguay"],
    "BO": ["bolivien", "bolivia"],
    "GT": ["guatemala"],
    "HN": ["honduras"],
    "NI": ["nicaragua"],
    "PA": ["panama"],
    "DO": ["dominikanische republik", "dominican republic"],
    "CN": ["china", "volksrepublik china"],
    "IN": ["indien", "india"],
    "PK": ["pakistan"],
    "BD": ["bangladesch", "bangladesh"],
    "LK": ["sri lanka"],
    "TH": ["thailand"],
    "VN": ["vietnam", "viet nam"],
    "ID": ["indonesien", "indonesia"],
    "MY": ["malaysia"],
    "PH": ["philippinen", "philippines"],
    "TW": ["taiwan"],
    "JP": ["japan"],
    "KR": ["suedkorea", "sudkorea", "südkorea", "south korea"],
    "AU": ["australien", "australia"],
    "NZ": ["neuseeland", "new zealand"],
    "ZA": ["suedafrika", "sudafrika", "südafrika", "south africa"],
    "MA": ["marokko", "morocco"],
    "TN": ["tunesien", "tunisia"],
    "EG": ["aegypten", "agypten", "ägypten", "egypt"],
    "KE": ["kenia", "kenya"],
    "ET": ["aethiopien", "athiopien", "äthiopien", "ethiopia"],
    "GH": ["ghana"],
    "CI": ["elfenbeinkueste", "elfenbeinkuste", "elfenbeinküste", "cote d ivoire", "ivory coast"],
    "SN": ["senegal"],
    "NG": ["nigeria"],
    "IL": ["israel"],
    "JO": ["jordanien", "jordan"],
    "LB": ["libanon", "lebanon"],
    "AE": ["vereinigte arabische emirate", "united arab emirates", "uae"],
    "SA": ["saudi arabien", "saudi arabia"],
    "IR": ["iran"],
}


def normalize_country_alias(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").casefold())
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


VALID_COUNTRY_CODES = set(COUNTRY_CODE_ALIASES)
COUNTRY_CODE_BY_ALIAS = {
    normalized_alias: code
    for code, aliases in COUNTRY_CODE_ALIASES.items()
    for normalized_alias in [normalize_country_alias(alias) for alias in aliases + [code]]
}
COUNTRY_ALIAS_PATTERNS = sorted(COUNTRY_CODE_BY_ALIAS.items(), key=lambda item: len(item[0]), reverse=True)


def write_excel(
    output_path: Path,
    matched_rows: list[dict[str, Any]],
    review_rows: list[dict[str, Any]],
    pair_rows: list[dict[str, Any]],
    attribute_rows: list[dict[str, Any]],
    logo_path: Path | None = None,
    offer_grouping_audit: list[dict[str, Any]] | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True) if output_path.parent != Path(".") else None
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Output"
    write_final_output_sheet(ws, build_final_output_rows(matched_rows), output_path, logo_path)
    week_label = output_week_label(output_path)
    write_final_output_sheet(
        wb.create_sheet("Final Output Short"),
        build_final_output_rows(matched_rows, short=True),
        output_path,
        logo_path,
        title=f"Wilhelm LIST Nachfolger | Wettbewerbsvergleich {week_label}",
        subtitle="Kompakte Angebotsansicht: Preis, Gültigkeit, Menge und Staffelpreise",
        table_name="Final_Output_Short_tbl",
        compact_rows=True,
    )
    write_sheet(wb.create_sheet("matched_products"), matched_rows, MATCHED_COLUMNS)
    write_sheet(wb.create_sheet("review_queue"), review_rows, REVIEW_COLUMNS)
    write_sheet(wb.create_sheet("pair_debug"), pair_rows, PAIR_COLUMNS)
    write_sheet(
        wb.create_sheet("offer_grouping_audit"),
        offer_grouping_audit or [],
        OFFER_GROUPING_AUDIT_COLUMNS,
    )
    attr_columns = list(attribute_rows[0].keys()) if attribute_rows else []
    write_sheet(wb.create_sheet("attribute_debug"), attribute_rows, attr_columns)
    wb.save(output_path)


def write_offer_grouping_audit(output_path: Path, audit_rows: list[dict[str, Any]]) -> Path:
    audit_path = output_path.with_name(f"{output_path.stem}_offer_grouping_audit.jsonl")
    with audit_path.open("w", encoding="utf-8") as handle:
        for row in audit_rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
    print(f"Offer grouping audit written: {audit_path}")
    return audit_path


def output_week_label(output_path: Path) -> str:
    for value in [output_path.stem, output_path.parent.name]:
        match = re.search(r"\bKW\s*(\d{1,2})\b", value, flags=re.IGNORECASE)
        if not match:
            match = re.search(r"KW(\d{1,2})", value, flags=re.IGNORECASE)
        if match:
            return f"KW{int(match.group(1)):02d}"
    return "KW"


def build_final_output_rows(matched_rows: list[dict[str, Any]], short: bool = False) -> list[dict[str, Any]]:
    rows = []
    for row in matched_rows:
        suffix = "_short" if short else ""
        rows.append({
            "Kategorie": category_label(row.get("category", "")),
            "Produkt": product_name_from_offer_cells(row),
            "_sort_brand": row.get("brand", ""),
            "Beschreibung": final_output_description(row.get("description", ""), row.get("brand", "")),
            "Herkunft": origin_country_codes(row.get("origin", "")),
            "Metro": row.get(f"Metro{suffix}", ""),
            "Selgros": row.get(f"Selgros{suffix}", ""),
            "Handelshof": row.get(f"Handelshof{suffix}", ""),
            "Edeka": row.get(f"Edeka{suffix}", ""),
            "EK": "",
            "VK": "",
            "Notizen": "",
        })
    return sorted(rows, key=final_output_sort_key)


def final_output_description(description: Any, brand: Any) -> str:
    lines = []
    description_text = wrap_excel_text(description, FINAL_TEXT_WRAP_WIDTH)
    brand_text = str(brand or "").strip()
    if description_text:
        lines.append(description_text)
    if brand_text:
        lines.append(wrap_excel_text(f"Brand: {brand_text},", FINAL_TEXT_WRAP_WIDTH))
    return "\n".join(line for line in lines if line)


def origin_country_codes(origin: Any) -> str:
    text = str(origin or "").strip()
    if not text:
        return ""
    codes: list[str] = []
    parts = [
        part.strip()
        for part in re.split(r"[,;/|+&\n]+|\b(?:und|oder|and|or)\b", text, flags=re.IGNORECASE)
        if part.strip()
    ]
    for part in parts or [text]:
        add_origin_codes(part, codes)
    if not codes:
        add_origin_codes(text, codes)
    return ", ".join(codes) if codes else text


def add_origin_codes(text: str, codes: list[str]) -> None:
    stripped = text.strip()
    if re.fullmatch(r"[A-Za-z]{2}", stripped):
        code = stripped.upper()
        if code in VALID_COUNTRY_CODES and code not in codes:
            codes.append(code)
        return
    normalized = normalize_country_alias(stripped)
    if not normalized:
        return
    exact_code = COUNTRY_CODE_BY_ALIAS.get(normalized)
    if exact_code:
        if exact_code not in codes:
            codes.append(exact_code)
        return
    for alias, code in COUNTRY_ALIAS_PATTERNS:
        if re.search(rf"(^| ){re.escape(alias)}($| )", normalized) and code not in codes:
            codes.append(code)


def product_name_from_offer_cells(row: dict[str, Any]) -> str:
    for supplier in SUPPLIER_ORDER:
        offer_text = str(row.get(supplier, "") or "")
        for line in offer_text.splitlines():
            if line.startswith("Produkt:"):
                product_name = line.removeprefix("Produkt:").strip()
                if product_name:
                    return product_name
    return str(row.get("product") or row.get("canonical_product_name") or "").strip()


def category_label(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    key = text.casefold().replace("-", "_").replace(" ", "_")
    if key in CATEGORY_LABELS:
        return CATEGORY_LABELS[key]
    return " ".join(part[:1].upper() + part[1:].lower() for part in text.replace("_", " ").split())


def category_order_index(value: Any) -> int:
    text = str(value or "").casefold()
    if text == "fisch":
        return 0
    if text == "fleisch":
        return 1
    if "obst" in text and ("gemüse" in text or "gemuese" in text):
        return 2
    return 10


def final_output_sort_key(row: dict[str, Any]) -> tuple:
    category = str(row.get("Kategorie", ""))
    return (
        category_order_index(category),
        category.casefold(),
        str(row.get("Produkt", "")).casefold(),
        str(row.get("_sort_brand", "")).casefold(),
    )


def final_output_cell_border(
    column: str,
    top: Side | None = None,
    bottom: Side | None = None,
) -> Border:
    return Border(
        top=top or Side(),
        bottom=bottom or Side(),
        right=Side(style="medium", color=LIST_GREEN) if column in FINAL_OUTPUT_VERTICAL_SEPARATOR_COLUMNS else Side(),
    )


def write_final_output_sheet(
    ws,
    rows: list[dict[str, Any]],
    output_path: Path,
    logo_path: Path | None,
    title: str = "LIST Goslar | Wettbewerbsvergleich",
    subtitle: str = "Automatisierter Artikel- und Preisvergleich aus aktuellen Wettbewerbsprospekten",
    table_name: str = "Final_Output_tbl",
    compact_rows: bool = False,
) -> None:
    ws.sheet_view.showGridLines = False
    for row_idx in range(1, 7):
        for col_idx in range(1, len(FINAL_OUTPUT_COLUMNS) + 1):
            ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=LIST_LIGHT)

    add_logo(ws, logo_path)
    last_col = get_column_letter(len(FINAL_OUTPUT_COLUMNS))
    ws.merge_cells(f"D1:{last_col}1")
    ws.merge_cells(f"D2:{last_col}2")
    ws.merge_cells(f"D3:{last_col}3")
    ws["D1"] = title
    ws["D1"].font = Font(bold=True, size=20, color=LIST_DARK)
    ws["D2"] = subtitle
    ws["D2"].font = Font(size=11, color="5D6B54")
    ws["D3"] = f"Erstellt am {time.strftime('%d.%m.%Y %H:%M')} | Datei: {output_path.name} | Artikelgruppen: {len(rows)}"
    ws["D3"].font = Font(size=9, color="6E7867")

    for col_idx, column in enumerate(FINAL_OUTPUT_COLUMNS[:3], 1):
        ws.cell(5, col_idx, column)
    for cell in ws[5]:
        cell.font = Font(bold=True, color=LIST_DARK)
        cell.fill = PatternFill("solid", fgColor="EAF3DF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[5].height = 22

    header_row = 7
    columns = FINAL_OUTPUT_COLUMNS
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(header_row, col_idx, column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=LIST_DARK)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = final_output_cell_border(
            column,
            bottom=Side(style="thin", color=LIST_GREEN),
        )
    for row_idx, row in enumerate(rows, header_row + 1):
        previous_category = rows[row_idx - header_row - 2].get("Kategorie") if row_idx > header_row + 1 else None
        current_category = row.get("Kategorie")
        top_side = Side(style="medium", color=LIST_GREEN) if previous_category is not None and current_category != previous_category else Side(style="hair", color=LIST_LINE)
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row_idx, col_idx, row.get(column, ""))
            cell.fill = PatternFill("solid", fgColor="FFFFFF" if row_idx % 2 else "FAFCF7")
            cell.border = final_output_cell_border(
                column,
                top=top_side,
                bottom=Side(style="hair", color=LIST_LINE),
            )
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if column in FINAL_OUTPUT_NUMBER_FORMATS:
                cell.number_format = FINAL_OUTPUT_NUMBER_FORMATS[column]
    ws.freeze_panes = f"A{header_row + 1}"
    if rows:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(rows)}"
        table = Table(displayName=table_name, ref=f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(rows)}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=False)
        ws.add_table(table)
    widths = final_output_column_widths(compact_rows)
    for idx, column in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(column, 20)
    for row_idx in range(header_row + 1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = final_output_row_height(ws, row_idx, columns, widths, compact_rows)


def final_output_row_height(ws, row_idx: int, columns: list[str], widths: dict[str, int], compact_rows: bool) -> float:
    max_lines = 1
    for col_idx, column in enumerate(columns, 1):
        value = str(ws.cell(row_idx, col_idx).value or "")
        if not value:
            continue
        max_lines = max(max_lines, estimated_excel_lines(value, widths.get(column, 20)))
    max_height = FINAL_OUTPUT_COMPACT_ROW_MAX_HEIGHT if compact_rows else FINAL_OUTPUT_FULL_ROW_MAX_HEIGHT
    calculated_height = FINAL_OUTPUT_ROW_LINE_HEIGHT * max_lines + FINAL_OUTPUT_ROW_PADDING
    return min(max_height, max(FINAL_OUTPUT_ROW_MIN_HEIGHT, calculated_height))


def final_output_column_widths(compact_rows: bool) -> dict[str, int]:
    if compact_rows:
        return {
            "Kategorie": 10,
            "Produkt": 18,
            "Beschreibung": 24,
            "Herkunft": 7,
            "Metro": 17,
            "Selgros": 17,
            "Handelshof": 17,
            "Edeka": 17,
            "EK": 9,
            "VK": 9,
            "Notizen": 16,
        }
    return {
        "Kategorie": 14,
        "Produkt": 26,
        "Beschreibung": 30,
        "Herkunft": 10,
        "Metro": 30,
        "Selgros": 30,
        "Handelshof": 30,
        "Edeka": 30,
        "EK": 10,
        "VK": 10,
        "Notizen": 24,
    }


def estimated_excel_lines(value: str, column_width: int) -> int:
    line_count = 0
    usable_width = max(8, column_width - 2)
    for line in value.splitlines() or [""]:
        line_count += max(1, math.ceil(len(line) / usable_width))
    return line_count


def add_logo(ws, logo_path: Path | None) -> None:
    image_path = prepare_logo_image(logo_path)
    if not image_path:
        ws["A1"] = "LIST"
        ws["A1"].font = Font(bold=True, size=22, color=LIST_RED)
        return
    try:
        img = OpenpyxlImage(str(image_path))
        # Excel image dimensions are set in pixels. At 96 DPI this is about
        # 2.22 inches wide by 1.1 inches high.
        img.width = 213
        img.height = 106
        ws.add_image(img, "A1")
    except Exception:
        ws["A1"] = "LIST"
        ws["A1"].font = Font(bold=True, size=22, color=LIST_RED)


def prepare_logo_image(logo_path: Path | None) -> Path | None:
    if not logo_path or not logo_path.exists():
        return None
    if logo_path.suffix.lower() in {".png", ".jpg", ".jpeg"}:
        return logo_path
    if logo_path.suffix.lower() != ".svg":
        return None
    png_path = Path(tempfile.gettempdir()) / f"{logo_path.stem}_{short_hash(str(logo_path))}.png"
    if png_path.exists() and png_path.stat().st_mtime >= logo_path.stat().st_mtime:
        return png_path
    try:
        subprocess.run(
            ["magick", str(logo_path), "-resize", "300x212", str(png_path)],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        return png_path if png_path.exists() else None
    except Exception:
        return None


MATCHED_COLUMNS = [
    "canonical_product_id", "canonical_product_name", "category", "product", "brand", "description", "base_product",
    "variant", "processing", "quality_class", "origin", "packaging", "unit_basis",
    "match_confidence", "review_needed", "match_reason",
    "Metro", "Selgros", "Handelshof", "Edeka",
]

FINAL_OUTPUT_COLUMNS = [
    "Kategorie", "Produkt", "Beschreibung", "Herkunft",
    "Metro", "Selgros", "Handelshof", "Edeka", "EK", "VK", "Notizen",
]

REVIEW_COLUMNS = [
    "review_id", "issue_type", "canonical_product_id", "canonical_product_name",
    "suppliers", "product_names", "prices", "price_per_kg_values", "issue_summary",
    "recommended_action", "llm_reason", "conflicting_attributes", "product_ids",
]

PAIR_COLUMNS = [
    "product_id_a", "product_id_b", "supplier_a", "supplier_b", "name_a", "name_b",
    "similarity", "hard_blocked", "block_reasons", "soft_warnings", "decision",
    "confidence", "canonical_name", "reason", "conflicting_attributes",
    "should_human_review", "second_judge_action", "second_judge_confidence", "second_judge_reason",
    "cluster_merge_blocked", "cluster_block_reasons",
]

OFFER_GROUPING_AUDIT_COLUMNS = [
    "cluster_id", "supplier", "rule", "kept_product_id", "merged_product_ids",
    "sources", "price", "valid_from", "valid_to", "final_amount",
]


def write_sheet(ws, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([excel_cell_value(row.get(col, "")) for col in columns])
    style_sheet(ws, len(rows), len(columns))


def excel_cell_value(value: Any) -> Any:
    if isinstance(value, (list, dict, tuple, set)):
        return json.dumps(list(value) if isinstance(value, set) else value, ensure_ascii=False)
    return value


def style_sheet(ws, row_count: int, col_count: int) -> None:
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor=LIST_DARK)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(bottom=Side(style="thin", color=LIST_GREEN))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if row_count > 0 and col_count > 0:
        ref = f"A1:{get_column_letter(col_count)}{row_count + 1}"
        table = Table(displayName=safe_table_name(ws.title), ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        try:
            ws.add_table(table)
        except ValueError:
            pass
    for idx in range(1, col_count + 1):
        header = str(ws.cell(1, idx).value or "")
        max_len = len(header)
        for row in range(2, min(ws.max_row, 80) + 1):
            max_len = max(max_len, min(len(str(ws.cell(row, idx).value or "")), 70))
        width = 32 if header in SUPPLIER_ORDER else min(max(max_len + 2, 10), 42)
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row_idx in range(2, ws.max_row + 1):
        fill = PatternFill("solid", fgColor="FFFFFF" if row_idx % 2 else "FAFCF7")
        for col_idx in range(1, col_count + 1):
            ws.cell(row_idx, col_idx).fill = fill
            ws.cell(row_idx, col_idx).border = Border(bottom=Side(style="hair", color=LIST_LINE))
    wrap_columns = set(SUPPLIER_ORDER + ["match_reason", "issue_summary", "llm_reason", "conflicting_attributes", "reason", "block_reasons", "soft_warnings"])
    for col_idx in range(1, col_count + 1):
        if ws.cell(1, col_idx).value in wrap_columns:
            for row in range(1, ws.max_row + 1):
                ws.cell(row, col_idx).alignment = Alignment(wrap_text=True, vertical="top")


def safe_table_name(name: str) -> str:
    return re.sub(r"\W+", "_", name).strip("_")[:25] + "_tbl"


if __name__ == "__main__":
    main()
