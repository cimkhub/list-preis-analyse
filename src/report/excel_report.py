import logging
import time
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.harmonize.categories import CATEGORY_ORDER, category_label, category_sort_key
from src.models import MatchedProduct, RawProduct
from src.utils.logging_setup import log_event

logger = logging.getLogger("birkenhof.report")

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
CATEGORY_FILL = PatternFill("solid", fgColor="D9E2F3")
CATEGORY_FONT = Font(bold=True, size=11, name="Arial")
DATA_FONT = Font(size=10, name="Arial")
BEST_FILL = PatternFill("solid", fgColor="C6EFCE")
WORST_FILL = PatternFill("solid", fgColor="FFC7CE")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_report(
    matched_products: list[MatchedProduct],
    unmatched_products: list[RawProduct],
    all_products: list[RawProduct],
    week: int,
    year: int,
    output_dir: str = "reports",
) -> str:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    filename = f"Artikelvergleich_KW{week:02d}_{year}.xlsx"
    filepath = output_path / filename
    started = time.perf_counter()

    wb = Workbook()

    _create_comparison_sheet(wb, matched_products, week, year)
    _create_details_sheet(wb, all_products)
    _create_unmatched_sheet(wb, unmatched_products)
    _create_summary_sheet(wb, matched_products, all_products, week, year)

    wb.save(str(filepath))
    logger.info(f"Report saved to {filepath}")
    log_event(
        logger,
        f"Excel report saved to {filepath}",
        event="excel_report",
        status="ok",
        output_path=str(filepath),
        week=week,
        year=year,
        matched_count=len(matched_products),
        unmatched_count=len(unmatched_products),
        total_products=len(all_products),
        duration_ms=round((time.perf_counter() - started) * 1000, 2),
    )
    return str(filepath)


def _create_comparison_sheet(wb: Workbook, products: list[MatchedProduct], week: int, year: int):
    ws = wb.active
    ws.title = "Vergleich"

    # Title
    ws.merge_cells("A1:N1")
    ws["A1"] = f"Artikelvergleich KW{week:02d}/{year}"
    ws["A1"].font = Font(bold=True, size=14, name="Arial")

    # Headers (matching reference Excel format)
    headers = [
        ("A", "Kategorie", 12),
        ("B", "Artikel", 40),
        ("C", "Metro", 18),
        ("D", "Metro Info", 15),
        ("E", "EDEKA Foodservice", 18),
        ("F", "EDEKA Info", 15),
        ("G", "Selgros", 12),
        ("H", "Selgros Zeitraum", 16),
        ("I", "Handelshof", 12),
        ("J", "Handelshof Zeitraum", 18),
        ("K", "Preis List EK", 13),
        ("L", "Preis List VK", 13),
        ("M", "List Angebot", 15),
        ("N", "Art-Nr + Bezeichnung", 22),
    ]

    for col_letter, header, width in headers:
        cell = ws[f"{col_letter}4"]
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[col_letter].width = width

    # Supplier sub-headers in row 5
    supplier_row = {
        "C": "in der Werbung bei:              bitte Preis eintragen",
    }
    for col, text in supplier_row.items():
        ws[f"{col}3"] = text
        ws[f"{col}3"].font = Font(italic=True, size=9, name="Arial")

    # Sort by category then name
    sorted_products = sorted(
        products,
        key=lambda p: (category_sort_key(p.category), p.canonical_name)
    )

    row = 5
    current_category = None

    for product in sorted_products:
        if product.category != current_category:
            current_category = product.category

        ws[f"A{row}"] = category_label(product.category)
        ws[f"A{row}"].font = CATEGORY_FONT

        ws[f"B{row}"] = product.canonical_name
        ws[f"B{row}"].font = DATA_FONT

        ws[f"C{row}"] = product.metro_price
        ws[f"D{row}"] = product.metro_info
        ws[f"E{row}"] = product.edeka_price
        ws[f"F{row}"] = product.edeka_info
        ws[f"G{row}"] = product.selgros_price
        ws[f"H{row}"] = product.selgros_period
        ws[f"I{row}"] = product.handelshof_price
        ws[f"J{row}"] = product.handelshof_period

        # Apply formatting to all data cells
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            if cell.font == Font():
                cell.font = DATA_FONT

        # Highlight best/worst prices
        _highlight_prices(ws, row, product)

        row += 1


def _highlight_prices(ws, row: int, product: MatchedProduct):
    """Highlight best (green) and worst (red) prices."""
    prices = {}

    # Parse Metro price (take first value as comparable price)
    if product.metro_price:
        try:
            first_price = product.metro_price.split("/")[0].replace(",", ".")
            prices["C"] = float(first_price)
        except (ValueError, IndexError):
            pass

    if product.edeka_price:
        prices["E"] = product.edeka_price
    if product.selgros_price:
        prices["G"] = product.selgros_price
    if product.handelshof_price:
        prices["I"] = product.handelshof_price

    if len(prices) < 2:
        return

    min_price = min(prices.values())
    max_price = max(prices.values())

    for col, price in prices.items():
        cell = ws[f"{col}{row}"]
        if price == min_price:
            cell.fill = BEST_FILL
        elif price == max_price and max_price > min_price * 1.05:
            cell.fill = WORST_FILL


def _create_details_sheet(wb: Workbook, products: list[RawProduct]):
    ws = wb.create_sheet("Details")

    headers = [
        "Lieferant", "Produkt", "Beschreibung", "Kategorie", "Herkunft",
        "Einheit", "Menge", "Preis", "Preis/kg", "Netto?",
        "Brutto", "Gültig von", "Gültig bis", "KW", "Quelle",
        "Seite", "Confidence"
    ]

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER

    for row_idx, p in enumerate(products, 2):
        data = [
            p.supplier, p.product_name, p.description, p.category, p.origin,
            p.unit, p.quantity, p.price, p.price_per_kg, p.price_is_net,
            p.price_gross,
            p.valid_from.isoformat() if p.valid_from else None,
            p.valid_to.isoformat() if p.valid_to else None,
            p.calendar_week, p.source_file, p.source_page,
            p.extraction_confidence,
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = BORDER

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _create_unmatched_sheet(wb: Workbook, products: list[RawProduct]):
    ws = wb.create_sheet("Nicht zugeordnet")

    headers = ["Lieferant", "Produkt", "Kategorie", "Preis", "Einheit"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER

    for row_idx, p in enumerate(products, 2):
        ws.cell(row=row_idx, column=1, value=p.supplier).font = DATA_FONT
        ws.cell(row=row_idx, column=2, value=p.product_name).font = DATA_FONT
        ws.cell(row=row_idx, column=3, value=p.category).font = DATA_FONT
        ws.cell(row=row_idx, column=4, value=p.price).font = DATA_FONT
        ws.cell(row=row_idx, column=5, value=p.unit).font = DATA_FONT

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10


def _create_summary_sheet(wb: Workbook, matched: list[MatchedProduct],
                          all_products: list[RawProduct], week: int, year: int):
    ws = wb.create_sheet("Zusammenfassung")
    bold = Font(bold=True, size=11, name="Arial")

    ws["A1"] = f"Zusammenfassung KW{week:02d}/{year}"
    ws["A1"].font = Font(bold=True, size=14, name="Arial")

    ws["A3"] = "Produkte pro Lieferant"
    ws["A3"].font = bold

    supplier_counts = {}
    for p in all_products:
        supplier_counts[p.supplier] = supplier_counts.get(p.supplier, 0) + 1

    row = 4
    for supplier, count in sorted(supplier_counts.items()):
        ws[f"A{row}"] = supplier.title()
        ws[f"B{row}"] = count
        row += 1

    row += 1
    ws[f"A{row}"] = "Gematchte Produkte"
    ws[f"A{row}"].font = bold
    ws[f"B{row}"] = len(matched)

    row += 1
    ws[f"A{row}"] = "Nicht zugeordnet"
    ws[f"A{row}"].font = bold
    ws[f"B{row}"] = len(all_products) - sum(
        1 for m in matched
        for s in ["metro_price", "edeka_price", "selgros_price", "handelshof_price"]
        if getattr(m, s) is not None
    )

    # Top price differences
    row += 2
    ws[f"A{row}"] = "Top 10 - Größte Preisunterschiede"
    ws[f"A{row}"].font = bold
    row += 1

    diffs = []
    for m in matched:
        prices = []
        if m.edeka_price:
            prices.append(m.edeka_price)
        if m.selgros_price:
            prices.append(m.selgros_price)
        if m.handelshof_price:
            prices.append(m.handelshof_price)
        if m.metro_price:
            try:
                prices.append(float(m.metro_price.split("/")[0].replace(",", ".")))
            except (ValueError, IndexError):
                pass

        if len(prices) >= 2:
            spread = ((max(prices) - min(prices)) / min(prices)) * 100
            diffs.append((m.canonical_name, spread, min(prices), max(prices)))

    diffs.sort(key=lambda x: x[1], reverse=True)

    ws[f"A{row}"] = "Produkt"
    ws[f"B{row}"] = "Differenz %"
    ws[f"C{row}"] = "Min Preis"
    ws[f"D{row}"] = "Max Preis"
    for cell in [ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"], ws[f"D{row}"]]:
        cell.font = bold

    for product, spread, min_p, max_p in diffs[:10]:
        row += 1
        ws[f"A{row}"] = product
        ws[f"B{row}"] = round(spread, 1)
        ws[f"C{row}"] = round(min_p, 2)
        ws[f"D{row}"] = round(max_p, 2)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
