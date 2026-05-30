#!/usr/bin/env python3
"""Local web dashboard for running and reviewing the price-analysis pipeline."""

from __future__ import annotations

import csv
import hashlib
import hmac
import json
import mimetypes
import os
import re
import secrets
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import uuid
from datetime import datetime
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.parse import parse_qs, unquote, urlparse
import zipfile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WEB_ROOT = ROOT / "web_dashboard"
DATA_ROOT = Path(os.environ.get("DASHBOARD_DATA_ROOT") or os.environ.get("DATA_ROOT") or ROOT / "data").resolve()
PARSED_ROOT = Path(os.environ.get("DASHBOARD_PARSED_ROOT") or os.environ.get("PARSED_ROOT") or ROOT / "parsed").resolve()
IMAGE_ROOT = Path(os.environ.get("DASHBOARD_IMAGE_ROOT") or os.environ.get("IMAGE_ROOT") or ROOT / "images").resolve()
RUN_STORE = DATA_ROOT / "dashboard_runs.json"
CLIENT_LOGO = Path(os.environ.get("DASHBOARD_CLIENT_LOGO", "/Users/lukas/Downloads/logo_sb_list_mitUZ.svg"))
DATA_ZIP_URL = os.environ.get("DASHBOARD_DATA_ZIP_URL", "").strip()
RUNTIME_ROOT = Path(os.environ.get("DASHBOARD_RUNTIME_ROOT") or DATA_ROOT.parent).resolve()
QA_MARK_DEFAULT = {"is_mistake": False, "comment": "", "updated_at": ""}
DASHBOARD_PASSWORD = os.environ.get("DASHBOARD_PASSWORD", "List2026")
DASHBOARD_MODE = os.environ.get("DASHBOARD_MODE", "local").strip().casefold()
REVIEW_ONLY = DASHBOARD_MODE in {"review", "readonly", "read-only", "customer"}
AUTH_COOKIE_NAME = "list_dashboard_auth"
AUTH_SECRET = os.environ.get("DASHBOARD_AUTH_SECRET") or secrets.token_hex(32)
AUTH_TOKEN = hmac.new(AUTH_SECRET.encode("utf-8"), DASHBOARD_PASSWORD.encode("utf-8"), hashlib.sha256).hexdigest()
COOKIE_SECURE = str(os.environ.get("DASHBOARD_COOKIE_SECURE") or ("1" if os.environ.get("PORT") else "0")).strip().casefold() in {"1", "true", "yes", "ja"}
PIPELINE_STEPS = [
    ("queued", "Lauf vorgemerkt", "Der Wochenlauf wird vorbereitet."),
    ("acquisition", "Prospekte sammeln", "Aktuelle Wettbewerber-PDFs werden heruntergeladen."),
    ("document_relevance", "Prospektrelevanz prüfen", "Nur lebensmittel- und markt-relevante PDFs werden behalten."),
    ("extraction", "Produkte extrahieren", "Produktnamen, Preise, Zeiträume, Mengen und Quellen werden ausgelesen."),
    ("product_relevance", "Produktrelevanz prüfen", "Produktzeilen und Gründe werden klassifiziert."),
    ("matching", "Wettbewerber zuordnen", "Vergleichbare Produkte werden gefunden und die LIST-Excel wird erstellt."),
    ("report", "Berichte erstellen", "Excel-Dateien und finale Artefakte werden geschrieben."),
    ("completed", "Abgeschlossen", "Der Wochenvergleich ist bereit."),
]
ACTIVE_RUNS: dict[str, dict] = {}
ACTIVE_LOCK = threading.Lock()


def main() -> None:
    bootstrap_runtime_data()
    port = int(os.environ.get("PORT") or os.environ.get("DASHBOARD_PORT", "8765"))
    host = os.environ.get("DASHBOARD_HOST") or ("0.0.0.0" if os.environ.get("PORT") else "127.0.0.1")
    server = ThreadingHTTPServer((host, port), DashboardHandler)
    public_host = "127.0.0.1" if host == "0.0.0.0" else host
    mode = "Review-only" if REVIEW_ONLY else "Lokal"
    print(f"Dashboard läuft unter http://{public_host}:{port} ({mode})", flush=True)
    server.serve_forever()


def bootstrap_runtime_data() -> None:
    if runtime_data_ready() and not truthy_env("DASHBOARD_DATA_SYNC_FORCE"):
        print("Runtime data found; skipping data ZIP sync.", flush=True)
        return
    if not DATA_ZIP_URL:
        print("No DASHBOARD_DATA_ZIP_URL set; starting without startup data sync.", flush=True)
        return

    RUNTIME_ROOT.mkdir(parents=True, exist_ok=True)
    print(f"Runtime data missing; downloading data ZIP from {redact_url(DATA_ZIP_URL)}", flush=True)
    with tempfile.TemporaryDirectory(prefix="dashboard-data-") as tmpdir:
        archive_path = Path(tmpdir) / "dashboard-data.zip"
        download_file(DATA_ZIP_URL, archive_path)
        extract_zip_safe(archive_path, RUNTIME_ROOT)
    print(f"Runtime data sync complete at {RUNTIME_ROOT}", flush=True)


def runtime_data_ready() -> bool:
    return (
        DATA_ROOT.exists()
        and PARSED_ROOT.exists()
        and IMAGE_ROOT.exists()
        and any(PARSED_ROOT.glob("KW*_*/all_suppliers*.csv"))
    )


def download_file(url: str, target: Path) -> None:
    request = Request(url, headers={"User-Agent": "list-railway-dashboard/1.0"})
    with urlopen(request, timeout=120) as response, target.open("wb") as handle:
        shutil.copyfileobj(response, handle)


def extract_zip_safe(archive_path: Path, destination: Path) -> None:
    destination = destination.resolve()
    with zipfile.ZipFile(archive_path) as archive:
        for member in archive.infolist():
            if member.is_dir():
                continue
            member_path = Path(member.filename)
            if member_path.is_absolute() or ".." in member_path.parts:
                raise RuntimeError(f"Unsafe ZIP path: {member.filename}")
            target = (destination / member_path).resolve()
            try:
                target.relative_to(destination)
            except ValueError as exc:
                raise RuntimeError(f"Unsafe ZIP path: {member.filename}") from exc
            target.parent.mkdir(parents=True, exist_ok=True)
            with archive.open(member) as source, target.open("wb") as handle:
                shutil.copyfileobj(source, handle)


def redact_url(url: str) -> str:
    parsed = urlparse(url)
    if not parsed.query:
        return url
    return parsed._replace(query="...").geturl()


def truthy_env(name: str) -> bool:
    return str(os.environ.get(name) or "").strip().casefold() in {"1", "true", "yes", "ja"}


class DashboardHandler(SimpleHTTPRequestHandler):
    server_version = "ListPriceDashboard/1.0"

    def log_message(self, format: str, *args) -> None:  # noqa: A002
        return

    def do_HEAD(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/login":
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            return
        if not self.is_authenticated():
            self.send_response(HTTPStatus.SEE_OTHER)
            self.send_header("Location", "/login")
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            return
        self.send_response(HTTPStatus.OK)
        self.send_header("Cache-Control", "no-store")
        self.end_headers()

    def do_GET(self) -> None:
        try:
            parsed = urlparse(self.path)
            if parsed.path == "/login":
                if self.is_authenticated():
                    return self.redirect("/")
                return self.serve_login_page()
            if parsed.path == "/logout":
                return self.logout()
            if parsed.path == "/client-logo":
                return self.serve_logo()
            if not self.is_authenticated():
                return self.require_login(parsed.path)
            if parsed.path == "/":
                return self.serve_static("index.html")
            if parsed.path.startswith("/assets/"):
                return self.serve_static(parsed.path.removeprefix("/assets/"))
            if parsed.path == "/file":
                return self.serve_file(parse_qs(parsed.query))
            if parsed.path.startswith("/api/"):
                return self.route_api_get(parsed.path, parse_qs(parsed.query))
            self.send_error(HTTPStatus.NOT_FOUND)
        except Exception as exc:
            self.send_json({"error": str(exc)}, status=500)

    def do_POST(self) -> None:
        try:
            parsed = urlparse(self.path)
            if parsed.path == "/login":
                return self.handle_login()
            if not self.is_authenticated():
                return self.require_login(parsed.path)
            if parsed.path == "/api/runs":
                if REVIEW_ONLY:
                    return self.send_json({"error": "Pipeline-Starts sind in der Kundenansicht deaktiviert."}, status=HTTPStatus.FORBIDDEN)
                payload = self.read_json()
                run = start_pipeline_run(payload)
                return self.send_json(run)
            if parsed.path == "/api/qa-mark":
                payload = self.read_json()
                return self.send_json(save_qa_mark(payload))
            self.send_error(HTTPStatus.NOT_FOUND)
        except Exception as exc:
            self.send_json({"error": str(exc)}, status=500)

    def route_api_get(self, path: str, query: dict[str, list[str]]) -> None:
        if path == "/api/config":
            return self.send_json({
                "review_only": REVIEW_ONLY,
                "mode": "review" if REVIEW_ONLY else "local",
                "can_start_pipeline": not REVIEW_ONLY,
            })
        if path == "/api/weeks":
            return self.send_json({"weeks": list_weeks()})
        if path == "/api/runs":
            return self.send_json({"runs": list_runs()})
        match = re.fullmatch(r"/api/runs/([^/]+)", path)
        if match:
            return self.send_json(get_run(match.group(1)))
        if path == "/api/products":
            return self.send_json(get_products(query))
        if path == "/api/pdf-relevance":
            return self.send_json(get_pdf_relevance(query))
        if path == "/api/matching":
            return self.send_json(get_matching(query))
        if path == "/api/preview":
            return self.send_json(get_preview(query))
        if path == "/api/summary":
            return self.send_json(get_summary(query))
        self.send_error(HTTPStatus.NOT_FOUND)

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0"))
        if not length:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def is_authenticated(self) -> bool:
        cookies = self.headers.get("Cookie", "")
        for part in cookies.split(";"):
            name, _, value = part.strip().partition("=")
            if name == AUTH_COOKIE_NAME and hmac.compare_digest(value, AUTH_TOKEN):
                return True
        return False

    def require_login(self, path: str) -> None:
        if path.startswith("/api/"):
            return self.send_json({"error": "Nicht angemeldet."}, status=HTTPStatus.UNAUTHORIZED)
        return self.redirect("/login")

    def handle_login(self) -> None:
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length).decode("utf-8") if length else ""
        form = parse_qs(body)
        password = first(form, "password", "")
        if hmac.compare_digest(password, DASHBOARD_PASSWORD):
            self.send_response(HTTPStatus.SEE_OTHER)
            self.send_header("Location", "/")
            self.send_header("Set-Cookie", auth_cookie(AUTH_TOKEN))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            return
        self.serve_login_page(error=True)

    def logout(self) -> None:
        self.send_response(HTTPStatus.SEE_OTHER)
        self.send_header("Location", "/login")
        self.send_header("Set-Cookie", auth_cookie("", max_age=0))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()

    def redirect(self, location: str) -> None:
        self.send_response(HTTPStatus.SEE_OTHER)
        self.send_header("Location", location)
        self.send_header("Cache-Control", "no-store")
        self.end_headers()

    def serve_login_page(self, error: bool = False) -> None:
        error_html = '<div class="error">Falsches Passwort. Bitte erneut versuchen.</div>' if error else ""
        html = f"""<!doctype html>
<html lang="de">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>LIST Preisvergleich Login</title>
    <style>
      :root {{
        --green: #73bf00;
        --green-dark: #1d301b;
        --line: #d9e6cf;
        --muted: #65745f;
        --red: #d92d20;
      }}
      * {{ box-sizing: border-box; }}
      body {{
        align-items: center;
        background: radial-gradient(circle at 18% 10%, rgba(115, 191, 0, 0.18), transparent 30%), #eef6e8;
        color: var(--green-dark);
        display: flex;
        font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
        justify-content: center;
        margin: 0;
        min-height: 100vh;
        padding: 24px;
      }}
      .login-card {{
        background: rgba(255, 255, 255, 0.94);
        border: 1px solid var(--line);
        border-radius: 18px;
        box-shadow: 0 24px 70px rgba(29, 48, 27, 0.16);
        max-width: 430px;
        padding: 32px;
        width: 100%;
      }}
      .logo-box {{
        align-items: center;
        background: #fff;
        border: 1px solid var(--line);
        border-radius: 14px;
        display: flex;
        height: 118px;
        justify-content: center;
        margin-bottom: 24px;
      }}
      .logo-box img {{
        max-height: 84px;
        max-width: 84%;
      }}
      h1 {{
        font-size: 28px;
        line-height: 1.1;
        margin: 0 0 8px;
      }}
      p {{
        color: var(--muted);
        line-height: 1.45;
        margin: 0 0 22px;
      }}
      label {{
        display: block;
        font-size: 13px;
        font-weight: 800;
        margin-bottom: 8px;
      }}
      input {{
        border: 1px solid var(--line);
        border-radius: 10px;
        font: inherit;
        padding: 13px 14px;
        width: 100%;
      }}
      input:focus {{
        border-color: var(--green);
        box-shadow: 0 0 0 4px rgba(115, 191, 0, 0.18);
        outline: none;
      }}
      button {{
        background: var(--green);
        border: 0;
        border-radius: 10px;
        color: #fff;
        cursor: pointer;
        font: inherit;
        font-weight: 900;
        margin-top: 16px;
        padding: 13px 16px;
        width: 100%;
      }}
      .error {{
        background: #fff1f0;
        border: 1px solid #ffd1cc;
        border-radius: 10px;
        color: var(--red);
        font-size: 13px;
        font-weight: 800;
        margin-bottom: 14px;
        padding: 10px 12px;
      }}
    </style>
  </head>
  <body>
    <main class="login-card">
      <div class="logo-box"><img src="/client-logo" alt="LIST Logo" /></div>
      <h1>Preisvergleich</h1>
      <p>Bitte Passwort eingeben, um die Anwendung zu öffnen.</p>
      {error_html}
      <form method="post" action="/login">
        <label for="password">Passwort</label>
        <input id="password" name="password" type="password" autocomplete="current-password" autofocus required />
        <button type="submit">Anmelden</button>
      </form>
    </main>
  </body>
</html>"""
        data = html.encode("utf-8")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def serve_static(self, relative: str) -> None:
        path = safe_join(WEB_ROOT, relative)
        if not path or not path.exists() or not path.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        data = path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def serve_logo(self) -> None:
        logo_path = next((path for path in [CLIENT_LOGO, WEB_ROOT / "client-logo.svg", WEB_ROOT / "client-logo.png"] if path.exists()), None)
        if not logo_path:
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        data = logo_path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mimetypes.guess_type(logo_path.name)[0] or "image/svg+xml")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def serve_file(self, query: dict[str, list[str]]) -> None:
        raw_path = first(query, "path")
        path = resolve_dashboard_path(unquote(raw_path))
        if not path or not path.exists() or not path.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(path.stat().st_size))
        self.end_headers()
        with path.open("rb") as handle:
            while chunk := handle.read(1024 * 256):
                self.wfile.write(chunk)

    def send_json(self, payload: dict, status: int = 200) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def start_pipeline_run(payload: dict) -> dict:
    week = int(payload.get("week") or 0)
    year = int(payload.get("year") or 0)
    if week < 1 or week > 53 or year < 2020:
        raise RuntimeError("Ungültige Kalenderwoche oder ungültiges Jahr.")
    run_id = datetime.now().strftime("%Y%m%d%H%M%S") + "-" + uuid.uuid4().hex[:6]
    run = {
        "id": run_id,
        "week": week,
        "year": year,
        "label": f"KW{week:02d} {year}",
        "status": "running",
        "step_key": "queued",
        "step_index": 0,
        "progress": 4,
        "started_at": iso_now(),
        "finished_at": "",
        "message": "Lauf vorgemerkt.",
        "logs": [],
        "artifacts": artifact_paths(week, year),
    }
    with ACTIVE_LOCK:
        ACTIVE_RUNS[run_id] = run
    persist_run(run)
    thread = threading.Thread(target=run_pipeline_process, args=(run_id,), daemon=True)
    thread.start()
    return run


def auth_cookie(value: str, max_age: int | None = None) -> str:
    parts = [f"{AUTH_COOKIE_NAME}={value}", "Path=/", "HttpOnly", "SameSite=Lax"]
    if max_age is not None:
        parts.append(f"Max-Age={max_age}")
    if COOKIE_SECURE:
        parts.append("Secure")
    return "; ".join(parts)


def run_pipeline_process(run_id: str) -> None:
    run = get_active_run(run_id)
    command = [sys.executable, "main.py", "--week", str(run["week"]), "--year", str(run["year"])]
    append_log(run_id, f"Starte Befehl: {' '.join(command)}")
    try:
        process = subprocess.Popen(
            command,
            cwd=ROOT,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )
        assert process.stdout is not None
        for line in process.stdout:
            line = line.strip()
            if line:
                append_log(run_id, line)
                update_step_from_line(run_id, line)
        return_code = process.wait()
        if return_code == 0:
            update_run(run_id, status="completed", step_key="completed", step_index=len(PIPELINE_STEPS) - 1, progress=100, message="Pipeline abgeschlossen.")
        else:
            update_run(run_id, status="failed", progress=current_progress(run_id), message=f"Pipeline fehlgeschlagen mit Exit-Code {return_code}.")
    except Exception as exc:
        append_log(run_id, f"Dashboard-Fehler: {exc}")
        update_run(run_id, status="failed", message=str(exc), progress=current_progress(run_id))
    finally:
        run = get_active_run(run_id)
        run["finished_at"] = iso_now()
        run["artifacts"] = artifact_paths(run["week"], run["year"])
        persist_run(run)


def update_step_from_line(run_id: str, line: str) -> None:
    markers = [
        ("--- ACQUISITION ---", "acquisition"),
        ("--- RELEVANCE + EXTRACTION ---", "document_relevance"),
        ("document extraction", "extraction"),
        ("--- PRODUCT RELEVANCE ---", "product_relevance"),
        ("--- COMPETITOR PRODUCT MATCHING ---", "matching"),
        ("--- REPORT GENERATION ---", "report"),
        ("Pipeline complete", "completed"),
    ]
    for marker, step_key in markers:
        if marker.lower() in line.lower():
            set_step(run_id, step_key, line)
            return
    if "Extracted" in line and "products" in line:
        set_step(run_id, "extraction", line)
    elif "Saved product relevance outputs" in line:
        set_step(run_id, "product_relevance", line)
    elif "Saved competitor matching output" in line:
        set_step(run_id, "matching", line)


def set_step(run_id: str, step_key: str, message: str) -> None:
    step_index = next((i for i, step in enumerate(PIPELINE_STEPS) if step[0] == step_key), 0)
    progress = int(8 + (step_index / max(1, len(PIPELINE_STEPS) - 1)) * 90)
    update_run(run_id, step_key=step_key, step_index=step_index, progress=progress, message=human_log_message(message))


def human_log_message(line: str) -> str:
    replacements = {
        "--- ACQUISITION ---": "Prospekte von Metro, Edeka, Selgros und Handelshof werden gesammelt.",
        "--- RELEVANCE + EXTRACTION ---": "Prospektrelevanz wird geprüft und Produkte aus akzeptierten PDFs werden extrahiert.",
        "--- PRODUCT RELEVANCE ---": "Es wird klassifiziert, welche Produktzeilen für den Vergleich relevant sind.",
        "--- COMPETITOR PRODUCT MATCHING ---": "Vergleichbare Produkte werden über Wettbewerber hinweg zugeordnet.",
        "--- REPORT GENERATION ---": "Die finalen Excel-Ausgaben werden geschrieben.",
    }
    if line in replacements:
        return replacements[line]
    lowered = line.lower()
    if "saved product relevance outputs" in lowered:
        return "Produktrelevanz-Ausgaben wurden gespeichert."
    if "saved competitor matching output" in lowered:
        return "Wettbewerber-Zuordnung wurde gespeichert."
    if "extracted" in lowered and "products" in lowered:
        return "Produkte wurden aus den relevanten PDFs extrahiert."
    if "pipeline complete" in lowered:
        return "Pipeline abgeschlossen."
    return line


def append_log(run_id: str, line: str) -> None:
    run = get_active_run(run_id)
    run["logs"].append({"time": iso_now(), "message": line})
    run["logs"] = run["logs"][-600:]
    persist_run(run)


def update_run(run_id: str, **updates) -> None:
    run = get_active_run(run_id)
    run.update(updates)
    persist_run(run)


def get_active_run(run_id: str) -> dict:
    with ACTIVE_LOCK:
        if run_id in ACTIVE_RUNS:
            return ACTIVE_RUNS[run_id]
    for run in load_run_store():
        if run["id"] == run_id:
            return run
    raise RuntimeError(f"Lauf nicht gefunden: {run_id}")


def current_progress(run_id: str) -> int:
    return int(get_active_run(run_id).get("progress") or 0)


def list_runs() -> list[dict]:
    stored = {run["id"]: run for run in load_run_store()}
    with ACTIVE_LOCK:
        stored.update(ACTIVE_RUNS)
    runs = list(stored.values())
    known = {(run["week"], run["year"]) for run in runs}
    for week_info in list_weeks():
        key = (week_info["week"], week_info["year"])
        if key not in known and week_info["has_relevant"]:
            runs.append({
                "id": f"historic-{week_info['year']}-{week_info['week']:02d}",
                "week": week_info["week"],
                "year": week_info["year"],
                "label": week_info["label"],
                "status": "completed",
                "progress": 100,
                "step_key": "completed",
                "step_index": len(PIPELINE_STEPS) - 1,
                "started_at": "",
                "finished_at": "",
                "message": "Historische Dateien auf der Festplatte gefunden.",
                "logs": [],
                "artifacts": artifact_paths(week_info["week"], week_info["year"]),
            })
    return sorted(runs, key=lambda item: (item.get("year", 0), item.get("week", 0), item.get("started_at", "")), reverse=True)


def get_run(run_id: str) -> dict:
    with ACTIVE_LOCK:
        if run_id in ACTIVE_RUNS:
            return ACTIVE_RUNS[run_id]
    for run in list_runs():
        if run["id"] == run_id:
            return run
    return {"error": "Lauf nicht gefunden"}


def load_run_store() -> list[dict]:
    if not RUN_STORE.exists():
        return []
    try:
        return json.loads(RUN_STORE.read_text(encoding="utf-8"))
    except Exception:
        return []


def persist_run(run: dict) -> None:
    RUN_STORE.parent.mkdir(parents=True, exist_ok=True)
    runs = {item["id"]: item for item in load_run_store()}
    runs[run["id"]] = run
    RUN_STORE.write_text(json.dumps(list(runs.values()), ensure_ascii=False, indent=2), encoding="utf-8")


def list_weeks() -> list[dict]:
    weeks = []
    for path in sorted(PARSED_ROOT.glob("KW*_*/"), reverse=True):
        match = re.match(r"KW(\d{1,2})_(\d{4})", path.name)
        if not match:
            continue
        week, year = int(match.group(1)), int(match.group(2))
        weeks.append({
            "week": week,
            "year": year,
            "label": f"KW{week:02d} {year}",
            "path": rel(path),
            "has_raw": (path / "all_suppliers.csv").exists(),
            "has_relevant": (path / "all_suppliers_relevant.csv").exists(),
            "has_matching": bool(find_matching_workbook(week, year)),
        })
    return weeks


def get_products(query: dict[str, list[str]]) -> dict:
    week, year = week_year_from_query(query)
    view = first(query, "view", "relevance")
    path = week_dir(week, year) / ("all_suppliers_relevant.csv" if view == "relevance" else "all_suppliers.csv")
    rows = read_csv_rows(path)
    for idx, row in enumerate(rows):
        row["_row_id"] = idx + 1
        row["_preview"] = preview_payload(row)
    apply_qa_marks(
        rows,
        week,
        year,
        "relevance" if view == "relevance" else "extraction",
        ["supplier", "source_file", "source_page", "product_name", "description", "price", "quantity"],
    )
    return {"week": week, "year": year, "view": view, "rows": rows, "stats": product_stats(rows)}


def get_pdf_relevance(query: dict[str, list[str]]) -> dict:
    week, year = week_year_from_query(query)
    rows = []
    for supplier_dir in sorted(DATA_ROOT.glob(f"*/{year}/{week:02d}")) + sorted(DATA_ROOT.glob(f"*/{year}/{week}")):
        decision_path = supplier_dir / "relevance_decisions.json"
        if not decision_path.exists():
            continue
        supplier = supplier_dir.parts[-3]
        try:
            decisions = json.loads(decision_path.read_text(encoding="utf-8"))
        except Exception:
            decisions = []
        for index, item in enumerate(decisions):
            row = dict(item)
            row["_row_id"] = len(rows) + 1
            row["supplier"] = supplier
            row["filename"] = row.get("filename") or Path(str(row.get("file_path") or "")).name
            row["skip_label"] = row.get("relevance_label", "") if not bool(row.get("is_relevant")) else ""
            row["skip_reason"] = row.get("relevance_reason", "") if not bool(row.get("is_relevant")) else ""
            row["_preview"] = pdf_relevance_preview_payload(row)
            rows.append(row)
    apply_qa_marks(rows, week, year, "pdfs", ["supplier", "file_path", "filename", "title", "tab"])
    return {
        "week": week,
        "year": year,
        "rows": rows,
        "stats": {
            "total": len(rows),
            "relevant": sum(1 for row in rows if bool(row.get("is_relevant"))),
            "skipped": sum(1 for row in rows if not bool(row.get("is_relevant"))),
            "skip_reasons": skip_reason_breakdown(rows),
            "suppliers": sorted({row.get("supplier", "") for row in rows if row.get("supplier")}),
        },
    }


def skip_reason_breakdown(rows: list[dict]) -> list[dict]:
    grouped: dict[str, dict] = {}
    for row in rows:
        if bool(row.get("is_relevant")):
            continue
        label = str(row.get("skip_label") or row.get("relevance_label") or "skipped").strip()
        reason = str(row.get("skip_reason") or row.get("relevance_reason") or "").strip()
        entry = grouped.setdefault(label, {"label": label, "count": 0, "example": reason})
        entry["count"] += 1
        if reason and not entry.get("example"):
            entry["example"] = reason
    return sorted(grouped.values(), key=lambda item: (-int(item["count"]), str(item["label"])))


def apply_qa_marks(rows: list[dict], week: int, year: int, scope: str, identity_fields: list[str]) -> None:
    stored_marks = load_qa_marks(week, year).get(scope, {})
    for idx, row in enumerate(rows):
        row.setdefault("_row_id", idx + 1)
        key = qa_row_key(scope, row, identity_fields, idx)
        mark = stored_marks.get(key, {})
        row["_qa_scope"] = scope
        row["_qa_key"] = key
        row["_qa_mark"] = {
            "is_mistake": bool(mark.get("is_mistake")),
            "comment": str(mark.get("comment") or ""),
            "updated_at": str(mark.get("updated_at") or ""),
        }


def qa_row_key(scope: str, row: dict, identity_fields: list[str], idx: int) -> str:
    values = {field: str(row.get(field) or "").strip() for field in identity_fields}
    values["_row_id"] = str(row.get("_row_id") or idx + 1)
    raw = json.dumps({"scope": scope, "values": values}, ensure_ascii=False, sort_keys=True)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:20]


def save_qa_mark(payload: dict) -> dict:
    week = int(payload.get("week") or 0)
    year = int(payload.get("year") or 0)
    if week < 1 or week > 53 or year < 2020:
        raise RuntimeError("Ungültige Kalenderwoche oder ungültiges Jahr.")
    scope = re.sub(r"[^a-zA-Z0-9_-]", "", str(payload.get("scope") or ""))
    key = re.sub(r"[^a-fA-F0-9]", "", str(payload.get("key") or ""))
    if not scope or not key:
        raise RuntimeError("Fehlender Prüfbereich oder Zeilenschlüssel.")
    comment = str(payload.get("comment") or "").strip()[:2000]
    is_mistake = bool(payload.get("is_mistake"))
    marks = load_qa_marks(week, year)
    scope_marks = marks.setdefault(scope, {})
    if is_mistake or comment:
        scope_marks[key] = {
            "is_mistake": is_mistake,
            "comment": comment,
            "updated_at": iso_now(),
        }
    else:
        scope_marks.pop(key, None)
    save_qa_marks(week, year, marks)
    return {
        "week": week,
        "year": year,
        "scope": scope,
        "key": key,
        "mark": scope_marks.get(key, dict(QA_MARK_DEFAULT)),
    }


def qa_mark_path(week: int, year: int) -> Path:
    return DATA_ROOT / f"qa_marks_KW{week:02d}_{year}.json"


def load_qa_marks(week: int, year: int) -> dict:
    path = qa_mark_path(week, year)
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def save_qa_marks(week: int, year: int, marks: dict) -> None:
    path = qa_mark_path(week, year)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(marks, ensure_ascii=False, indent=2), encoding="utf-8")


def pdf_relevance_preview_payload(row: dict[str, object]) -> dict:
    file_path = safe_absolute_or_relative_path(str(row.get("file_path") or ""))
    image_path = find_relevance_preview(file_path)
    if image_path:
        return {"type": "image", "url": f"/file?path={rel(image_path)}", "path": rel(image_path)}
    if file_path and file_path.exists():
        return {"type": "pdf", "url": f"/file?path={rel(file_path)}", "path": rel(file_path)}
    return {"type": "none", "url": "", "path": ""}


def find_relevance_preview(file_path: Path | None) -> Path | None:
    if not file_path:
        return None
    try:
        supplier = file_path.parts[-4]
    except Exception:
        return None
    base = IMAGE_ROOT / "_relevance" / supplier / file_path.stem
    for candidate in [base / "preview-01.png", base / "preview-1.png"]:
        if candidate.exists():
            return candidate
    matches = sorted(base.glob("preview-*.png"))
    return matches[0] if matches else None


def get_matching(query: dict[str, list[str]]) -> dict:
    week, year = week_year_from_query(query)
    workbook = find_matching_workbook(week, year)
    if not workbook:
        return {"week": week, "year": year, "rows": [], "review": [], "available": False}
    sheet = first(query, "sheet", "Final Output")
    if not excel_has_sheet(workbook, sheet):
        sheet = default_matching_sheet(workbook, week)
    rows = read_excel_sheet(workbook, sheet, header_row=7)
    review = read_excel_sheet(workbook, "review_queue", header_row=1) if excel_has_sheet(workbook, "review_queue") else []
    pair_debug = read_excel_sheet(workbook, "pair_debug", header_row=1) if excel_has_sheet(workbook, "pair_debug") else []
    matching_scope = "matching_short" if sheet == "Final Output Short" else "matching_final"
    apply_qa_marks(rows, week, year, matching_scope, ["Kategorie", "Produkt", "Marke", "Metro", "Selgros", "Handelshof", "Edeka"])
    apply_qa_marks(review, week, year, "matching_review", ["review_id", "canonical_product_id", "canonical_product_name", "product_ids", "issue_summary"])
    return {
        "week": week,
        "year": year,
        "workbook": rel(workbook),
        "sheet": sheet,
        "rows": rows,
        "review": review,
        "pair_debug": pair_debug[:1000],
        "available": True,
        "stats": matching_stats(rows, review),
    }


def get_summary(query: dict[str, list[str]]) -> dict:
    week, year = week_year_from_query(query)
    raw_rows = read_csv_rows(week_dir(week, year) / "all_suppliers.csv")
    relevant_rows = read_csv_rows(week_dir(week, year) / "all_suppliers_relevant.csv")
    matching = get_matching({"week": [str(week)], "year": [str(year)]})
    return {
        "week": week,
        "year": year,
        "raw_count": len(raw_rows),
        "relevant_count": sum(1 for row in relevant_rows if yes(row.get("Relevant")) and yes(row.get("Relevant Time"))),
        "total_relevance_rows": len(relevant_rows),
        "matching_rows": len(matching.get("rows") or []),
        "review_count": len(matching.get("review") or []),
        "artifacts": artifact_paths(week, year),
    }


def get_preview(query: dict[str, list[str]]) -> dict:
    source = first(query, "source_file")
    page = first(query, "page")
    payload = preview_payload({"source_file": source, "source_page": page})
    return payload


def preview_payload(row: dict[str, str]) -> dict:
    source = row.get("source_file") or ""
    page = str(row.get("source_page") or "").strip()
    source_path = resolve_dashboard_path(source)
    image_path = find_page_image(source_path, page) if source_path else None
    if image_path:
        return {"type": "image", "url": f"/file?path={rel(image_path)}", "path": rel(image_path)}
    if source_path and source_path.exists():
        suffix = f"#page={page}" if page else ""
        return {"type": "pdf", "url": f"/file?path={rel(source_path)}{suffix}", "path": rel(source_path)}
    return {"type": "none", "url": "", "path": ""}


def safe_absolute_or_relative_path(value: str) -> Path | None:
    if not value:
        return None
    return resolve_dashboard_path(value)


def find_page_image(source_path: Path | None, page: str) -> Path | None:
    if not source_path:
        return None
    try:
        supplier, year, week = source_path.parts[-4], source_path.parts[-3], int(source_path.parts[-2])
    except Exception:
        return None
    stem = source_path.stem
    page_int = int(float(page)) if page else 1
    base = IMAGE_ROOT / supplier / year / f"KW{week:02d}" / stem
    for name in [f"page-{page_int:02d}.png", f"page-{page_int}.png"]:
        candidate = base / name
        if candidate.exists():
            return candidate
    return None


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def read_excel_sheet(path: Path, sheet: str, header_row: int) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [str(cell.value or "").strip() for cell in ws[header_row]]
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        rows.append({headers[idx] or f"column_{idx + 1}": "" if value is None else str(value) for idx, value in enumerate(row)})
    return rows


def excel_has_sheet(path: Path, sheet: str) -> bool:
    wb = load_workbook(path, read_only=True, data_only=True)
    return sheet in wb.sheetnames


def default_matching_sheet(path: Path, week: int) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    for candidate in [f"KW{week:02d}", "Final Output Short", "Final Output"]:
        if candidate in wb.sheetnames:
            return candidate
    return wb.sheetnames[0] if wb.sheetnames else ""


def product_stats(rows: list[dict[str, str]]) -> dict:
    suppliers = sorted({row.get("supplier", "") for row in rows if row.get("supplier")})
    relevant = sum(1 for row in rows if yes(row.get("Relevant")))
    time_relevant = sum(1 for row in rows if yes(row.get("Relevant Time")))
    return {"total": len(rows), "suppliers": suppliers, "relevant": relevant, "time_relevant": time_relevant}


def matching_stats(rows: list[dict[str, str]], review: list[dict[str, str]]) -> dict:
    competitor_cols = ["Metro", "Selgros", "Handelshof", "Edeka"]
    matched_multi = 0
    unmatched = 0
    for row in rows:
        present = sum(1 for col in competitor_cols if row.get(col))
        matched_multi += present >= 2
        unmatched += present <= 1
    return {"total": len(rows), "matched_multi": matched_multi, "single_supplier": unmatched, "review": len(review)}


def artifact_paths(week: int, year: int) -> dict:
    base = week_dir(week, year)
    workbook = find_matching_workbook(week, year)
    return {
        "raw_csv": rel(base / "all_suppliers.csv") if (base / "all_suppliers.csv").exists() else "",
        "relevant_csv": rel(base / "all_suppliers_relevant.csv") if (base / "all_suppliers_relevant.csv").exists() else "",
        "relevant_xlsx": rel(base / "all_suppliers_relevant.xlsx") if (base / "all_suppliers_relevant.xlsx").exists() else "",
        "matching_xlsx": rel(workbook) if workbook else "",
    }


def find_matching_workbook(week: int, year: int) -> Path | None:
    base = week_dir(week, year)
    candidates = [
        base / f"Artikelvergleich KW{week:02d}.xlsx",
        base / "matched_competitor_products.xlsx",
    ]
    return next((path for path in candidates if path.exists()), None)


def week_dir(week: int, year: int) -> Path:
    return PARSED_ROOT / f"KW{week:02d}_{year}"


def week_year_from_query(query: dict[str, list[str]]) -> tuple[int, int]:
    week = int(first(query, "week", "0"))
    year = int(first(query, "year", "0"))
    if week < 1 or year < 2020:
        weeks = list_weeks()
        if not weeks:
            raise RuntimeError("Keine Wochendaten gefunden.")
        return weeks[0]["week"], weeks[0]["year"]
    return week, year


def yes(value: object) -> bool:
    return str(value or "").strip().casefold() in {"ja", "yes", "true", "1", "x", "relevant"}


def first(query: dict[str, list[str]], key: str, default: str = "") -> str:
    values = query.get(key)
    return values[0] if values else default


def safe_join(root: Path, relative: str) -> Path | None:
    if not relative:
        return None
    path = (root / relative).resolve()
    try:
        path.relative_to(root.resolve())
    except ValueError:
        return None
    return path


def resolve_dashboard_path(value: str) -> Path | None:
    if not value:
        return None
    raw = unquote(str(value).split("#", 1)[0])
    path = Path(raw)
    if path.is_absolute():
        resolved = path.resolve()
        if is_allowed_file_root(resolved):
            return resolved
        remapped = remap_known_root_from_absolute(path)
        return remapped if remapped and is_allowed_file_root(remapped) else None
    normalized = raw.lstrip("/")
    candidates = []
    if normalized == "data" or normalized.startswith("data/"):
        relative = normalized.split("/", 1)[1] if "/" in normalized else ""
        candidates.append((DATA_ROOT / relative).resolve())
    elif normalized == "parsed" or normalized.startswith("parsed/"):
        relative = normalized.split("/", 1)[1] if "/" in normalized else ""
        candidates.append((PARSED_ROOT / relative).resolve())
    elif normalized == "images" or normalized.startswith("images/"):
        relative = normalized.split("/", 1)[1] if "/" in normalized else ""
        candidates.append((IMAGE_ROOT / relative).resolve())
    candidates.append((ROOT / normalized).resolve())
    for candidate in candidates:
        if is_allowed_file_root(candidate):
            return candidate
    return None


def is_allowed_file_root(path: Path) -> bool:
    for root in [ROOT, DATA_ROOT, PARSED_ROOT, IMAGE_ROOT]:
        try:
            path.relative_to(root.resolve())
            return True
        except ValueError:
            continue
    return False


def remap_known_root_from_absolute(path: Path) -> Path | None:
    parts = path.parts
    for folder, root in [("data", DATA_ROOT), ("parsed", PARSED_ROOT), ("images", IMAGE_ROOT)]:
        if folder not in parts:
            continue
        index = len(parts) - 1 - parts[::-1].index(folder)
        return (root / Path(*parts[index + 1:])).resolve()
    return None


def rel(path: Path) -> str:
    resolved = path.resolve()
    for prefix, root in [("data", DATA_ROOT), ("parsed", PARSED_ROOT), ("images", IMAGE_ROOT), ("", ROOT)]:
        try:
            suffix = resolved.relative_to(root.resolve())
            return str(Path(prefix) / suffix) if prefix else str(suffix)
        except ValueError:
            continue
    return str(resolved)


def iso_now() -> str:
    return datetime.now().isoformat(timespec="seconds")


if __name__ == "__main__":
    main()
