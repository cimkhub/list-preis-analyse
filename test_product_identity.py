import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

from add_market_forecast_signals import finalize_customer_workbook

from match_competitor_products import (
    Caches,
    FINAL_OUTPUT_COLUMNS,
    PAIR_CACHE_VERSION,
    apply_second_judgement,
    build_clusters,
    build_output_rows,
    generate_same_supplier_duplicate_candidates,
    judge_pairs,
    make_pair_key,
    write_excel,
)
from src.harmonize.product_identity import (
    build_identity_clusters,
    build_offer_records,
    dedupe_observation_rows,
    deterministic_canonical_name,
    make_observation_id,
    make_offer_id,
    normalize_origin,
    structural_compatibility,
)


def row(key: str, name: str, **overrides):
    item = {
        "supplier": "selgros",
        "supplier_norm": "Selgros",
        "location": "Braunschweig",
        "product_name": name,
        "product_name_original": name,
        "product_name_de": name,
        "product": name,
        "description": "",
        "category": "obst_gemuese",
        "product_family": "obst_gemuese",
        "temperature_state": "unknown",
        "processing_state": "raw_plain",
        "calibre": "",
        "origin": "",
        "brand": "",
        "source_brand": "",
        "brand_evidence": "",
        "brand_evidence_source": "unknown",
        "relevance_brand": "",
        "relevance_policy_group": "",
        "relevance_route": "core_fresh",
        "quantity": "1",
        "unit": "kg",
        "price": "5.99",
        "price_per_kg": "",
        "price_is_net": "False",
        "price_gross": "",
        "price_tiers": "",
        "price_basis": "unit",
        "package_count": "",
        "package_size_value": "",
        "package_size_unit": "",
        "total_content_value": "",
        "total_content_unit": "",
        "packaging_type": "",
        "packaging_raw": "",
        "valid_from": "2026-08-10",
        "valid_to": "2026-08-15",
        "calendar_week": "32",
        "year": "2026",
        "source_document_sha256": "sha256:selgros-kw32",
        "source_file": "selgros_kw32.pdf",
        "source_page": "1",
        "source_item_id": key,
        "source_item_index": key,
        "source_row_index": key,
    }
    item.update(overrides)
    return item


def prepare(*items):
    prepared, audit = dedupe_observation_rows(items)
    assert not audit
    return prepared


def attrs_for(items):
    return {
        item["product_id"]: {
            "product_id": item["product_id"],
            "base_product": item["product_name_original"],
            "product_family": item.get("product_family", "unknown"),
            "fresh_or_frozen": item.get("temperature_state", "unknown"),
            "processing": item.get("processing_state", "unknown"),
            "origin": item.get("origin") or None,
            "calibre": item.get("calibre") or None,
            "attribute_confidence": 100,
        }
        for item in items
    }


def cluster_members(clusters):
    return sorted(tuple(sorted(cluster["product_ids"])) for cluster in clusters)


def test_observation_id_is_order_independent_and_ignores_derived_translation():
    original = row("one", "Pork Back", product_name_de="Schweinerücken", product="Schweinerücken")
    changed_translation = {
        **original,
        "product_name_de": "Schweineruecken",
        "product": "Schweineruecken",
        "source_row_index": "999",
        "source_item_index": "999",
        "source_file": "renamed.pdf",
    }

    assert make_observation_id(original) == make_observation_id(changed_translation)
    assert make_observation_id(original) != make_observation_id({**original, "price": "6.49"})

    first, _ = dedupe_observation_rows([original, row("two", "Forelle")])
    second, _ = dedupe_observation_rows(list(reversed([original, row("two", "Forelle")])))
    assert [item["product_id"] for item in first] == [item["product_id"] for item in second]


def test_exact_observation_dedupe_has_audit_link_but_keeps_commercial_variants():
    exact = row("chanterelle-a", "Pfifferlinge", price="28.99", quantity="3", unit="Kiste")
    other_offer = row("chanterelle-b", "Pfifferlinge", price="11.99", quantity="1", unit="Korb")

    kept, audit = dedupe_observation_rows([exact, dict(exact), other_offer])

    assert len(kept) == 2
    assert len(audit) == 1
    assert audit[0]["removed_observation_id"] == audit[0]["kept_observation_id"]
    assert audit[0]["reason"] == "exact_observation_duplicate"


def test_offer_fingerprint_excludes_source_and_aggregates_source_refs():
    first, second = prepare(
        row("offer-a", "Pfifferlinge", source_file="one.pdf", source_page="4"),
        row("offer-b", "Pfifferlinge", source_file="two.pdf", source_page="9"),
    )
    records = build_offer_records("pv_pfifferlinge", [second, first])

    assert len(records) == 1
    assert len(records[0].observation_ids) == 2
    assert {ref["source_page"] for ref in records[0].source_refs} == {"4", "9"}
    assert make_offer_id("pv_pfifferlinge", first) == make_offer_id("pv_pfifferlinge", second)
    assert make_offer_id("pv_pfifferlinge", first) != make_offer_id(
        "pv_pfifferlinge", {**first, "unit": "Kiste"}
    )


def test_origin_temperature_line_brand_and_calibre_variant_rules():
    assert normalize_origin("UA, RO").values == normalize_origin("Rumänien / Ukraine").values
    assert structural_compatibility(row("a", "Pfifferlinge", origin="UA, RO"), row("b", "Pfifferlinge", origin="RO/UA")).compatible
    assert not structural_compatibility(row("a", "Pfifferlinge", origin="UA"), row("b", "Pfifferlinge", origin="UA, RO")).compatible
    assert not structural_compatibility(row("a", "Tomaten", origin="EU"), row("b", "Tomaten", origin="Deutschland")).compatible
    assert structural_compatibility(row("a", "Tomaten", origin=""), row("b", "Tomaten", origin="Deutschland")).compatible

    assert not structural_compatibility(
        row("a", "Rinder-Roastbeef", temperature_state="fresh"),
        row("b", "Rinder-Roastbeef", temperature_state="frozen"),
    ).compatible
    assert not structural_compatibility(row("a", "Duroc Schweinenacken"), row("b", "Schweinenacken")).compatible
    assert structural_compatibility(
        row("a", "Forelle", calibre="200-400 g"),
        row("b", "Forelle", calibre="300-400 g"),
    ).compatible

    packaged = {"relevance_route": "packaged_exception", "relevance_policy_group": "milk"}
    assert not structural_compatibility(
        row("a", "Milch", relevance_brand="ARO", **packaged),
        row("b", "Milch", relevance_brand="Milram", **packaged),
    ).compatible
    assert structural_compatibility(
        row("a", "Milch", relevance_brand="", **packaged),
        row("b", "Milch", relevance_brand="Milram", **packaged),
    ).compatible
    assert structural_compatibility(row("a", "ASC Wildlachsfilet"), row("b", "MSC Wildlachsfilet")).compatible


def test_kw32_grouping_cases_preserve_family_variant_and_offer_levels():
    items = prepare(
        row("pf-a", "Pfifferlinge", origin="UA, RO", price="28.99", quantity="3", unit="Kiste"),
        row("pf-b", "Pfifferlinge", origin="RO, UA", price="11.99", quantity="1", unit="Korb"),
        row("gz-a", "Gemüsezwiebeln", origin="Spanien", calibre="80-110", valid_from="2026-08-03", valid_to="2026-08-08"),
        row("gz-b", "Gemüsezwiebeln", origin="Spanien", calibre="80-110", valid_from="2026-08-10", valid_to="2026-08-15"),
        row("ka-a", "Karotte", origin="Deutschland", valid_from="2026-08-03", valid_to="2026-08-08"),
        row("ka-b", "Möhre", origin="Deutschland", valid_from="2026-08-10", valid_to="2026-08-15"),
        row("rb-f", "Rinder-Roastbeef", category="fleisch", product_family="fleisch", origin="Uruguay", temperature_state="fresh"),
        row("rb-tk", "Rinder-Roastbeef", category="tk", product_family="fleisch", origin="Argentinien", temperature_state="frozen"),
        row("sn-g", "Schweinenacken", category="fleisch", product_family="fleisch", origin="Spanien"),
        row("sn-d", "Duroc Schweinenacken", category="fleisch", product_family="fleisch", origin="Spanien"),
        row("fo-a", "Forelle", category="fisch", product_family="fisch", calibre="200-400 g"),
        row("fo-b", "Forelle", category="fisch", product_family="fisch", calibre="300-400 g"),
    )
    clusters = build_identity_clusters(items, attrs_for(items))
    by_names = [
        ({next(item["source_item_id"] for item in items if item["product_id"] == pid) for pid in cluster["product_ids"]}, cluster)
        for cluster in clusters
    ]

    def matching(expected):
        return next(cluster for names, cluster in by_names if names == set(expected))

    for expected in [("pf-a", "pf-b"), ("gz-a", "gz-b"), ("ka-a", "ka-b"), ("fo-a", "fo-b")]:
        cluster = matching(expected)
        product_rows = [item for item in items if item["product_id"] in cluster["product_ids"]]
        assert len(build_offer_records(cluster["variant_id"], product_rows)) == 2

    roastbeef = [cluster for names, cluster in by_names if names <= {"rb-f", "rb-tk"}]
    assert len(roastbeef) == 2
    assert len({cluster["product_family_id"] for cluster in roastbeef}) == 1
    nacken = [cluster for names, cluster in by_names if names <= {"sn-g", "sn-d"}]
    assert len(nacken) == 2
    assert len({cluster["product_family_id"] for cluster in nacken}) == 1


def test_complete_cluster_compatibility_blocks_missing_attribute_bridges():
    items = prepare(
        row("a", "Pfifferlinge", origin="Spanien"),
        row("b", "Pfifferlinge", origin=""),
        row("c", "Pfifferlinge", origin="Deutschland"),
    )
    attributes = attrs_for(items)
    first = build_identity_clusters(items, attributes)
    second = build_identity_clusters(list(reversed(items)), attributes)

    assert cluster_members(first) == cluster_members(second)
    id_by_source = {item["source_item_id"]: item["product_id"] for item in items}
    assert not any({id_by_source["a"], id_by_source["c"]} <= set(cluster["product_ids"]) for cluster in first)


def test_hard_structural_conflict_wins_before_cache_and_after_second_merge(tmp_path):
    items = prepare(
        row("fresh", "Rinder-Roastbeef", temperature_state="fresh"),
        row("frozen", "Rinder-Roastbeef", temperature_state="frozen"),
    )
    df = pd.DataFrame(items)
    attributes = attrs_for(items)
    a, b = items[0]["product_id"], items[1]["product_id"]
    candidate = {
        "product_id_a": a,
        "product_id_b": b,
        "supplier_a": "Selgros",
        "supplier_b": "Selgros",
        "name_a": "Rinder-Roastbeef",
        "name_b": "Rinder-Roastbeef",
        "similarity": 1.0,
    }
    caches = Caches(tmp_path)
    caches.pairs[make_pair_key(a, b)] = {
        "pair_key": make_pair_key(a, b),
        "schema_version": PAIR_CACHE_VERSION,
        "decision": "exact_match",
        "confidence": 100,
        "should_human_review": False,
    }

    judged, count = judge_pairs([candidate], df, attributes, caches, None, False, True, 1)
    assert count == 1
    assert judged[0]["hard_blocked"] is True
    assert judged[0]["decision"] == "same_family_not_comparable"

    after_second = apply_second_judgement(judged[0], {
        "final_action": "MERGE",
        "confidence": 100,
        "manual_review_needed": False,
    })
    assert after_second == judged[0]


def test_same_supplier_same_page_is_a_candidate_without_deleting_offers():
    items = prepare(
        row("one", "Pfifferlinge", source_page="7", price="28.99", quantity="3", unit="Kiste"),
        row("two", "Pfifferlinge", source_page="7", price="11.99", quantity="1", unit="Korb"),
    )
    attributes = attrs_for(items)
    embeddings = {item["product_id"]: np.array([1.0, 0.0]) for item in items}

    candidates = generate_same_supplier_duplicate_candidates(
        pd.DataFrame(items), attributes, embeddings, top_k=25
    )

    assert len(candidates) == 1
    assert candidates[0]["supplier_a"] == candidates[0]["supplier_b"] == "Selgros"


def test_source_backed_names_block_red_snapper_and_lachsforelle_degradation():
    red_snapper = row(
        "snapper",
        "Red Snapper Filet",
        product_name_de="Rotbarsch Filet",
        product="Rotbarsch Filet",
    )
    trout = row(
        "trout",
        "Lachsforelle",
        product_name_de="Forelle Rot",
        product="Forelle Rot",
    )

    assert deterministic_canonical_name([red_snapper]) == "Red Snapper Filet"
    assert deterministic_canonical_name([trout]) == "Lachsforelle"
    assert not structural_compatibility(red_snapper, row("rotbarsch", "Rotbarsch Filet")).compatible


def test_same_supplier_offers_render_as_two_stable_blocks_and_excel_contract(tmp_path):
    items = prepare(
        row("pf-a", "Pfifferlinge", origin="UA, RO", price="28.99", quantity="3", unit="Kiste"),
        row("pf-b", "Pfifferlinge", origin="RO, UA", price="11.99", quantity="1", unit="Korb"),
    )
    attributes = attrs_for(items)

    def render(product_rows):
        df = pd.DataFrame(product_rows)
        clusters = build_clusters(df, attributes, [])
        return build_output_rows(df, attributes, [], clusters)[0]

    matched = render(items)
    reversed_matched = render(list(reversed(items)))
    assert len(matched) == 1
    assert matched[0]["Selgros"].split("\n\n") == reversed_matched[0]["Selgros"].split("\n\n")
    assert len(matched[0]["Selgros"].split("\n\n")) == 2
    assert len(matched[0]["Selgros_short"].split("\n\n")) == 2
    assert "28,99 EUR" in matched[0]["Selgros"]
    assert "11,99 EUR" in matched[0]["Selgros"]
    assert matched[0]["review_needed"] == "FALSE"

    path = tmp_path / "Artikelvergleich KW32.xlsx"
    audit = [{
        "removed_observation_id": "obs_same",
        "kept_observation_id": "obs_same",
        "reason": "exact_observation_duplicate",
        "removed_source_row_index": "2",
        "kept_source_row_index": "1",
    }]
    write_excel(path, matched, [], [], [], None, observation_audit=audit)
    workbook = load_workbook(path, read_only=True)

    assert [cell.value for cell in workbook["Final Output"][7]] == FINAL_OUTPUT_COLUMNS
    assert [cell.value for cell in workbook["Final Output Short"][7]] == FINAL_OUTPUT_COLUMNS
    assert "product_family_id" in [cell.value for cell in workbook["matched_products"][1]]
    assert [cell.value for cell in workbook["observation_dedupe"][1]][:3] == [
        "removed_observation_id", "kept_observation_id", "reason"
    ]
    assert json.loads(matched[0]["offer_source_refs"])


def test_customer_finalization_keeps_audit_very_hidden_with_one_visible_sheet():
    workbook = Workbook()
    workbook.active.title = "Final Output Short"
    workbook.create_sheet("Final Output")
    for name in [
        "matched_products",
        "review_queue",
        "pair_debug",
        "attribute_debug",
        "observation_dedupe",
    ]:
        workbook.create_sheet(name)
    workbook.create_sheet("temporary")

    final_name = finalize_customer_workbook(
        workbook,
        Path("Artikelvergleich KW32.xlsx"),
    )

    assert final_name == "KW32"
    assert [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"] == ["KW32"]
    assert "Final Output" not in workbook.sheetnames
    assert "temporary" not in workbook.sheetnames
    for name in [
        "matched_products",
        "review_queue",
        "pair_debug",
        "attribute_debug",
        "observation_dedupe",
    ]:
        assert workbook[name].sheet_state == "veryHidden"
